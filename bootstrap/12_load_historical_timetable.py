"""
Bootstrap 12 — Load AY2526 historical timetable from SIT Excel exports as Draft entries.

Sources:
  2510 DSC Year 1-2 Timetable.xlsx  -> Tri 1  (Sep–Nov 2025)
  2520 DSC Year 1-3 Timetable 1.xlsx -> Tri 2  (Jan–Apr 2026)
  2530 DSC Year 1-2 Timetable 1.xlsx -> Tri 3  (May–Jul 2026)

Entries are imported as DRAFT (is_published=False).
Use Admin -> Timetable -> Publish to make them visible to professors/students.

Run:
    python bootstrap/12_load_historical_timetable.py
"""

import sys, os, re
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import pandas as pd
import openpyxl
from datetime import date, time, timedelta, datetime

from app import create_app, db
from app.models.course import Course
from app.models.class_session import ClassSession
from app.models.timeslot import TimeSlot
from app.models.room import Room
from app.models.professor import Professor
from app.models.timetable_entry import TimetableEntry
from app.models.academic_calendar import AcademicCalendar
from app.models.class_session_professor import ClassSessionProfessor


def link_professor(session, prof):
    """Create the ClassSessionProfessor link if it doesn't already exist.
    Without this, find_professor()'s match was only ever used for the
    unmatched-name report - never actually saved (a real bug, found and
    fixed 2026-07-10)."""
    if not prof:
        return
    existing = ClassSessionProfessor.query.filter_by(
        session_id=session.id, professor_id=prof.id).first()
    if existing:
        return
    is_first = ClassSessionProfessor.query.filter_by(session_id=session.id).count() == 0
    db.session.add(ClassSessionProfessor(
        session_id=session.id, professor_id=prof.id,
        is_primary=is_first, display_order=0 if is_first else 1,
    ))

# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
BASE = os.environ.get('HISTORICAL_TIMETABLE_DIR', '').strip()
FILES = [
    (os.path.join(BASE, '2510 DSC Year 1-2 Timetable.xlsx'),    'Sheet', 1, 'AY2526-T1', 'AY2526'),
    (os.path.join(BASE, '2520 DSC Year 1-3 Timetable 1.xlsx'),  'raw',   2, 'AY2526-T2', 'AY2526'),
    (os.path.join(BASE, '2530 DSC Year 1-2 Timetable 1.xlsx'),  None,    3, 'AY2526-T3', 'AY2526'),  # grid only
]
if not BASE or not os.path.isdir(BASE):
    raise SystemExit(
        'Set HISTORICAL_TIMETABLE_DIR to the folder containing the three source workbooks.'
    )

# ---------------------------------------------------------------------------
# Session type mapping
# ---------------------------------------------------------------------------
TYPE_MAP = {
    'lecture':    'lecture',
    'lectorial':  'lecture',
    'tutorial':   'tutorial',
    'laboratory': 'lab',
    'seminar':    'seminar',
    'workshop':   'workshop',
    'quiz':       'quiz',
    'projects':   None,   # skip - no session_type equivalent, not scheduled by the solver
}

def map_session_type(raw):
    return TYPE_MAP.get(raw.strip().lower())

# ---------------------------------------------------------------------------
# Date parsing  "03-Sep-25" or "03-Sep-2025"
# ---------------------------------------------------------------------------
def parse_dates(date_str):
    if not date_str or str(date_str).strip().upper() in ('TBA', 'NAN', ''):
        return []
    results = []
    for part in str(date_str).split(','):
        part = part.strip()
        if not part:
            continue
        for fmt in ('%d-%b-%y', '%d-%b-%Y'):
            try:
                results.append(datetime.strptime(part, fmt).date())
                break
            except ValueError:
                pass
    return results

# ---------------------------------------------------------------------------
# Room matching: strip -SR### suffix, try E2-XX-XX code
# ---------------------------------------------------------------------------
_room_cache = None

def _load_rooms():
    global _room_cache
    if _room_cache is None:
        _room_cache = {r.room_code.upper(): r for r in Room.query.all()}
    return _room_cache

def find_room(room_str):
    if not room_str or room_str.strip().lower() in ('online', 'nan', ''):
        return None  # online — no room needed
    rooms = _load_rooms()
    # Take first room if comma-separated
    code = room_str.split(',')[0].strip()
    # Try exact match
    if code.upper() in rooms:
        return rooms[code.upper()]
    # Strip -SR### suffix: "E2-07-01-SR259" -> "E2-07-01"
    stripped = re.sub(r'-SR\d+$', '', code).upper()
    if stripped in rooms:
        return rooms[stripped]
    # Try matching on just the SR part
    sr_match = re.search(r'SR(\d+)', code, re.IGNORECASE)
    if sr_match:
        sr_code = f'SR{sr_match.group(1)}'
        for k, v in rooms.items():
            if k.endswith(sr_code):
                return v
    return None   # unmatched

# ---------------------------------------------------------------------------
# Timeslot matching: prefer duration that matches the session type
# ---------------------------------------------------------------------------
_ts_cache = None
_ts_by_day_start = None   # (day, start_time) -> [TimeSlot, ...]

# Expected duration (minutes) per session type.
# Extend this dict if new session types are added.
SESSION_EXPECTED_DURATION = {
    'lab':      180,
    'lecture':  120,
    'tutorial': 120,
    'seminar':  120,
}

def _load_timeslots():
    global _ts_cache, _ts_by_day_start
    if _ts_cache is None:
        _ts_cache = {}
        _ts_by_day_start = {}
        for ts in TimeSlot.query.all():
            key = (ts.day_of_week.lower(), ts.start_time, ts.end_time)
            _ts_cache[key] = ts
            ds_key = (ts.day_of_week.lower(), ts.start_time)
            _ts_by_day_start.setdefault(ds_key, []).append(ts)
    return _ts_cache, _ts_by_day_start

def find_timeslot(day_str, start_t, end_t, session_type=None):
    """Find the best matching timeslot for a session.

    Prefers the timeslot whose duration matches the expected duration for the
    given session_type (e.g. labs → 180 min, lectures → 120 min).  Falls back
    to exact (day, start, end) if no duration-preferred candidate exists.
    """
    if not day_str or not start_t:
        return None
    slots, by_day_start = _load_timeslots()
    day = day_str.strip().lower()
    preferred_dur = SESSION_EXPECTED_DURATION.get(session_type) if session_type else None

    # Candidates sharing the same day + start time
    candidates = by_day_start.get((day, start_t), [])

    if candidates and preferred_dur:
        def _dur(ts):
            return (ts.end_time.hour * 60 + ts.end_time.minute) - \
                   (ts.start_time.hour * 60 + ts.start_time.minute)
        # Sort: exact duration match first, then closest
        candidates = sorted(candidates, key=lambda ts: abs(_dur(ts) - preferred_dur))
        return candidates[0]

    # Exact match by (day, start, end) — no session_type hint given
    if end_t:
        exact = slots.get((day, start_t, end_t))
        if exact:
            return exact

    # Fallback: any candidate with the right start time
    return candidates[0] if candidates else None

def parse_time(t_val):
    """Convert '09:00:00' string or time object -> time."""
    if isinstance(t_val, time):
        return t_val
    if isinstance(t_val, str):
        try:
            parts = t_val.strip().split(':')
            return time(int(parts[0]), int(parts[1]))
        except Exception:
            pass
    return None

# ---------------------------------------------------------------------------
# Professor matching: word-inclusion (case-insensitive)
# ---------------------------------------------------------------------------
_prof_cache = None

def _load_professors():
    global _prof_cache
    if _prof_cache is None:
        _prof_cache = []
        for p in Professor.query.join(Professor.user).all():
            if p.user:
                _prof_cache.append((p, p.user.name.upper()))
    return _prof_cache

def find_professor(name_str):
    if not name_str or str(name_str).strip().upper() in ('NAN', ''):
        return None
    # Take first if comma-separated (primary professor)
    name = name_str.split(',')[0].strip().upper()
    words = set(name.split())
    profs = _load_professors()
    # Whole-word match, not substring - "WANG YU" must not match "WANG
    # FENGYU" just because "YU" is a substring of "FENGYU" (real bug found
    # 2026-07-10, caused a false professor double-booking).
    exact, partial = None, None
    for prof, db_name in profs:
        db_words = set(db_name.split())
        if words == db_words:
            exact = prof
            break
        if partial is None and words <= db_words:
            partial = prof
    return exact or partial

# ---------------------------------------------------------------------------
# ClassSession matching: by module_code + session_type + trimester_num
# Cycle through available sessions for parallel groups (LEC, LEC2, LEC3...)
# ---------------------------------------------------------------------------
_session_assignment = {}   # (module_code, session_type, tri_num, group_variant) -> session_id

def find_session(module_code, session_type, trimester_num, group_variant):
    """
    Map a historical group variant (LEC/LEC2/LEC3/P1/P2...) to a ClassSession.
    Cycles through available sessions so parallel groups get different sessions.
    """
    key = (module_code.upper(), session_type, trimester_num, group_variant)
    if key in _session_assignment:
        sid = _session_assignment[key]
        return ClassSession.query.get(sid)

    # A module code can have several Course rows (one per trimester it's
    # offered in) - pick the one that actually has a matching session for
    # THIS trimester/type, not just the first Course row found overall.
    candidate_courses = Course.query.filter(
        Course.module_code.ilike(module_code)
    ).all()
    if not candidate_courses:
        return None

    sessions = []
    for course in candidate_courses:
        sessions = ClassSession.query.filter_by(
            course_id=course.id,
            session_type=session_type,
            trimester=trimester_num,
        ).all()
        if sessions:
            break
    if not sessions:
        return None

    # Count how many group_variants have already been assigned for this module+type
    already = sum(
        1 for k in _session_assignment
        if k[0] == module_code.upper() and k[1] == session_type and k[2] == trimester_num
    )
    # Cycle through available sessions
    chosen = sessions[already % len(sessions)]
    _session_assignment[key] = chosen.id
    return chosen

# ---------------------------------------------------------------------------
# Week number calculation relative to trimester start
# ---------------------------------------------------------------------------
def week_number_for_date(d, tri_start_monday):
    delta = (d - tri_start_monday).days
    if delta < 0:
        return None
    return delta // 7 + 1

def monday_of(d):
    return d - timedelta(days=d.weekday())

# ---------------------------------------------------------------------------
# Academic calendar seeding
# ---------------------------------------------------------------------------
def ensure_academic_calendar(trimester_key, tri_start_date, all_dates):
    existing = AcademicCalendar.query.filter_by(trimester=trimester_key).first()
    if existing:
        return
    # Determine week range from actual data dates
    if not all_dates:
        return
    start_monday = monday_of(min(all_dates))
    end_date = max(all_dates)
    total_weeks = (monday_of(end_date) - start_monday).days // 7 + 1

    import holidays as holidays_lib
    years = set(range(start_monday.year, end_date.year + 1))
    sg_ph = set(holidays_lib.Singapore(years=years).keys())

    current = start_monday
    for wk in range(1, total_weeks + 1):
        week_days = [current + timedelta(days=d) for d in range(5)]
        ph_days   = [d for d in week_days if d in sg_ph]
        is_ph     = len(ph_days) > 0
        ph_names  = ', '.join(
            holidays_lib.Singapore(years={d.year}).get(d, '') for d in ph_days
        )
        db.session.add(AcademicCalendar(
            trimester        = trimester_key,
            week_number      = wk,
            start_date       = current,
            end_date         = current + timedelta(days=4),
            is_term_break    = False,
            is_public_holiday= is_ph,
            notes            = f'Public holiday: {ph_names}' if is_ph else None,
        ))
        current += timedelta(weeks=1)
    db.session.commit()
    print(f'  [calendar] Created {total_weeks} weeks for {trimester_key}')

# ---------------------------------------------------------------------------
# Parse structured sheet (2510 "Sheet" / 2520 "raw")
# ---------------------------------------------------------------------------
def parse_structured_sheet(filepath, sheet_name, tri_num, tri_key, ay):
    print(f'\n--- Parsing {os.path.basename(filepath)} [{sheet_name}] -> {tri_key} ---')
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=0)

    entries_created = 0
    skipped_no_ts   = []
    skipped_no_sess = []
    unmatched_rooms = set()
    unmatched_profs = set()
    used_entries    = set()   # (session_id, week_num)

    all_dates_seen  = []

    rows_by_group = {}   # group_variant -> list of row dicts
    for _, row in df.iterrows():
        name = str(row.get('Name', '')).strip()
        if not name or name == 'nan':
            continue

        # Module code: "DSC1001-2510-ENG-UGRD-PU-LEC/All" -> "DSC1001"
        module_code = name.split('-')[0].strip()

        # Group variant: the real parallel-section id (P1/P2/S5/S6/Q1/All)
        # is the token AFTER the last slash, not before it. "SEM/S5" and
        # "SEM2/S5" are the same section S5 continuing into a later
        # calendar block; "SEM/S5" and "SEM/S6" are different parallel
        # sections meeting at the same block. Grouping on the pre-slash
        # tag (the old behaviour) merged different sections - e.g. all of
        # S5/S6/S7/S8 for UCS1001 landed in one "SEM" bucket sharing one
        # ClassSession, wrongly cross-linking every section's real
        # professor onto every other section (real bug, found 2026-07-10).
        variant_match = re.search(r'[/-](\w+)$', name)
        group_variant = variant_match.group(1).upper() if variant_match else 'ALL'

        activity_raw  = str(row.get('Activity Type Name', '')).strip()
        session_type  = map_session_type(activity_raw)
        if not session_type:
            continue

        dates_raw  = row.get('Activity Dates (Individual)', '')
        day_str    = str(row.get('Scheduled Days', '')).strip()
        start_raw  = row.get('Scheduled Start Time', '')
        end_raw    = row.get('Scheduled End Time', '')
        room_raw   = str(row.get('Allocated Location Name', '')).strip()
        staff_raw  = str(row.get('Allocated Staff Name', '')).strip()

        dates   = parse_dates(dates_raw)
        start_t = parse_time(start_raw)
        end_t   = parse_time(end_raw)
        if not dates or not start_t or not end_t:
            continue

        all_dates_seen.extend(dates)

        rows_by_group.setdefault((module_code, session_type, group_variant), []).append({
            'dates': dates, 'day': day_str,
            'start_t': start_t, 'end_t': end_t,
            'room_raw': room_raw, 'staff_raw': staff_raw,
        })

    # Seed academic calendar from all dates seen
    if all_dates_seen:
        tri_start = monday_of(min(all_dates_seen))
        ensure_academic_calendar(tri_key, tri_start, all_dates_seen)
    else:
        tri_start = date.today()

    # Now create entries
    for (module_code, session_type, group_variant), row_list in rows_by_group.items():
        session = find_session(module_code, session_type, tri_num, group_variant)
        if not session:
            skipped_no_sess.append(f'{module_code} {session_type} (variant={group_variant})')
            continue

        for row in row_list:
            ts = find_timeslot(row['day'], row['start_t'], row['end_t'], session_type=session_type)
            if not ts:
                skipped_no_ts.append(
                    f"{module_code} {session_type} — "
                    f"{row['day']} {row['start_t']}-{row['end_t']}"
                )
                continue

            room = find_room(row['room_raw'])
            if row['room_raw'].lower() not in ('online', 'nan', '') and not room:
                unmatched_rooms.add(row['room_raw'])

            prof = find_professor(row['staff_raw'])
            if row['staff_raw'] not in ('', 'nan') and not prof:
                unmatched_profs.add(row['staff_raw'].split(',')[0].strip())
            link_professor(session, prof)

            for d in row['dates']:
                wk = week_number_for_date(d, tri_start)
                if not wk:
                    continue
                pair = (session.id, wk)
                if pair in used_entries:
                    continue
                used_entries.add(pair)
                db.session.add(TimetableEntry(
                    class_session_id  = session.id,
                    timeslot_id       = ts.id,
                    room_id           = room.id if room else None,
                    week_number       = wk,
                    trimester         = tri_key,
                    academic_year     = ay,
                    is_published      = True,
                    is_manually_edited= False,
                    is_backbone       = True,
                ))
                entries_created += 1

    db.session.commit()
    _print_report(tri_key, entries_created, skipped_no_ts, skipped_no_sess,
                  unmatched_rooms, unmatched_profs)

# ---------------------------------------------------------------------------
# Parse grid sheet (2530 — no structured tab)
# ---------------------------------------------------------------------------
def parse_grid_sheet(filepath, sheet_name, tri_num, tri_key, ay):
    print(f'\n--- Parsing {os.path.basename(filepath)} [{sheet_name}] grid -> {tri_key} ---')
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    # Build col -> date from row-2 merged cells (e.g. "Mon 04-May-2026")
    col_to_date = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row != 2:
            continue
        val = ws.cell(mr.min_row, mr.min_col).value
        if not isinstance(val, str):
            continue
        # "Mon 04-May-2026"  or  "Fri 24-Jul-2026"
        m = re.search(r'(\d{2}-[A-Za-z]{3}-\d{4})', val)
        if not m:
            continue
        try:
            d = datetime.strptime(m.group(1), '%d-%b-%Y').date()
            for c in range(mr.min_col, mr.max_col + 1):
                col_to_date[c] = d
        except ValueError:
            pass

    # Build row -> start_time from col-A "08:00-08:30" labels
    row_to_start = {}
    for r in range(3, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, str) and ':' in val:
            try:
                start_str = val.split('-')[0].strip()
                h, m2 = map(int, start_str.split(':'))
                row_to_start[r] = time(h, m2)
            except Exception:
                pass

    # Process merged session cells (skip row ≤ 2 and col ≤ 1)
    entries_created = 0
    skipped_no_ts   = []
    skipped_no_sess = []
    unmatched_rooms = set()
    unmatched_profs = set()
    used_entries    = set()
    all_dates_seen  = []
    all_rows = []

    for mr in ws.merged_cells.ranges:
        if mr.min_row <= 2 or mr.min_col <= 1:
            continue
        val = ws.cell(mr.min_row, mr.min_col).value
        if not isinstance(val, str) or '|' not in val:
            continue

        # "DSC2101 All | \nE6-03-16-SR343 | \nKHO CHOON SIONG"
        parts = [p.strip() for p in re.split(r'\|\s*\n?', val)]
        if len(parts) < 2:
            continue

        module_group = parts[0].strip()
        room_raw     = parts[1].strip() if len(parts) > 1 else ''
        staff_raw    = parts[2].strip() if len(parts) > 2 else ''

        module_code  = module_group.split()[0].strip()
        group        = module_group.split()[1].strip() if len(module_group.split()) > 1 else 'All'

        session_date = col_to_date.get(mr.min_col)
        if not session_date:
            continue

        all_dates_seen.append(session_date)

        start_t = row_to_start.get(mr.min_row)
        # End time = start of the row AFTER the merged range
        end_row_start = row_to_start.get(mr.max_row + 1)
        if not start_t:
            continue
        if end_row_start:
            end_t = end_row_start
        else:
            # Calculate: (max_row - min_row + 1) * 30 min
            total_min = (mr.max_row - mr.min_row + 1) * 30
            eh = start_t.hour + (start_t.minute + total_min) // 60
            em = (start_t.minute + total_min) % 60
            end_t = time(eh % 24, em)

        day_name = session_date.strftime('%A')
        all_rows.append((module_code, group, session_date, day_name, start_t, end_t, room_raw, staff_raw))

    if not all_dates_seen:
        print('  No session data found.')
        return

    tri_start = monday_of(min(all_dates_seen))
    ensure_academic_calendar(tri_key, tri_start, all_dates_seen)

    # Infer session_type from duration
    def infer_type(start_t, end_t, module_code, group):
        mins = (end_t.hour * 60 + end_t.minute) - (start_t.hour * 60 + start_t.minute)
        if mins <= 0:
            mins += 24 * 60
        if group.upper() in ('ALL',):
            return 'lecture'  # All-group sessions are usually lectures
        if mins >= 180:
            return 'lab'
        if mins >= 120:
            return 'tutorial'
        return 'tutorial'

    for (module_code, group, session_date, day_name, start_t, end_t, room_raw, staff_raw) in all_rows:
        session_type = infer_type(start_t, end_t, module_code, group)
        session = find_session(module_code, session_type, tri_num, group)
        if not session:
            # Try other types
            for alt in ('lecture', 'tutorial', 'lab', 'seminar'):
                if alt != session_type:
                    session = find_session(module_code, alt, tri_num, group)
                    if session:
                        break
        if not session:
            skipped_no_sess.append(f'{module_code} {session_type} (group={group})')
            continue

        ts = find_timeslot(day_name, start_t, end_t, session_type=session_type)
        if not ts:
            skipped_no_ts.append(f'{module_code} — {day_name} {start_t}-{end_t}')
            continue

        room = find_room(room_raw)
        if room_raw.lower() not in ('online', 'nan', '') and not room:
            unmatched_rooms.add(room_raw)

        prof = find_professor(staff_raw)
        if staff_raw not in ('', 'nan') and not prof:
            unmatched_profs.add(staff_raw.split(',')[0].strip())
        link_professor(session, prof)

        wk = week_number_for_date(session_date, tri_start)
        if not wk:
            continue
        pair = (session.id, wk)
        if pair in used_entries:
            continue
        used_entries.add(pair)
        db.session.add(TimetableEntry(
            class_session_id  = session.id,
            timeslot_id       = ts.id,
            room_id           = room.id if room else None,
            week_number       = wk,
            trimester         = tri_key,
            academic_year     = ay,
            is_published      = False,
            is_manually_edited= False,
            is_backbone       = True,
        ))
        entries_created += 1

    db.session.commit()
    _print_report(tri_key, entries_created, skipped_no_ts, skipped_no_sess,
                  unmatched_rooms, unmatched_profs)

# ---------------------------------------------------------------------------
# Report helper
# ---------------------------------------------------------------------------
def _print_report(tri_key, created, skipped_ts, skipped_sess, unmatched_rooms, unmatched_profs):
    print(f'  Entries created : {created}')
    if skipped_ts:
        print(f'  Skipped (no timeslot match, {len(skipped_ts)}):')
        for s in sorted(set(skipped_ts))[:10]:
            print(f'    - {s}')
        if len(skipped_ts) > 10:
            print(f'    ...and {len(skipped_ts)-10} more')
    if skipped_sess:
        print(f'  Skipped (no ClassSession in DB, {len(set(skipped_sess))}):')
        for s in sorted(set(skipped_sess))[:10]:
            print(f'    - {s}')
    if unmatched_rooms:
        print(f'  Unmatched rooms (need to add to DB or verify):')
        for r in sorted(unmatched_rooms):
            print(f'    - {r}')
    if unmatched_profs:
        print(f'  Unmatched professors (imported entry without professor link):')
        for p in sorted(unmatched_profs):
            print(f'    - {p}')

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
app = create_app()
with app.app_context():
    # Clear existing AY2526 historical entries
    existing = TimetableEntry.query.filter_by(academic_year='AY2526').count()
    if existing > 0:
        print(f'Clearing {existing} existing AY2526 entries...')
        TimetableEntry.query.filter_by(academic_year='AY2526').delete()
        AcademicCalendar.query.filter(
            AcademicCalendar.trimester.in_(['AY2526-T1', 'AY2526-T2', 'AY2526-T3'])
        ).delete()
        db.session.commit()

    for filepath, sheet_name, tri_num, tri_key, ay in FILES:
        if not os.path.exists(filepath):
            print(f'MISSING: {filepath}')
            continue

        if sheet_name:
            # Structured sheet (2510, 2520)
            parse_structured_sheet(filepath, sheet_name, tri_num, tri_key, ay)
        else:
            # Grid format (2530) — parse each year sheet
            wb = openpyxl.load_workbook(filepath)
            for sheet in wb.sheetnames:
                if sheet.lower().startswith('year'):
                    parse_grid_sheet(filepath, sheet, tri_num, tri_key, ay)

    print('\n=== Done ===')
    total = TimetableEntry.query.filter_by(academic_year='AY2526').count()
    print(f'Total AY2526 historical entries in DB: {total}')
    print()
    print('Next steps:')
    print('  1. Review in Admin > Timetable — select AY2526-T1 / T2 / T3 tabs')
    print('  2. Verify unmatched rooms/profs above and update data if needed')
    print('  3. Click Publish when ready — professors and students will then see it')
