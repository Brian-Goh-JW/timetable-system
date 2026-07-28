import io
import os
import unittest
import zipfile
from datetime import date, time, timedelta
from types import SimpleNamespace

os.environ['DATABASE_URL'] = 'sqlite:///:memory:'
os.environ['FLASK_SECRET_KEY'] = 'unit-test-secret'

import openpyxl
from flask import g

from app import create_app, db
from app.engine.solver import (
    _get_or_create_calendar,
    _overlapping_timeslot_indices,
    _parallel_section_alternatives,
    _room_compatible,
    _sessions_share_students,
    solve,
)
from app.models.academic_calendar import AcademicCalendar
from app.models.class_session import ClassSession
from app.models.course import Course
from app.models.event import Event
from app.models.programme import Programme
from app.models.professor import Professor
from app.models.room import Room
from app.models.class_session_professor import ClassSessionProfessor
from app.models.solve_run import SolveRun
from app.models.student_group import StudentGroup
from app.models.timeslot import TimeSlot
from app.models.timetable_entry import TimetableEntry
from app.models.user import User
from app.models.student_enrollment import StudentSectionAssignment
from app.models.student_enrollment import StudentEnrollment
from app.routes.admin import (
    _audit_generated_hard_conflicts,
    _check_week_conflicts,
    _load_solve_run,
)
from app.services.events import event_affects_occurrence
from app.utils.timetable import (
    apply_explicit_student_sections, select_preferred_layer,
    select_student_sections,
)
from app.utils.xlsx import restore_worksheet_extensions


class RegressionTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        cls.context = cls.app.app_context()
        cls.context.push()
        db.create_all()

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        db.drop_all()
        cls.context.pop()

    def setUp(self):
        db.session.rollback()
        for table in reversed(db.metadata.sorted_tables):
            db.session.execute(table.delete())
        db.session.commit()
        db.session.remove()
        g.pop('_login_user', None)

    def _base_data(self):
        programme = Programme(code='TST', name='Test', cluster='Test')
        group = StudentGroup(
            programme=programme, year_level=1,
            group_label='TST-Y1', intake_size=30,
        )
        course = Course(
            programme=programme, module_code='TST1001', trimester=1,
            title='Test', year_level=1, delivery_mode='f2f',
            sessions_per_week=1, total_hours=24, split_count=1,
        )
        db.session.add_all([programme, group, course])
        db.session.flush()
        return programme, group, course

    def _admin_client(self):
        admin = User(
            name='Admin', email='admin@sit.edu.sg', role='admin',
            password_hash='unused',
        )
        db.session.add(admin)
        db.session.commit()
        client = self.app.test_client()
        with client.session_transaction() as session:
            session['_user_id'] = str(admin.id)
            session['_fresh'] = True
        return client, admin

    def test_availability_expands_to_real_clock_overlaps(self):
        p2 = SimpleNamespace(
            day_of_week='Monday', start_time=time(12), end_time=time(14)
        )
        lab_pm2 = SimpleNamespace(
            day_of_week='Monday', start_time=time(13), end_time=time(15)
        )
        tuesday = SimpleNamespace(
            day_of_week='Tuesday', start_time=time(13), end_time=time(15)
        )

        self.assertEqual(
            {0, 1}, _overlapping_timeslot_indices([p2, lab_pm2, tuesday], p2)
        )

    def test_calendar_rebuilds_stale_twelve_week_target(self):
        start = date(2026, 5, 4)
        for number in range(1, 13):
            week_start = start + timedelta(weeks=number - 1)
            db.session.add(AcademicCalendar(
                trimester='AY2526-T3', week_number=number,
                start_date=week_start, end_date=week_start + timedelta(days=4),
                is_term_break=number == 7, is_public_holiday=False,
            ))
        db.session.commit()

        calendar = _get_or_create_calendar('AY2526-T3', start, {7})

        self.assertEqual(list(range(1, 14)), [week.week_number for week in calendar])

    def test_parallel_sections_do_not_claim_the_same_students(self):
        _, group, course = self._base_data()
        first = ClassSession(
            course=course, session_type='tutorial', delivery_mode='f2f',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='1,2,3', group_label='T1',
        )
        second = ClassSession(
            course=course, session_type='tutorial', delivery_mode='f2f',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='1,2,3', group_label='T2',
        )
        sequential = ClassSession(
            course=course, session_type='tutorial', delivery_mode='f2f',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='4,5,6', group_label='T3',
        )
        db.session.add_all([first, second, sequential])
        db.session.commit()

        self.assertTrue(_parallel_section_alternatives(first, second))
        self.assertFalse(_sessions_share_students(first, second))
        self.assertFalse(_parallel_section_alternatives(first, sequential))
        self.assertTrue(_sessions_share_students(first, sequential))
        self.assertEqual(15, first.effective_group_size)
        self.assertEqual(30, sequential.effective_group_size)

    def test_disjoint_week_sessions_can_reuse_one_slot_and_room(self):
        _, group, course = self._base_data()
        slot = TimeSlot(
            day_of_week='Monday', period_label='P1',
            start_time=time(9), end_time=time(11),
        )
        room = Room(
            room_code='TST-LT1', building='TST',
            room_type='lecture', capacity=60, is_active=True,
        )
        professor_user = User(
            name='Test Professor', email='professor@sit.edu.sg',
            role='professor', password_hash='unused',
        )
        professor = Professor(
            user=professor_user, staff_id='TST-P1', department='Test',
        )
        first = ClassSession(
            course=course, session_type='tutorial', delivery_mode='f2f',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='1', group_label='All',
            fixed_timeslot=slot, fixed_room=room,
        )
        second = ClassSession(
            course=course, session_type='tutorial', delivery_mode='f2f',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='3', group_label='All',
            fixed_timeslot=slot, fixed_room=room,
        )
        db.session.add_all([
            slot, room, professor_user, professor, first, second,
        ])
        db.session.flush()
        db.session.add_all([
            ClassSessionProfessor(
                session_id=first.id, professor_id=professor.id,
                is_primary=True,
            ),
            ClassSessionProfessor(
                session_id=second.id, professor_id=professor.id,
                is_primary=True,
            ),
        ])
        db.session.commit()

        success, _, _ = solve(
            'AY2728-T1', date(2027, 9, 6), {7},
            trimester_num=1, academic_year='AY2728',
            session_id_filter={first.id, second.id},
            max_time_seconds=5,
        )

        self.assertTrue(success)
        generated = TimetableEntry.query.filter_by(
            trimester='AY2728-T1', is_backbone=False,
        ).all()
        self.assertEqual(
            {first.id, second.id},
            {entry.class_session_id for entry in generated},
        )
        self.assertEqual({slot.id}, {entry.timeslot_id for entry in generated})
        self.assertEqual({room.id}, {entry.room_id for entry in generated})

    def test_student_projection_selects_one_parallel_section(self):
        _, group, course = self._base_data()
        slot = TimeSlot(
            day_of_week='Monday', period_label='P1',
            start_time=time(9), end_time=time(11),
        )
        first = ClassSession(
            course=course, session_type='tutorial', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='1,2,3', group_label='T1',
        )
        second = ClassSession(
            course=course, session_type='tutorial', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='1,2,3', group_label='T2',
        )
        sequential = ClassSession(
            course=course, session_type='tutorial', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='4,5,6', group_label='T3',
        )
        db.session.add_all([slot, first, second, sequential])
        db.session.flush()
        entries = [
            TimetableEntry(
                class_session_id=session.id, timeslot_id=slot.id,
                week_number=1 if session is not sequential else 4,
                trimester='AY2526-T1', is_backbone=False,
            )
            for session in (first, second, sequential)
        ]
        db.session.add_all(entries)
        db.session.commit()

        selected = select_student_sections(entries, selector_key=1)
        self.assertEqual(
            {second.id, sequential.id},
            {entry.class_session_id for entry in selected},
        )

    def test_manual_conflict_detects_overlapping_period_labels(self):
        _, group, course = self._base_data()
        p2 = TimeSlot(
            day_of_week='Monday', period_label='P2',
            start_time=time(12), end_time=time(14),
        )
        lab_pm2 = TimeSlot(
            day_of_week='Monday', period_label='Lab PM2',
            start_time=time(13), end_time=time(15),
        )
        first = ClassSession(
            course=course, session_type='lecture', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
        )
        second = ClassSession(
            course=course, session_type='lecture', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
        )
        db.session.add_all([p2, lab_pm2, first, second])
        db.session.flush()
        db.session.add(TimetableEntry(
            class_session_id=first.id, timeslot_id=lab_pm2.id,
            week_number=1, trimester='AY2526-T1', is_backbone=False,
        ))
        db.session.commit()

        conflicts = _check_week_conflicts(
            1, 'AY2526-T1', p2.id, None, None, group.id,
        )

        self.assertTrue(any('Student group double-booking' in item for item in conflicts))

    def test_publish_rejects_invalid_generated_room(self):
        client, _ = self._admin_client()
        _, group, course = self._base_data()
        slot = TimeSlot(
            day_of_week='Monday', period_label='P1',
            start_time=time(9), end_time=time(11),
        )
        class_session = ClassSession(
            course=course, session_type='lecture', delivery_mode='f2f',
            duration_hours=2, student_group=group, trimester=1,
        )
        db.session.add_all([slot, class_session])
        db.session.flush()
        entry = TimetableEntry(
            class_session_id=class_session.id, timeslot_id=slot.id,
            room_id=None, week_number=1, trimester='AY2526-T1',
            is_backbone=False, is_published=False,
        )
        db.session.add(entry)
        db.session.commit()

        response = client.post('/admin/timetable', data={
            'action': 'publish', 'trimester': 'AY2526-T1',
        })
        db.session.refresh(entry)

        self.assertEqual(200, response.status_code)
        self.assertFalse(entry.is_published)

    def test_professor_import_rejects_email_owned_by_admin(self):
        client, admin = self._admin_client()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Professors'
        sheet.append([
            'Staff ID', 'Name', 'Email', 'Department',
            'Temporary Password (New Staff Only)',
        ])
        sheet.append(['P999', 'Duplicate', admin.email, 'Test', 'Test1234'])
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = client.post(
            '/admin/professors/import',
            data={'file': (stream, 'professors.xlsx')},
            content_type='multipart/form-data',
        )

        self.assertEqual(302, response.status_code)
        self.assertEqual(0, db.session.query(User).filter_by(role='professor').count())

    def test_student_export_never_includes_password_hashes(self):
        client, _ = self._admin_client()
        _, group, _ = self._base_data()
        student = User(
            name='Student One', email='student.one@sit.edu.sg', role='student',
            student_group=group,
        )
        student.set_password('LocalTest123!')
        db.session.add(student)
        db.session.commit()

        response = client.get('/admin/students/export')
        workbook = openpyxl.load_workbook(io.BytesIO(response.data), data_only=False)
        sheet = workbook['Students']

        self.assertEqual(200, response.status_code)
        self.assertIn('students.xlsx', response.headers['Content-Disposition'])
        self.assertEqual(
            [
                'Account ID', 'Name', 'Email', 'Student Group Label',
                'Temporary Password (new or reset only)',
            ],
            [cell.value for cell in sheet[1]],
        )
        self.assertEqual(student.id, sheet['A2'].value)
        self.assertEqual('TST-Y1', sheet['D2'].value)
        self.assertIsNone(sheet['E2'].value)
        exported_values = {
            str(cell.value)
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        self.assertNotIn(student.password_hash, exported_values)

    def test_student_enrolment_export_is_valid_workbook(self):
        client, _ = self._admin_client()
        _, group, course = self._base_data()
        student = User(
            name='Enrolled Student', email='enrolled@example.com', role='student',
            password_hash='unused', student_group=group,
        )
        session = ClassSession(
            course=course, session_type='tutorial', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='1,2,3', group_label='T1',
        )
        db.session.add_all([student, session])
        db.session.flush()
        db.session.add(StudentEnrollment(
            user_id=student.id, course_id=course.id,
            academic_year='AY2526', trimester=1,
        ))
        db.session.add(StudentSectionAssignment(
            user_id=student.id, class_session_id=session.id,
            academic_year='AY2526', trimester=1,
        ))
        db.session.commit()

        response = client.get('/admin/students/enrolments/export')
        workbook = openpyxl.load_workbook(io.BytesIO(response.data))
        sheet = workbook['Enrolments']

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            ['Student Email', 'Academic Year', 'Trimester', 'Module Code', 'Section Group'],
            [cell.value for cell in sheet[1]],
        )
        self.assertEqual(
            ['enrolled@example.com', 'AY2526', 1, 'TST1001', 'T1'],
            [cell.value for cell in sheet[2]],
        )

    def test_student_group_export_roundtrips_as_preview(self):
        client, _ = self._admin_client()
        programme, parent, _ = self._base_data()
        db.session.add(StudentGroup(
            programme=programme, year_level=1,
            group_label='TST-Y1-Lab', intake_size=15, parent=parent,
        ))
        db.session.commit()

        exported = client.get('/admin/student-groups/export')
        response = client.post(
            '/admin/student-groups/import',
            data={
                'file': (
                    io.BytesIO(exported.data),
                    'student_groups.xlsx',
                ),
                'import_mode': 'preview',
            },
            content_type='multipart/form-data',
            follow_redirects=True,
        )

        self.assertEqual(200, response.status_code)
        self.assertIn('Preview only', response.get_data(as_text=True))

    def test_student_import_creates_updates_and_assigns_groups(self):
        client, _ = self._admin_client()
        _, group, _ = self._base_data()
        existing = User(
            name='Old Name', email='old.student@sit.edu.sg', role='student',
            password_hash='existing-password-hash',
        )
        db.session.add(existing)
        db.session.commit()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Students'
        sheet.append([
            'Account ID', 'Name', 'Email', 'Student Group Label',
            'Temporary Password (new or reset only)',
        ])
        sheet.append([
            existing.id, 'Updated Student', 'updated.student@sit.edu.sg',
            group.group_label, '',
        ])
        sheet.append([
            '', 'New Student', 'new.student@sit.edu.sg',
            group.group_label, 'LocalTest123!',
        ])
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = client.post(
            '/admin/students/import',
            data={'file': (stream, 'students.xlsx')},
            content_type='multipart/form-data',
        )
        db.session.refresh(existing)
        created = User.query.filter_by(email='new.student@sit.edu.sg').one()

        self.assertEqual(302, response.status_code)
        self.assertEqual('Updated Student', existing.name)
        self.assertEqual('updated.student@sit.edu.sg', existing.email)
        self.assertEqual(group.id, existing.student_group_id)
        self.assertEqual('existing-password-hash', existing.password_hash)
        self.assertEqual('student', created.role)
        self.assertEqual(group.id, created.student_group_id)
        self.assertTrue(created.check_password('LocalTest123!'))

    def test_student_import_rejects_entire_file_on_role_email_collision(self):
        client, admin = self._admin_client()
        existing = User(
            name='Original Student', email='original.student@sit.edu.sg',
            role='student', password_hash='unchanged',
        )
        db.session.add(existing)
        db.session.commit()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Students'
        sheet.append([
            'Account ID', 'Name', 'Email', 'Student Group Label',
            'Temporary Password (new or reset only)',
        ])
        sheet.append([
            existing.id, 'Should Not Save', existing.email, '', '',
        ])
        sheet.append([
            '', 'Invalid New Student', admin.email, '', 'LocalTest123!',
        ])
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = client.post(
            '/admin/students/import',
            data={'file': (stream, 'students.xlsx')},
            content_type='multipart/form-data',
        )
        db.session.refresh(existing)

        self.assertEqual(302, response.status_code)
        self.assertEqual('Original Student', existing.name)
        self.assertEqual(1, User.query.filter_by(role='student').count())

    def test_subgroup_regeneration_preserves_assigned_student(self):
        client, _ = self._admin_client()
        programme, _, _ = self._base_data()
        parent = StudentGroup(
            programme=programme, year_level=2,
            group_label='TST-Y2', intake_size=20,
        )
        child = StudentGroup(
            programme=programme, year_level=2,
            group_label='TST-Y2-A', intake_size=20, parent=parent,
        )
        student = User(
            name='Student', email='student@sit.edu.sg', role='student',
            password_hash='unused', student_group=child,
        )
        db.session.add_all([parent, child, student])
        db.session.commit()
        child_id = child.id

        response = client.post(
            f'/admin/student-groups/{parent.id}/generate-subgroups',
            data={'num_subgroups': '2'},
        )
        db.session.refresh(student)

        self.assertEqual(302, response.status_code)
        self.assertEqual(child_id, student.student_group_id)
        self.assertIsNotNone(db.session.get(StudentGroup, child_id))

    def test_layer_selection_falls_back_per_session(self):
        rows = [
            SimpleNamespace(class_session_id=1, is_backbone=True, id=1),
            SimpleNamespace(class_session_id=1, is_backbone=False, id=2),
            SimpleNamespace(class_session_id=2, is_backbone=True, id=3),
        ]

        self.assertEqual({2, 3}, {row.id for row in select_preferred_layer(rows)})

    def test_course_sessions_get_does_not_expand_split_rows(self):
        client, _ = self._admin_client()
        _, group, course = self._base_data()
        course.split_count = 2
        db.session.add(ClassSession(
            course=course, session_type='tutorial', delivery_mode='f2f',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='1,2,3', group_label='All',
        ))
        db.session.commit()

        response = client.get(f'/admin/courses/{course.id}/sessions')

        self.assertEqual(200, response.status_code)
        self.assertEqual(1, ClassSession.query.filter_by(course_id=course.id).count())

    def test_timetable_entry_suggestions_returns_json(self):
        client, _ = self._admin_client()
        _, group, course = self._base_data()
        slot = TimeSlot(
            day_of_week='Monday', period_label='P1',
            start_time=time(9), end_time=time(11),
        )
        session = ClassSession(
            course=course, session_type='lecture', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='1,2,3', group_label='All',
        )
        db.session.add_all([slot, session])
        db.session.flush()
        entry = TimetableEntry(
            class_session_id=session.id, timeslot_id=slot.id,
            week_number=1, trimester='AY2526-T1',
            academic_year='AY2526', is_backbone=False,
        )
        db.session.add(entry)
        db.session.commit()

        response = client.get(
            f'/admin/timetable/entries/{entry.id}/suggestions'
        )

        self.assertEqual(200, response.status_code)
        self.assertIsInstance(response.get_json()['suggestions'], list)

    def test_saved_solve_stats_are_loaded(self):
        db.session.add(SolveRun(
            trimester='AY2526-T1', solver_status='Feasible',
            stats_json='{"entries_created": 4}',
        ))
        db.session.commit()

        self.assertEqual(4, _load_solve_run('AY2526-T1')['entries_created'])

    def test_publication_audit_rejects_term_break_occurrence(self):
        _, group, course = self._base_data()
        slot = TimeSlot(
            day_of_week='Monday', period_label='P1',
            start_time=time(9), end_time=time(11),
        )
        session = ClassSession(
            course=course, session_type='lecture', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
            teaching_weeks='1,2,3,4,5,6,7,8', group_label='All',
        )
        calendar = AcademicCalendar(
            trimester='AY2526-T1', week_number=7,
            start_date=date(2025, 10, 13), end_date=date(2025, 10, 17),
            is_term_break=True, is_public_holiday=False,
        )
        db.session.add_all([slot, session, calendar])
        db.session.flush()
        entry = TimetableEntry(
            class_session_id=session.id, timeslot_id=slot.id,
            week_number=7, trimester='AY2526-T1',
            academic_year='AY2526', is_backbone=False,
        )
        db.session.add(entry)
        db.session.commit()

        conflicts = _audit_generated_hard_conflicts([entry])

        self.assertTrue(any('term-break Week 7' in item for item in conflicts))
        client, _ = self._admin_client()
        response = client.get('/admin/timetable/report?trimester=AY2526-T1')
        page = response.get_data(as_text=True)
        self.assertEqual(200, response.status_code)
        self.assertIn('This timetable must not be published or exported.', page)
        self.assertNotIn('This timetable works.', page)

    def test_template2_extension_is_restored(self):
        base_path = os.path.join(self.app.root_path, 'static', 'template2_base.xlsx')
        workbook = openpyxl.load_workbook(base_path)
        for sheet_name in (
            'Timetable', 'Course Code', 'Location', 'Staff',
            'Sheet4', 'Sheet1', 'Sheet2', 'Sheet3',
        ):
            self.assertEqual(
                1, workbook[sheet_name].max_row,
                f'{sheet_name} must contain headers only in the tracked template',
            )
        output = io.BytesIO()
        workbook.save(output)
        restore_worksheet_extensions(base_path, output)

        with zipfile.ZipFile(output) as archive:
            worksheet_xml = b''.join(
                archive.read(name) for name in archive.namelist()
                if name.startswith('xl/worksheets/') and name.endswith('.xml')
            )
        self.assertIn(b'x14:dataValidations', worksheet_xml)
        # The copied extension uses xr:uid attributes whose namespace was
        # inherited from the original worksheet root.  The rebuilt workbook
        # must remain valid XML and reopen successfully after restoration.
        output.seek(0)
        reopened = openpyxl.load_workbook(output)
        self.assertIn('Timetable', reopened.sheetnames)

    def test_security_headers_are_present(self):
        response = self.app.test_client().get('/login')

        self.assertEqual('DENY', response.headers['X-Frame-Options'])
        self.assertEqual('nosniff', response.headers['X-Content-Type-Options'])
        self.assertIn("frame-ancestors 'none'", response.headers['Content-Security-Policy'])

    def test_course_scoped_recurring_event_matches_only_its_course(self):
        programme, group, course = self._base_data()
        other = Course(
            programme=programme, module_code='TST1002', trimester=1,
            title='Other', year_level=1, delivery_mode='f2f',
            sessions_per_week=1, total_hours=24, split_count=1,
        )
        first = ClassSession(
            course=course, session_type='lecture', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
        )
        second = ClassSession(
            course=other, session_type='lecture', delivery_mode='online',
            duration_hours=2, student_group=group, trimester=1,
        )
        event = Event(
            name='Annual course event', event_date=date(2025, 9, 1),
            is_full_day=True, scope='course', course=course,
            outcome='reschedule', trimester=1, is_recurring=True,
        )
        db.session.add_all([other, first, second, event])
        db.session.commit()

        occurrence = dict(
            occurrence_date=date(2026, 9, 1), timeslot_id=1,
            academic_year='AY2627', trimester=1,
        )
        self.assertTrue(event_affects_occurrence(event, first, **occurrence))
        self.assertFalse(event_affects_occurrence(event, second, **occurrence))

    def test_explicit_student_section_overrides_fallback_selection(self):
        _, group, course = self._base_data()
        student = User(
            name='Student', email='section@example.com', role='student',
            password_hash='unused', student_group=group,
        )
        slot = TimeSlot(
            day_of_week='Monday', period_label='P1',
            start_time=time(9), end_time=time(11),
        )
        sessions = [
            ClassSession(
                course=course, session_type='tutorial', delivery_mode='online',
                duration_hours=2, student_group=group, trimester=1,
                teaching_weeks='1,2,3', group_label=label,
            ) for label in ('T1', 'T2')
        ]
        db.session.add_all([student, slot, *sessions])
        db.session.flush()
        entries = [
            TimetableEntry(
                class_session_id=session.id, timeslot_id=slot.id,
                week_number=1, trimester='AY2526-T1', is_backbone=False,
            ) for session in sessions
        ]
        db.session.add_all(entries)
        db.session.add(StudentSectionAssignment(
            user_id=student.id, class_session_id=sessions[0].id,
            academic_year='AY2526', trimester=1,
        ))
        db.session.commit()

        selected = apply_explicit_student_sections(
            entries, student.id, 'AY2526', 1
        )
        self.assertEqual({sessions[0].id}, {row.class_session_id for row in selected})

    def test_room_equipment_and_accessibility_are_hard_requirements(self):
        session = SimpleNamespace(
            session_type='lecture', accessibility_required=True,
            required_equipment_tags={'lecture capture'},
            student_group=None, effective_group_size=1,
        )
        suitable = SimpleNamespace(
            room_type='lecture', is_accessible=True,
            equipment_tags={'lecture capture'}, capacity=30,
        )
        inaccessible = SimpleNamespace(
            room_type='lecture', is_accessible=False,
            equipment_tags={'lecture capture'}, capacity=30,
        )
        missing_equipment = SimpleNamespace(
            room_type='lecture', is_accessible=True,
            equipment_tags=set(), capacity=30,
        )
        self.assertTrue(_room_compatible(suitable, session))
        self.assertFalse(_room_compatible(inaccessible, session))
        self.assertFalse(_room_compatible(missing_equipment, session))

    def test_inactive_account_cannot_login(self):
        user = User(
            name='Inactive', email='inactive@example.com', role='student',
            active=False,
        )
        user.set_password('TestPassword123')
        db.session.add(user)
        db.session.commit()

        response = self.app.test_client().post('/login', data={
            'email': user.email, 'password': 'TestPassword123',
        })
        self.assertEqual(200, response.status_code)
        self.assertIn(b'Incorrect email or password', response.data)


if __name__ == '__main__':
    unittest.main()
