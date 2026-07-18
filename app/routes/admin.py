import math
import threading
import uuid
import time as _time
from flask import Blueprint, render_template, redirect, url_for, flash, request, abort, jsonify
from flask_login import login_required, current_user
from app import db
from app.models.course import Course
from app.models.professor import Professor
from app.models.room import Room
from app.models.timetable_flag import TimetableFlag
from app.models.availability_declaration import AvailabilityDeclaration
from app.models.user import User
from app.models.student_group import StudentGroup
from app.models.programme import Programme
from app.models.class_session import ClassSession
from app.models.class_session_professor import ClassSessionProfessor
from app.models.timetable_entry import TimetableEntry
from app.models.timeslot import TimeSlot
from app.models.audit_log import AuditLog

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# In-memory store for background solver tasks: {task_id: {status, ...}}
# Entries expire after 2 hours (cleaned up lazily on new task start).
_solver_tasks: dict = {}
_TASK_TTL = 7200  # seconds

def _save_solve_run(trimester, stats):
    """Persist a successful solve()'s stats for this trimester (one row per
    trimester, upserted). The primary generate flow is async - it posts to
    /timetable/solve-async, then the page redirects to a plain GET, so the
    stats dict from that POST is gone by the time the page renders. Storing
    in the DB (not just in-memory) means the constraint summary and
    Scheduling Report survive a server restart and don't depend on which
    process ran the solve."""
    import json
    from datetime import datetime, timezone
    from app.models.solve_run import SolveRun
    row = SolveRun.query.filter_by(trimester=trimester).first()
    if row is None:
        row = SolveRun(trimester=trimester)
        db.session.add(row)
    row.solver_status = stats.get('solver_status', 'Feasible')
    row.stats_json = json.dumps(stats)
    row.updated_at = datetime.now(timezone.utc)
    db.session.commit()


def _load_solve_run(trimester):
    """Return the last persisted solve() stats dict for this trimester, or
    {} if none exists yet."""
    import json
    from app.models.solve_run import SolveRun
    row = SolveRun.query.filter_by(trimester=trimester).first()
    if row is None:
        return {}
    try:
        return json.loads(row.stats_json)
    except (TypeError, ValueError):
        return {}

# Official SIT academic calendar - week 1 start dates (all Mondays).
# Source: https://www.singaporetech.edu.sg/admissions/undergraduate/academic-calendar-sit-and-joint-programmes
SIT_ACADEMIC_CALENDAR = {
    'AY2425': {1: '2024-09-02', 2: '2025-01-06', 3: '2025-05-05'},
    'AY2526': {1: '2025-09-01', 2: '2026-01-05', 3: '2026-05-04'},
    'AY2627': {1: '2026-08-31', 2: '2027-01-04', 3: '2027-05-03'},
}

# ---------------------------------------------------------------------------
# Template 2 export mapping tables - hand-transcribed, not sourced from any
# uploaded data. Shared between timetable_export_template2() and system_info()
# so the "assumed values" disclosure page always reflects what the export
# actually uses (single source of truth - no risk of the two drifting apart).
# ---------------------------------------------------------------------------
T2_CLASS_TYPE = {
    'lecture': 'Lecture', 'lectorial': 'Lectorial', 'tutorial': 'Tutorial',
    'lab': 'Laboratory', 'seminar': 'Seminar', 'workshop': 'Workshop', 'quiz': 'Quiz',
}
T2_ACT_CODE = {
    'lecture': 'LEC', 'lectorial': 'LET', 'tutorial': 'TUT',
    'lab': 'LAB', 'seminar': 'SEM', 'workshop': 'WOR', 'quiz': 'QUZ',
}
# Both tables verified 2026-07-10 against Ms. Yang's own "Class Type" reference
# sheet (Worksheet in ITP Project Requirements (Template 2).xlsx) - her sheet
# names Laboratory (not "Lab") and codes workshop as WOR (not "WRK").
T2_DAY_ABBR = {
    'Monday': 'Mon', 'Tuesday': 'Tue', 'Wednesday': 'Wed',
    'Thursday': 'Thu', 'Friday': 'Fri',
}
T2_CLUSTER_ABBR = {
    'ENG': 'ENG', 'Engineering': 'ENG', 'ICT': 'ICT',
    'University-Wide': 'UWM', 'Business': 'BUS', 'Health': 'HLS',
}
T2_PROG_SECTOR = {
    'ASE': ('PUNGGOL', 'PU'), 'CVE': ('PUNGGOL', 'PU'), 'SDE': ('PUNGGOL', 'PU'),
    'NAME': ('PUNGGOL', 'PU'), 'RSE': ('PUNGGOL', 'PU'), 'EDE': ('PUNGGOL', 'PU'),
    'EEE': ('PUNGGOL', 'PU'), 'EPE': ('PUNGGOL', 'PU'), 'METS': ('PUNGGOL', 'PU'),
    'MEC': ('PUNGGOL', 'PU'), 'MDME': ('PUNGGOL', 'PU'), 'SBE': ('PUNGGOL', 'PU'),
    'ESE': ('PUNGGOL', 'PU'), 'DSC': ('PUNGGOL', 'PU'), 'CPC': ('PUNGGOL', 'PU'),
    'ISE': ('PUNGGOL', 'PU'),
}
# All SIT programmes are now consolidated at the Punggol campus (confirmed by
# Brian 2026-07-10) - the previous per-programme DOVER/TP/SP/NYP/NP mapping
# modelled the old distributed-campus setup and is no longer correct.
T2_PROG_SECTOR_DEFAULT = ('PUNGGOL', 'PU')  # fallback for any unmapped programme code


@admin_bp.before_request
@login_required
def require_admin():
    """Reject any non-admin user trying to access admin routes."""
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))
    if current_user.role != 'admin':
        abort(403)


def _data_quality_gaps():
    """Modules missing a split count, and sessions with no professor assigned -
    shared by the Dashboard's Needs Attention panel and System Info's Data
    Gaps tab so the two can't silently drift apart (found hand-duplicated in
    both places, 2026-07-15)."""
    from sqlalchemy import exists as sa_exists

    has_session = sa_exists().where(ClassSession.course_id == Course.id)
    courses_missing_split = Course.query.filter(
        Course.delivery_mode.in_(['f2f', 'hybrid']),
        Course.split_count.is_(None),
        ~has_session,
    ).order_by(Course.year_level, Course.module_code).all()

    no_prof_sessions = ClassSession.query.filter(
        ~sa_exists().where(ClassSessionProfessor.session_id == ClassSession.id)
    ).all()

    return courses_missing_split, no_prof_sessions


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@admin_bp.route('/dashboard')
@login_required
def dashboard():
    courses_missing_split, no_prof_sessions = _data_quality_gaps()

    stats = {
        'total_courses':         Course.query.count(),
        'courses_missing_split': len(courses_missing_split),
        'total_professors':      Professor.query.count(),
        'total_rooms':           Room.query.filter_by(is_active=True).count(),
        'open_flags':            TimetableFlag.query.filter_by(status='open').count(),
        'pending_declarations':  AvailabilityDeclaration.query.filter_by(status='pending').count(),
    }

    # Per-trimester status - real data, not a placeholder. Every trimester
    # that has ANY TimetableEntry row is "generated"; every one where at
    # least one entry is_published is "published". Room utilisation is
    # "how much of the room inventory actually gets used", not room-slot
    # combos vs the full room x timeslot universe (that denominator made a
    # perfectly normal week look like 0.7% utilised - found 2026-07-11).
    from app.models.solve_run import SolveRun
    total_active_rooms = Room.query.filter_by(is_active=True).count()
    trimester_status = []
    all_trimesters = sorted({
        t[0] for t in db.session.query(TimetableEntry.trimester).distinct().all()
    })
    for tri in all_trimesters:
        sessions_scheduled = db.session.query(TimetableEntry.class_session_id)\
            .filter(TimetableEntry.trimester == tri).distinct().count()
        rooms_used = db.session.query(TimetableEntry.room_id)\
            .filter(TimetableEntry.trimester == tri, TimetableEntry.room_id.isnot(None))\
            .distinct().count()
        room_util = round(rooms_used / total_active_rooms * 100, 1) if total_active_rooms else 0
        profs_covered = db.session.query(ClassSessionProfessor.professor_id)\
            .join(TimetableEntry, TimetableEntry.class_session_id == ClassSessionProfessor.session_id)\
            .filter(TimetableEntry.trimester == tri).distinct().count()
        is_published = db.session.query(TimetableEntry.id)\
            .filter(TimetableEntry.trimester == tri, TimetableEntry.is_published.is_(True)).first() is not None
        solve_row = SolveRun.query.filter_by(trimester=tri).first()
        trimester_status.append({
            'trimester':          tri,
            'sessions_scheduled': sessions_scheduled,
            'room_util_pct':      room_util,
            'profs_covered':      profs_covered,
            'is_published':       is_published,
            'generated_at':       (solve_row.updated_at or solve_row.created_at) if solve_row else None,
        })

    kpis = trimester_status  # renamed conceptually to "per-trimester KPIs" - see template

    # Needs Attention - one consolidated list instead of scattering the same
    # kind of information (things that need admin action) across several
    # always-visible cards, most of which show nothing most of the time.
    from app.engine.checker import get_blocking_issues
    no_prof_count = len(no_prof_sessions)
    blockers, _warnings = get_blocking_issues()
    needs_attention = [
        {'label': 'Modules missing a split count', 'count': stats['courses_missing_split'],
         'href': url_for('admin.courses'), 'icon': 'bi-diagram-2'},
        {'label': 'Classes with no professor assigned', 'count': no_prof_count,
         'href': url_for('admin.system_info'), 'icon': 'bi-person-x'},
        {'label': 'Blocking issues (generation would fail)', 'count': len(blockers),
         'href': url_for('admin.timetable'), 'icon': 'bi-exclamation-octagon'},
        {'label': 'Open conflict flags', 'count': stats['open_flags'],
         'href': url_for('admin.timetable_flags'), 'icon': 'bi-flag'},
        {'label': 'Pending availability declarations', 'count': stats['pending_declarations'],
         'href': url_for('admin.declarations'), 'icon': 'bi-clock-history'},
    ]

    # Recent Activity - manual timetable edits made through the admin UI
    # (bulk/script-driven changes don't go through here, so this reflects
    # what a human deliberately changed, not every DB write).
    recent_activity = (AuditLog.query
                        .order_by(AuditLog.timestamp.desc())
                        .limit(8).all())

    return render_template(
        'admin/dashboard.html',
        stats=stats,
        courses_missing_split=courses_missing_split,
        kpis=kpis,
        needs_attention=needs_attention,
        recent_activity=recent_activity,
    )


# ---------------------------------------------------------------------------
# Courses
# ---------------------------------------------------------------------------

SESSION_TYPES = ('lecture', 'lab', 'seminar', 'tutorial', 'lectorial', 'workshop', 'quiz')
SESSION_DELIVERY_MODES = ('f2f', 'online')
COURSE_DELIVERY_MODES = ('f2f', 'online', 'hybrid')

# Template 2's "Group" column letter per session_type - confirmed from Ms Yang's
# reference file (Lecture/Lectorial/Tutorial/Workshop/Quiz all observed directly
# in real rows); Lab and Seminar have no real example in that file, so P and S
# are a disclosed best-fit (P avoids colliding with Lecture's L; S matches
# "Seminar" itself) rather than a confirmed convention.
GROUP_LABEL_PREFIX = {
    'lecture': 'L', 'lectorial': 'L', 'tutorial': 'T',
    'lab': 'P', 'seminar': 'S', 'workshop': 'W', 'quiz': 'Q',
}


def _recompute_group_labels(course_id, session_type):
    """Template 2's "Group" column: a letter (matching Class Type) + a
    sequential number identifying which parallel split-section a session
    belongs to - e.g. a Tutorial split into 4 parallel groups is T1/T2/T3/T4.
    "All" means unsplit (the whole cohort attends together), and is what
    every session defaulted to before this existed, since nothing ever
    computed a real value. Quiz is the one confirmed exception - the
    reference file always numbers it (Q1), never "All", even for a single
    quiz group. Ordered by id for determinism. Called after anything that
    changes how many sessions of one type a module has, so labels never
    drift stale."""
    siblings = (ClassSession.query
                .filter_by(course_id=course_id, session_type=session_type)
                .order_by(ClassSession.id)
                .all())
    if not siblings:
        return
    letter = GROUP_LABEL_PREFIX.get(session_type, 'X')
    if len(siblings) == 1 and session_type != 'quiz':
        siblings[0].group_label = 'All'
    else:
        for i, s in enumerate(siblings, start=1):
            s.group_label = f'{letter}{i}'


@admin_bp.route('/courses')
@login_required
def courses():
    all_courses = Course.query.order_by(Course.year_level, Course.module_code).all()
    return render_template('admin/courses.html', courses=all_courses)


@admin_bp.route('/courses/add', methods=['GET', 'POST'])
@login_required
def course_add():
    programmes = Programme.query.order_by(Programme.code).all()
    form = request.form

    if request.method == 'POST':
        module_code = request.form.get('module_code', '').strip().upper()
        title = request.form.get('title', '').strip()
        prog_id_raw = request.form.get('programme_id', '').strip()
        year_level_raw = request.form.get('year_level', '').strip()
        trimester_raw = request.form.get('trimester', '').strip()
        delivery_mode = request.form.get('delivery_mode', '').strip().lower()
        split_count_raw = request.form.get('split_count', '').strip()
        remarks = request.form.get('remarks', '').strip()
        official_year_range = request.form.get('official_year_range', '').strip()

        errors = []
        if not module_code:
            errors.append('Module Code is required.')
        if not title:
            errors.append('Module Title is required.')

        programme = Programme.query.get(int(prog_id_raw)) if prog_id_raw.isdigit() else None
        if not programme:
            errors.append('Please select a programme.')

        year_level = int(year_level_raw) if year_level_raw.isdigit() else None
        if not year_level or year_level < 1:
            errors.append('Please select a year level.')

        trimester = int(trimester_raw) if trimester_raw.isdigit() else None

        if delivery_mode not in COURSE_DELIVERY_MODES:
            errors.append('Please select a delivery mode.')

        split_count = None
        if delivery_mode in ('f2f', 'hybrid') and split_count_raw.isdigit() and int(split_count_raw) > 0:
            split_count = int(split_count_raw)

        if not errors and programme:
            existing = Course.query.filter_by(
                module_code=module_code, programme_id=programme.id, trimester=trimester
            ).first()
            if existing:
                errors.append(
                    f'{module_code} already exists for {programme.code}'
                    + (f' Trimester {trimester}' if trimester else '') + '.'
                )

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/course_add.html', programmes=programmes, form=form)

        course = Course(
            programme_id=programme.id, module_code=module_code, title=title,
            year_level=year_level, trimester=trimester, delivery_mode=delivery_mode,
            sessions_per_week=1, total_hours=0, split_count=split_count,
            remarks=remarks or None, official_year_range=official_year_range or None,
        )
        db.session.add(course)
        db.session.commit()

        flash(f'{module_code} added - add its first class session below.', 'success')
        return redirect(url_for('admin.course_sessions', course_id=course.id))

    return render_template('admin/course_add.html', programmes=programmes, form=form)


@admin_bp.route('/courses/<int:course_id>/edit', methods=['GET', 'POST'])
@login_required
def course_edit(course_id):
    course = Course.query.get_or_404(course_id)

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        remarks = request.form.get('remarks', '').strip()
        split_count_raw = request.form.get('split_count', '').strip()
        official_year_range = request.form.get('official_year_range', '').strip()

        # Validate
        if not title:
            flash('Module title cannot be empty.', 'danger')
            return render_template('admin/course_edit.html', course=course)

        split_count = None
        if course.delivery_mode in ('f2f', 'hybrid'):
            if split_count_raw == '':
                split_count = None      # Admin left it blank - still not set
            else:
                try:
                    split_count = int(split_count_raw)
                    if split_count < 1:
                        raise ValueError
                except ValueError:
                    flash('Split count must be a whole number of 1 or more.', 'danger')
                    return render_template('admin/course_edit.html', course=course)

        course.title = title
        course.remarks = remarks or None
        course.split_count = split_count
        course.official_year_range = official_year_range or None
        db.session.commit()

        flash(f'{course.module_code} updated successfully.', 'success')
        return redirect(url_for('admin.courses'))

    return render_template('admin/course_edit.html', course=course)


# ---------------------------------------------------------------------------
# Courses - page-level Import / Export (same pattern as Professors/Rooms/
# Student Groups above). Scoped to exactly what the Edit form already
# supports - Title, Remarks, Split Count - since that's the only thing a
# single Course can be edited to on this page; there's no "Add Course" form
# either (modules are only ever created via Template 1's bulk import), so
# new-row creation isn't offered here - matching rows must already exist.
# ---------------------------------------------------------------------------

@admin_bp.route('/courses/export')
@login_required
def course_export():
    import io
    import openpyxl
    from openpyxl.styles import Font
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Modules'
    headers = ['Module Code', 'Programme Code', 'Trimester', 'Year Level', 'Title',
              'Delivery Mode', 'Split Count', 'Official Year Range', 'Remarks']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    all_courses = Course.query.join(Course.programme).order_by(
        Course.year_level, Course.module_code).all()
    for c in all_courses:
        ws.append([c.module_code, c.programme.code, c.trimester, c.year_level, c.title,
                  c.delivery_mode, c.split_count, c.official_year_range or '', c.remarks or ''])

    for i, width in enumerate([16, 16, 12, 12, 34, 14, 14, 20, 40], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    notes = wb.create_sheet('Read Me')
    notes.append(['How to use this file'])
    notes['A1'].font = Font(bold=True, size=13)
    for line in [
        '',
        'This import can only update Title, Split Count, Official Year Range, and Remarks '
        '- the same fields the Edit Module page lets you change. Module Code, Programme '
        'Code, Trimester, Year Level, and Delivery Mode are shown for context (they\'re '
        'how rows are matched) but can\'t be changed here.',
        'Official Year Range is a reference note only (e.g. "Year 2-4", from SIT\'s own '
        'module catalog) - it never affects scheduling, which always uses Year Level.',
        'Rows are matched by Module Code + Programme Code + Trimester together - the '
        'same module code can legitimately appear more than once (different programmes '
        'or trimesters), so all three must match an existing module.',
        'New modules can\'t be created through this import - modules are only ever '
        'created through a full bulk import (Template 1, or Admin Tools > Import), '
        'since a module needs its class sessions set up at the same time.',
        'Split Count can be left blank to clear it. If any row fails validation, nothing '
        'in this file is imported - fix the error and re-upload the whole file.',
    ]:
        notes.append([line])
    notes.column_dimensions['A'].width = 100

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='modules.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/courses/import', methods=['POST'])
@login_required
def course_import():
    import pandas as pd

    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('admin.courses'))

    try:
        df = pd.read_excel(file, sheet_name='Modules', dtype=str).fillna('')
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('admin.courses'))

    df.columns = df.columns.str.strip().str.lower()
    required_cols = {'module code', 'programme code', 'trimester', 'title', 'split count',
                     'official year range', 'remarks'}
    missing = required_cols - set(df.columns)
    if missing:
        flash(f'Missing column(s): {", ".join(sorted(missing))}. Re-download the template and try again.', 'danger')
        return redirect(url_for('admin.courses'))

    programmes_by_code = {p.code: p for p in Programme.query.all()}
    existing_by_key = {}
    for c in Course.query.join(Course.programme).all():
        existing_by_key[(c.module_code, c.programme.code, c.trimester)] = c

    rows = []
    errors = []
    seen_keys = {}

    for i, row in df.iterrows():
        excel_row = i + 2
        module_code = row.get('module code', '').strip().upper()
        prog_code = row.get('programme code', '').strip().upper()
        tri_raw = row.get('trimester', '').strip()
        title = row.get('title', '').strip()
        split_raw = row.get('split count', '').strip()
        official_year_range = row.get('official year range', '').strip()
        remarks = row.get('remarks', '').strip()

        if not module_code and not prog_code and not title:
            continue  # fully blank row

        row_errors = []
        if not module_code: row_errors.append('Module Code is required')
        if not prog_code:   row_errors.append('Programme Code is required')
        if prog_code and prog_code not in programmes_by_code:
            row_errors.append(f'Programme Code "{prog_code}" was not found')
        if not title:       row_errors.append('Title is required')

        trimester = None
        if tri_raw:
            try:
                trimester = int(tri_raw)
            except ValueError:
                row_errors.append('Trimester must be a number (1, 2, or 3)')

        split_count = None
        if split_raw:
            try:
                split_count = int(split_raw)
                if split_count < 1:
                    raise ValueError
            except ValueError:
                row_errors.append('Split Count must be blank or a whole number of 1 or more')

        key = (module_code, prog_code, trimester)
        if key in seen_keys:
            row_errors.append(f'Same Module Code + Programme Code + Trimester also used on row {seen_keys[key]}')
        else:
            seen_keys[key] = excel_row

        existing_course = existing_by_key.get(key)
        if existing_course is None:
            row_errors.append(
                'No matching course found for this Module Code + Programme Code + Trimester - '
                'new courses can\'t be created through this import'
            )

        if row_errors:
            errors.append(f'Row {excel_row} ({module_code or "?"}): ' + '; '.join(row_errors))
            continue

        rows.append({
            'excel_row': excel_row, 'existing': existing_course,
            'title': title, 'split_count': split_count,
            'official_year_range': official_year_range or None, 'remarks': remarks or None,
        })

    if errors:
        flash(f'Import rejected - {len(errors)} problem(s) found. Nothing was changed.', 'danger')
        for e in errors[:15]:
            flash(e, 'warning')
        if len(errors) > 15:
            flash(f'...and {len(errors) - 15} more. Fix these and re-upload the whole file.', 'warning')
        return redirect(url_for('admin.courses'))

    updated = 0
    for r in rows:
        course = r['existing']
        course.title = r['title']
        course.split_count = r['split_count']
        course.official_year_range = r['official_year_range']
        course.remarks = r['remarks']
        updated += 1

    db.session.commit()
    flash(f'Import complete - {updated} module(s) updated.', 'success')
    return redirect(url_for('admin.courses'))


# ---------------------------------------------------------------------------
# Professors
# ---------------------------------------------------------------------------

@admin_bp.route('/professors')
@login_required
def professors():
    all_professors = (Professor.query
                      .join(Professor.user)
                      .order_by(User.name)
                      .all())
    return render_template('admin/professors.html', professors=all_professors)


@admin_bp.route('/professors/add', methods=['GET', 'POST'])
@login_required
def professor_add():
    if request.method == 'POST':
        name       = request.form.get('name', '').strip()
        email      = request.form.get('email', '').strip().lower()
        staff_id   = request.form.get('staff_id', '').strip()
        department = request.form.get('department', '').strip()
        password   = request.form.get('password', '').strip()

        errors = []
        if not name:       errors.append('Name is required.')
        if not email:      errors.append('Email is required.')
        if not staff_id:   errors.append('Staff ID is required.')
        if not department: errors.append('Department is required.')
        if not password:   errors.append('Temporary password is required.')

        if User.query.filter_by(email=email).first():
            errors.append('An account with this email already exists.')
        if Professor.query.filter_by(staff_id=staff_id).first():
            errors.append('A professor with this Staff ID already exists.')

        if errors:
            departments = sorted({p.department for p in Professor.query.all() if p.department})
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/professor_add.html', form=request.form, departments=departments)

        # Create User account + Professor profile in one transaction
        user = User(name=name, email=email, role='professor')
        user.set_password(password)
        db.session.add(user)
        db.session.flush()      # get user.id before committing

        professor = Professor(user_id=user.id, staff_id=staff_id, department=department)
        db.session.add(professor)
        db.session.commit()

        flash(f'Professor {name} added successfully.', 'success')
        return redirect(url_for('admin.professors'))

    departments = sorted({p.department for p in Professor.query.all() if p.department})
    return render_template('admin/professor_add.html', form={}, departments=departments)


@admin_bp.route('/professors/<int:professor_id>/edit', methods=['GET', 'POST'])
@login_required
def professor_edit(professor_id):
    professor = Professor.query.get_or_404(professor_id)
    user = professor.user

    if request.method == 'POST':
        name         = request.form.get('name', '').strip()
        email        = request.form.get('email', '').strip().lower()
        staff_id     = request.form.get('staff_id', '').strip()
        department   = request.form.get('department', '').strip()
        new_password = request.form.get('new_password', '').strip()

        errors = []
        if not name:       errors.append('Name is required.')
        if not email:      errors.append('Email is required.')
        if not staff_id:   errors.append('Staff ID is required.')
        if not department: errors.append('Department is required.')

        existing_email = User.query.filter_by(email=email).first()
        if existing_email and existing_email.id != user.id:
            errors.append('Another account is already using this email.')

        existing_sid = Professor.query.filter_by(staff_id=staff_id).first()
        if existing_sid and existing_sid.id != professor.id:
            errors.append('Another professor already has this Staff ID.')

        departments = sorted({p.department for p in Professor.query.all() if p.department})
        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/professor_edit.html', professor=professor, departments=departments)

        user.name            = name
        user.email           = email
        professor.staff_id   = staff_id
        professor.department = department

        if new_password:
            user.set_password(new_password)
            flash('Password has been reset.', 'info')

        db.session.commit()
        flash(f'{name} updated successfully.', 'success')
        return redirect(url_for('admin.professors'))

    departments = sorted({p.department for p in Professor.query.all() if p.department})
    return render_template('admin/professor_edit.html', professor=professor, departments=departments)


# ---------------------------------------------------------------------------
# Professors - page-level Import / Export (additive to the CRUD above, not a
# replacement - the download carries every current professor, not a blank
# template, so bulk edits happen "in place" in Excel before re-uploading.
# Reference implementation for the same pattern on other entity pages.
# ---------------------------------------------------------------------------

@admin_bp.route('/professors/export')
@login_required
def professor_export():
    import io
    import openpyxl
    from openpyxl.styles import Font
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Professors'
    headers = ['Staff ID', 'Name', 'Email', 'Department', 'Temporary Password (new staff only)']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for p in Professor.query.join(Professor.user).order_by(User.name).all():
        ws.append([p.staff_id, p.user.name, p.user.email, p.department, ''])

    for i, width in enumerate([14, 28, 32, 24, 32], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    notes = wb.create_sheet('Read Me')
    notes.append(['How to use this file'])
    notes['A1'].font = Font(bold=True, size=13)
    for line in [
        '',
        'To update an existing professor: edit their row directly (matched by Staff ID - '
        'don\'t change it unless you mean to). Leave "Temporary Password" blank to keep their '
        'current password.',
        'To add a new professor: add a new row with a Staff ID that doesn\'t already exist. '
        '"Temporary Password" is required for new rows.',
        'To remove a professor: delete them from the Professors page directly - this import '
        'never deletes anyone, even if you remove their row here.',
        'If any row fails validation, nothing in this file is imported - fix the error and '
        're-upload the whole file.',
    ]:
        notes.append([line])
    notes.column_dimensions['A'].width = 100

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='professors.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/professors/import', methods=['POST'])
@login_required
def professor_import():
    import pandas as pd

    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('admin.professors'))

    try:
        df = pd.read_excel(file, sheet_name='Professors', dtype=str).fillna('')
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('admin.professors'))

    df.columns = df.columns.str.strip().str.lower()
    required_cols = {'staff id', 'name', 'email', 'department'}
    missing = required_cols - set(df.columns)
    if missing:
        flash(f'Missing column(s): {", ".join(sorted(missing))}. Re-download the template and try again.', 'danger')
        return redirect(url_for('admin.professors'))

    existing_by_staff_id = {p.staff_id: p for p in Professor.query.join(Professor.user).all()}
    existing_by_email = {p.user.email.lower(): p for p in Professor.query.join(Professor.user).all()}

    rows = []
    errors = []
    seen_staff_ids = {}
    seen_emails = {}

    for i, row in df.iterrows():
        excel_row = i + 2  # header is row 1
        staff_id = row.get('staff id', '').strip()
        name = row.get('name', '').strip()
        email = row.get('email', '').strip().lower()
        department = row.get('department', '').strip()
        temp_password = row.get('temporary password (new staff only)', '').strip()

        if not staff_id and not name and not email and not department:
            continue  # fully blank row - skip silently

        row_errors = []
        if not staff_id:   row_errors.append('Staff ID is required')
        if not name:       row_errors.append('Name is required')
        if not email:      row_errors.append('Email is required')
        elif '@' not in email:
            row_errors.append('Email looks invalid')
        if not department: row_errors.append('Department is required')

        if staff_id:
            if staff_id in seen_staff_ids:
                row_errors.append(f'Staff ID also used on row {seen_staff_ids[staff_id]}')
            else:
                seen_staff_ids[staff_id] = excel_row
        if email:
            if email in seen_emails:
                row_errors.append(f'Email also used on row {seen_emails[email]}')
            else:
                seen_emails[email] = excel_row

        existing_prof = existing_by_staff_id.get(staff_id) if staff_id else None
        is_new = existing_prof is None

        if email:
            email_owner = existing_by_email.get(email)
            if email_owner and (is_new or email_owner.id != existing_prof.id):
                row_errors.append(f'Email is already used by another professor ({email_owner.staff_id})')

        if is_new and not temp_password:
            row_errors.append('Temporary Password is required for a new Staff ID')

        if row_errors:
            errors.append(f'Row {excel_row} (Staff ID "{staff_id or "?"}"): ' + '; '.join(row_errors))
            continue

        rows.append({
            'excel_row': excel_row, 'staff_id': staff_id, 'name': name, 'email': email,
            'department': department, 'temp_password': temp_password,
            'is_new': is_new, 'existing': existing_prof,
        })

    if errors:
        flash(f'Import rejected - {len(errors)} problem(s) found. Nothing was changed.', 'danger')
        for e in errors[:15]:
            flash(e, 'warning')
        if len(errors) > 15:
            flash(f'...and {len(errors) - 15} more. Fix these and re-upload the whole file.', 'warning')
        return redirect(url_for('admin.professors'))

    created = updated = 0
    for r in rows:
        if r['is_new']:
            user = User(name=r['name'], email=r['email'], role='professor')
            user.set_password(r['temp_password'])
            db.session.add(user)
            db.session.flush()
            db.session.add(Professor(user_id=user.id, staff_id=r['staff_id'], department=r['department']))
            created += 1
        else:
            prof = r['existing']
            prof.user.name = r['name']
            prof.user.email = r['email']
            prof.department = r['department']
            if r['temp_password']:
                prof.user.set_password(r['temp_password'])
            updated += 1

    db.session.commit()
    flash(f'Import complete - {created} professor(s) added, {updated} updated.', 'success')
    return redirect(url_for('admin.professors'))


# ---------------------------------------------------------------------------
# Students
# ---------------------------------------------------------------------------

@admin_bp.route('/students')
@login_required
def students():
    all_students = (User.query
                    .filter_by(role='student')
                    .order_by(User.name)
                    .all())
    return render_template('admin/students.html', students=all_students)


@admin_bp.route('/students/add', methods=['GET', 'POST'])
@login_required
def student_add():
    all_groups = (StudentGroup.query
                  .order_by(StudentGroup.year_level, StudentGroup.group_label)
                  .all())

    if request.method == 'POST':
        name     = request.form.get('name', '').strip()
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()
        group_id = request.form.get('student_group_id', '').strip()

        errors = []
        if not name:     errors.append('Full name is required.')
        if not email:    errors.append('Email address is required.')
        if not password: errors.append('Password is required.')

        if email and User.query.filter_by(email=email).first():
            errors.append(f'An account with email {email} already exists.')

        group_id_num, err = _parse_id(group_id, 'Student group')
        if err:
            errors.append(err)
        elif group_id_num is not None and StudentGroup.query.get(group_id_num) is None:
            errors.append('That student group no longer exists - please pick another.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/student_add.html',
                                   all_groups=all_groups, form=request.form)

        student = User(
            name             = name,
            email            = email,
            role             = 'student',
            student_group_id = group_id_num,
        )
        student.set_password(password)
        db.session.add(student)
        db.session.commit()
        flash(f'Student account created for {name}.', 'success')
        return redirect(url_for('admin.students'))

    return render_template('admin/student_add.html', all_groups=all_groups, form={})


@admin_bp.route('/students/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def student_edit(user_id):
    student = User.query.get_or_404(user_id)
    if student.role != 'student':
        abort(404)

    all_groups = (StudentGroup.query
                  .order_by(StudentGroup.year_level, StudentGroup.group_label)
                  .all())

    if request.method == 'POST':
        name     = request.form.get('name', '').strip()
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()
        group_id = request.form.get('student_group_id', '').strip()

        errors = []
        if not name:  errors.append('Full name is required.')
        if not email: errors.append('Email address is required.')

        existing = User.query.filter_by(email=email).first()
        if existing and existing.id != student.id:
            errors.append(f'Another account is already using {email}.')

        group_id_num, err = _parse_id(group_id, 'Student group')
        if err:
            errors.append(err)
        elif group_id_num is not None and StudentGroup.query.get(group_id_num) is None:
            errors.append('That student group no longer exists - please pick another.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/student_edit.html',
                                   student=student, all_groups=all_groups)

        student.name             = name
        student.email            = email
        student.student_group_id = group_id_num
        if password:
            student.set_password(password)
            flash('Password has been reset.', 'info')

        db.session.commit()
        flash(f'{name} updated successfully.', 'success')
        return redirect(url_for('admin.students'))

    return render_template('admin/student_edit.html',
                           student=student, all_groups=all_groups)


@admin_bp.route('/students/<int:user_id>/delete', methods=['POST'])
@login_required
def student_delete(user_id):
    student = User.query.get_or_404(user_id)
    if student.role != 'student':
        abort(404)
    name = student.name
    db.session.delete(student)
    db.session.commit()
    flash(f'Student account for {name} deleted.', 'success')
    return redirect(url_for('admin.students'))


# ---------------------------------------------------------------------------
# Rooms
# ---------------------------------------------------------------------------

@admin_bp.route('/rooms')
@login_required
def rooms():
    building_filter = request.args.get('building', '')
    query = Room.query.order_by(Room.building, Room.room_code)
    if building_filter:
        query = query.filter_by(building=building_filter)
    all_rooms = query.all()

    buildings = [r[0] for r in db.session.query(Room.building).distinct().order_by(Room.building).all()]

    return render_template('admin/rooms.html',
                           rooms=all_rooms,
                           buildings=buildings,
                           active_building=building_filter)


@admin_bp.route('/rooms/add', methods=['GET', 'POST'])
@login_required
def room_add():
    if request.method == 'POST':
        room_code = request.form.get('room_code', '').strip().upper()
        building  = request.form.get('building', '').strip().upper()
        capacity  = request.form.get('capacity', '').strip()
        room_type = request.form.get('room_type', '').strip()

        errors = []
        if not room_code: errors.append('Room code is required.')
        if not building:  errors.append('Building is required.')
        if not room_type: errors.append('Room type is required.')

        try:
            capacity = int(capacity)
            if capacity < 1:
                raise ValueError
        except (ValueError, TypeError):
            errors.append('Capacity must be a whole number of 1 or more.')
            capacity = ''

        if Room.query.filter_by(room_code=room_code).first():
            errors.append(f'A room with code {room_code} already exists.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/room_add.html', form=request.form)

        db.session.add(Room(
            room_code=room_code,
            building=building,
            capacity=capacity,
            room_type=room_type,
            is_active=True,
        ))
        db.session.commit()
        flash(f'Room {room_code} added successfully.', 'success')
        return redirect(url_for('admin.rooms'))

    return render_template('admin/room_add.html', form={})


@admin_bp.route('/rooms/<int:room_id>/edit', methods=['GET', 'POST'])
@login_required
def room_edit(room_id):
    room = Room.query.get_or_404(room_id)

    if request.method == 'POST':
        room_code = request.form.get('room_code', '').strip().upper()
        building  = request.form.get('building', '').strip().upper()
        capacity  = request.form.get('capacity', '').strip()
        room_type = request.form.get('room_type', '').strip()

        errors = []
        if not room_code: errors.append('Room code is required.')
        if not building:  errors.append('Building is required.')
        if not room_type: errors.append('Room type is required.')

        try:
            capacity = int(capacity)
            if capacity < 1:
                raise ValueError
        except (ValueError, TypeError):
            errors.append('Capacity must be a whole number of 1 or more.')
            capacity = ''

        existing = Room.query.filter_by(room_code=room_code).first()
        if existing and existing.id != room.id:
            errors.append(f'Another room with code {room_code} already exists.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/room_edit.html', room=room)

        room.room_code = room_code
        room.building  = building
        room.capacity  = capacity
        room.room_type = room_type
        db.session.commit()

        flash(f'Room {room_code} updated successfully.', 'success')
        return redirect(url_for('admin.rooms'))

    return render_template('admin/room_edit.html', room=room)


@admin_bp.route('/rooms/<int:room_id>/toggle', methods=['POST'])
@login_required
def room_toggle(room_id):
    room = Room.query.get_or_404(room_id)
    room.is_active = not room.is_active
    db.session.commit()
    status = 'activated' if room.is_active else 'deactivated'
    flash(f'Room {room.room_code} {status}.', 'info')
    return redirect(url_for('admin.rooms', building=request.form.get('building', '')))


@admin_bp.route('/rooms/<int:room_id>/delete', methods=['POST'])
@login_required
def room_delete(room_id):
    room = Room.query.get_or_404(room_id)

    if room.timetable_entries:
        flash(f'Room {room.room_code} cannot be deleted - it has assigned timetable entries. Deactivate it instead.', 'danger')
        return redirect(url_for('admin.rooms'))

    code = room.room_code
    db.session.delete(room)
    db.session.commit()
    flash(f'Room {code} deleted.', 'success')
    return redirect(url_for('admin.rooms'))


# ---------------------------------------------------------------------------
# Rooms - page-level Import / Export (same pattern as Professors above:
# additive to CRUD, download carries all current data, whole-file validation).
# ---------------------------------------------------------------------------

ROOM_TYPES = ('lecture', 'lab', 'seminar')


@admin_bp.route('/rooms/export')
@login_required
def room_export():
    import io
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rooms'
    headers = ['Room Code', 'Building', 'Room Type', 'Capacity', 'Active']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    all_rooms = Room.query.order_by(Room.building, Room.room_code).all()
    for r in all_rooms:
        ws.append([r.room_code, r.building, r.room_type, r.capacity, 'Yes' if r.is_active else 'No'])

    for i, width in enumerate([16, 18, 14, 12, 10], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Real dropdowns - Room Type and Active are genuinely closed sets in this
    # system (unlike e.g. a professor's Department, which is free text).
    last_row = max(ws.max_row, 2) + 200   # headroom for new rows
    dv_type = DataValidation(type='list', formula1=f'"{",".join(ROOM_TYPES)}"', allow_blank=False)
    dv_type.add(f'C2:C{last_row}')
    ws.add_data_validation(dv_type)
    dv_active = DataValidation(type='list', formula1='"Yes,No"', allow_blank=False)
    dv_active.add(f'E2:E{last_row}')
    ws.add_data_validation(dv_active)

    notes = wb.create_sheet('Read Me')
    notes.append(['How to use this file'])
    notes['A1'].font = Font(bold=True, size=13)
    for line in [
        '',
        'To update an existing room: edit its row directly (matched by Room Code - '
        'don\'t change it unless you mean to).',
        'To add a new room: add a new row with a Room Code that doesn\'t already exist.',
        'To remove a room: delete or deactivate it from the Rooms page directly - this '
        'import never deletes anyone, even if you remove their row here.',
        'Room Type and Active are dropdowns - pick from the list, don\'t type a custom value.',
        'If any row fails validation, nothing in this file is imported - fix the error and '
        're-upload the whole file.',
    ]:
        notes.append([line])
    notes.column_dimensions['A'].width = 100

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='rooms.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/rooms/import', methods=['POST'])
@login_required
def room_import():
    import pandas as pd

    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('admin.rooms'))

    try:
        df = pd.read_excel(file, sheet_name='Rooms', dtype=str).fillna('')
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('admin.rooms'))

    df.columns = df.columns.str.strip().str.lower()
    required_cols = {'room code', 'building', 'room type', 'capacity', 'active'}
    missing = required_cols - set(df.columns)
    if missing:
        flash(f'Missing column(s): {", ".join(sorted(missing))}. Re-download the template and try again.', 'danger')
        return redirect(url_for('admin.rooms'))

    existing_by_code = {r.room_code: r for r in Room.query.all()}

    rows = []
    errors = []
    seen_codes = {}

    for i, row in df.iterrows():
        excel_row = i + 2
        code = row.get('room code', '').strip().upper()
        building = row.get('building', '').strip().upper()
        room_type = row.get('room type', '').strip().lower()
        capacity_raw = row.get('capacity', '').strip()
        active_raw = row.get('active', '').strip().lower()

        if not code and not building and not room_type and not capacity_raw:
            continue  # fully blank row

        row_errors = []
        if not code:      row_errors.append('Room Code is required')
        if not building:  row_errors.append('Building is required')
        if room_type not in ROOM_TYPES:
            row_errors.append(f'Room Type must be one of {", ".join(ROOM_TYPES)}')
        capacity = None
        try:
            capacity = int(capacity_raw)
            if capacity < 1:
                raise ValueError
        except ValueError:
            row_errors.append('Capacity must be a whole number of 1 or more')
        if active_raw not in ('yes', 'no'):
            row_errors.append('Active must be Yes or No')

        if code:
            if code in seen_codes:
                row_errors.append(f'Room Code also used on row {seen_codes[code]}')
            else:
                seen_codes[code] = excel_row

        if row_errors:
            errors.append(f'Row {excel_row} (Room Code "{code or "?"}"): ' + '; '.join(row_errors))
            continue

        rows.append({
            'excel_row': excel_row, 'code': code, 'building': building, 'room_type': room_type,
            'capacity': capacity, 'is_active': active_raw == 'yes',
            'existing': existing_by_code.get(code),
        })

    if errors:
        flash(f'Import rejected - {len(errors)} problem(s) found. Nothing was changed.', 'danger')
        for e in errors[:15]:
            flash(e, 'warning')
        if len(errors) > 15:
            flash(f'...and {len(errors) - 15} more. Fix these and re-upload the whole file.', 'warning')
        return redirect(url_for('admin.rooms'))

    created = updated = 0
    for r in rows:
        if r['existing']:
            room = r['existing']
            room.building = r['building']
            room.room_type = r['room_type']
            room.capacity = r['capacity']
            room.is_active = r['is_active']
            updated += 1
        else:
            db.session.add(Room(room_code=r['code'], building=r['building'], room_type=r['room_type'],
                                capacity=r['capacity'], is_active=r['is_active']))
            created += 1

    db.session.commit()
    flash(f'Import complete - {created} room(s) added, {updated} updated.', 'success')
    return redirect(url_for('admin.rooms'))


# ---------------------------------------------------------------------------
# Student Groups
# ---------------------------------------------------------------------------

@admin_bp.route('/student-groups')
@login_required
def student_groups():
    # Only top-level groups (no parent); sub-groups shown nested under them
    top_level = (StudentGroup.query
                 .filter_by(parent_id=None)
                 .order_by(StudentGroup.year_level, StudentGroup.group_label)
                 .all())
    return render_template('admin/student_groups.html', groups=top_level)


@admin_bp.route('/student-groups/add', methods=['GET', 'POST'])
@login_required
def student_group_add():
    programmes = Programme.query.order_by(Programme.code).all()

    if request.method == 'POST':
        programme_id = request.form.get('programme_id', '').strip()
        year_level   = request.form.get('year_level', '').strip()
        intake_size  = request.form.get('intake_size', '').strip()

        errors = []
        programme = None
        if not programme_id:
            errors.append('Programme is required.')
        else:
            try:
                programme = Programme.query.get(int(programme_id))
            except ValueError:
                errors.append('Programme was not a valid selection - please choose from the dropdown.')
            else:
                if programme is None:
                    errors.append('That programme no longer exists - please pick another.')

        try:
            year_level = int(year_level)
            if year_level < 1:
                raise ValueError
        except (ValueError, TypeError):
            errors.append('Year level must be a positive number.')
            year_level = ''

        try:
            intake_size = int(intake_size)
            if intake_size < 1:
                raise ValueError
        except (ValueError, TypeError):
            errors.append('Intake size must be a positive number.')
            intake_size = ''

        group_label = f'{programme.code}-Y{year_level}' if programme and year_level != '' else ''

        if group_label and StudentGroup.query.filter_by(group_label=group_label).first():
            errors.append(f'A group with label {group_label} already exists.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/student_group_add.html',
                                   programmes=programmes, form=request.form)

        group = StudentGroup(
            programme_id=programme.id,
            year_level=year_level,
            group_label=group_label,
            intake_size=intake_size,
            parent_id=None,
        )
        db.session.add(group)
        db.session.commit()
        flash(f'Student group {group_label} created successfully.', 'success')
        return redirect(url_for('admin.student_groups'))

    return render_template('admin/student_group_add.html',
                           programmes=programmes, form={})


@admin_bp.route('/student-groups/<int:group_id>/edit', methods=['GET', 'POST'])
@login_required
def student_group_edit(group_id):
    group = StudentGroup.query.get_or_404(group_id)
    programmes = Programme.query.order_by(Programme.code).all()

    if request.method == 'POST':
        intake_size = request.form.get('intake_size', '').strip()

        errors = []
        try:
            intake_size = int(intake_size)
            if intake_size < 1:
                raise ValueError
        except (ValueError, TypeError):
            errors.append('Intake size must be a positive number.')
            intake_size = ''

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/student_group_edit.html', group=group)

        group.intake_size = intake_size
        db.session.commit()
        flash(f'Group {group.group_label} updated successfully.', 'success')
        return redirect(url_for('admin.student_groups'))

    return render_template('admin/student_group_edit.html', group=group)


@admin_bp.route('/student-groups/<int:group_id>/generate-subgroups', methods=['POST'])
@login_required
def student_group_generate(group_id):
    parent = StudentGroup.query.get_or_404(group_id)
    num_raw = request.form.get('num_subgroups', '').strip()

    try:
        num = int(num_raw)
        if num < 2:
            raise ValueError
    except (ValueError, TypeError):
        flash('Number of sub-groups must be 2 or more.', 'danger')
        return redirect(url_for('admin.student_groups'))

    # Guard: refuse if existing sub-groups already have sessions assigned (would cause FK violation)
    existing_subs = StudentGroup.query.filter_by(parent_id=parent.id).all()
    if existing_subs:
        sub_ids = [s.id for s in existing_subs]
        blocking = ClassSession.query.filter(
            ClassSession.student_group_id.in_(sub_ids)
        ).first()
        if blocking:
            flash(
                f'Cannot regenerate sub-groups for {parent.group_label} - '
                f'existing sub-groups have sessions assigned. '
                f'Clear all session assignments first.',
                'danger'
            )
            return redirect(url_for('admin.student_groups'))

    # Delete existing sub-groups for this parent before regenerating
    StudentGroup.query.filter_by(parent_id=parent.id).delete()

    sub_size = math.ceil(parent.intake_size / num)
    labels = [chr(65 + i) for i in range(num)]   # A, B, C, ...

    for label in labels:
        db.session.add(StudentGroup(
            programme_id=parent.programme_id,
            year_level=parent.year_level,
            group_label=f'{parent.group_label}-{label}',
            intake_size=sub_size,
            parent_id=parent.id,
        ))

    db.session.commit()
    flash(f'Generated {num} sub-groups for {parent.group_label}.', 'success')
    return redirect(url_for('admin.student_groups'))


@admin_bp.route('/student-groups/<int:group_id>/delete', methods=['POST'])
@login_required
def student_group_delete(group_id):
    group = StudentGroup.query.get_or_404(group_id)

    # Check this group AND all its sub-groups for assigned sessions
    all_ids = [group.id] + [sub.id for sub in group.sub_groups]
    blocking = ClassSession.query.filter(
        ClassSession.student_group_id.in_(all_ids)
    ).first()
    if blocking:
        flash(
            f'Group {group.group_label} cannot be deleted - '
            f'it or its sub-groups have sessions assigned. '
            f'Clear all session assignments first.',
            'danger'
        )
        return redirect(url_for('admin.student_groups'))

    label = group.group_label
    # Delete sub-groups first if this is a parent
    StudentGroup.query.filter_by(parent_id=group.id).delete()
    db.session.delete(group)
    db.session.commit()
    flash(f'Group {label} deleted.', 'success')
    return redirect(url_for('admin.student_groups'))


# ---------------------------------------------------------------------------
# Student Groups - page-level Import / Export (same pattern as Professors/
# Rooms above). Scoped to match exactly what the CRUD already supports:
# only Intake Size can be changed for an existing group (top-level or
# sub-group), and only new TOP-LEVEL groups can be created - sub-groups
# still have to go through "Generate Sub-groups" on the page itself, since
# that's what keeps their auto-numbered labels and split sizes consistent.
# ---------------------------------------------------------------------------

@admin_bp.route('/student-groups/export')
@login_required
def student_group_export():
    import io
    import openpyxl
    from openpyxl.styles import Font
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Student Groups'
    headers = ['Group Label', 'Programme Code', 'Year Level', 'Intake Size', 'Parent Group Label']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    all_groups = StudentGroup.query.join(StudentGroup.programme).order_by(
        StudentGroup.year_level, StudentGroup.group_label).all()
    for g in all_groups:
        ws.append([g.group_label, g.programme.code, g.year_level, g.intake_size,
                  g.parent.group_label if g.parent else ''])

    for i, width in enumerate([20, 16, 12, 14, 20], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    notes = wb.create_sheet('Read Me')
    notes.append(['How to use this file'])
    notes['A1'].font = Font(bold=True, size=13)
    for line in [
        '',
        'To update an existing group\'s Intake Size: edit that row directly (matched by '
        'Group Label). Programme Code, Year Level, and Parent Group Label are shown for '
        'context but can\'t be changed here - they\'re read-only once a group is created.',
        'To add a new top-level group: add a new row with a Group Label that doesn\'t '
        'already exist, leave Parent Group Label blank, and set Programme Code + Year '
        'Level - Group Label must exactly match the auto-generated form '
        '"<Programme Code>-Y<Year Level>" (e.g. DSC-Y2), the same way the Add Group page '
        'generates it.',
        'Sub-groups (e.g. DSC-Y2-A) can\'t be created through this import - use '
        '"Generate Sub-groups" on the Student Groups page instead, so split sizes and '
        'lettering stay consistent.',
        'If any row fails validation, nothing in this file is imported - fix the error and '
        're-upload the whole file.',
    ]:
        notes.append([line])
    notes.column_dimensions['A'].width = 100

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='student_groups.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/student-groups/import', methods=['POST'])
@login_required
def student_group_import():
    import pandas as pd

    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('admin.student_groups'))

    try:
        df = pd.read_excel(file, sheet_name='Student Groups', dtype=str).fillna('')
    except Exception as e:
        flash(f'Could not read file: {e}', 'danger')
        return redirect(url_for('admin.student_groups'))

    df.columns = df.columns.str.strip().str.lower()
    required_cols = {'group label', 'programme code', 'year level', 'intake size', 'parent group label'}
    missing = required_cols - set(df.columns)
    if missing:
        flash(f'Missing column(s): {", ".join(sorted(missing))}. Re-download the template and try again.', 'danger')
        return redirect(url_for('admin.student_groups'))

    existing_by_label = {g.group_label: g for g in StudentGroup.query.all()}
    programmes_by_code = {p.code: p for p in Programme.query.all()}

    rows = []
    errors = []
    seen_labels = {}

    for i, row in df.iterrows():
        excel_row = i + 2
        label = row.get('group label', '').strip().upper()
        prog_code = row.get('programme code', '').strip().upper()
        year_raw = row.get('year level', '').strip()
        intake_raw = row.get('intake size', '').strip()
        parent_label = row.get('parent group label', '').strip().upper()

        if not label and not prog_code and not year_raw and not intake_raw:
            continue  # fully blank row

        row_errors = []
        if not label:
            row_errors.append('Group Label is required')
        if label in seen_labels:
            row_errors.append(f'Group Label also used on row {seen_labels.get(label)}')
        elif label:
            seen_labels[label] = excel_row

        intake_size = None
        try:
            intake_size = int(intake_raw)
            if intake_size < 1:
                raise ValueError
        except ValueError:
            row_errors.append('Intake Size must be a whole number of 1 or more')

        existing_group = existing_by_label.get(label) if label else None

        if existing_group is None:
            # New group - top-level only, label must be the auto-generated form
            if parent_label:
                row_errors.append(
                    'Sub-groups can\'t be created via import - use "Generate Sub-groups" '
                    'on the Student Groups page for a new sub-group'
                )
            if not prog_code or prog_code not in programmes_by_code:
                row_errors.append(f'Programme Code "{prog_code}" was not found')
            year_level = None
            try:
                year_level = int(year_raw)
                if year_level < 1:
                    raise ValueError
            except ValueError:
                row_errors.append('Year Level must be a whole number of 1 or more')

            if prog_code in programmes_by_code and year_level:
                expected_label = f'{prog_code}-Y{year_level}'
                if label != expected_label:
                    row_errors.append(
                        f'Group Label should be "{expected_label}" for Programme {prog_code} '
                        f'Year {year_level} - labels are auto-generated, not freely chosen'
                    )

            if row_errors:
                errors.append(f'Row {excel_row} (Group Label "{label or "?"}"): ' + '; '.join(row_errors))
                continue

            rows.append({
                'excel_row': excel_row, 'is_new': True, 'label': label,
                'programme_id': programmes_by_code[prog_code].id, 'year_level': year_level,
                'intake_size': intake_size,
            })
        else:
            if row_errors:
                errors.append(f'Row {excel_row} (Group Label "{label}"): ' + '; '.join(row_errors))
                continue
            rows.append({
                'excel_row': excel_row, 'is_new': False, 'existing': existing_group,
                'intake_size': intake_size,
            })

    if errors:
        flash(f'Import rejected - {len(errors)} problem(s) found. Nothing was changed.', 'danger')
        for e in errors[:15]:
            flash(e, 'warning')
        if len(errors) > 15:
            flash(f'...and {len(errors) - 15} more. Fix these and re-upload the whole file.', 'warning')
        return redirect(url_for('admin.student_groups'))

    created = updated = 0
    for r in rows:
        if r['is_new']:
            db.session.add(StudentGroup(
                programme_id=r['programme_id'], year_level=r['year_level'],
                group_label=r['label'], intake_size=r['intake_size'], parent_id=None,
            ))
            created += 1
        else:
            r['existing'].intake_size = r['intake_size']
            updated += 1

    db.session.commit()
    flash(f'Import complete - {created} group(s) added, {updated} updated.', 'success')
    return redirect(url_for('admin.student_groups'))


# ---------------------------------------------------------------------------
# Session Assignment
# ---------------------------------------------------------------------------

@admin_bp.route('/courses/<int:course_id>/sessions')
@login_required
def course_sessions(course_id):
    course = Course.query.get_or_404(course_id)

    # Sync f2f session count to split_count (expand up or trim down).
    from collections import defaultdict
    target_count = course.split_count if course.split_count and course.split_count > 0 else 1
    f2f_sessions = [s for s in course.class_sessions if s.delivery_mode == 'f2f']
    type_counts = defaultdict(list)
    for s in f2f_sessions:
        type_counts[s.session_type].append(s)

    changed = False
    for stype, sessions in type_counts.items():
        # Expand: create sessions if below target
        while len(sessions) < target_count:
            new_session = ClassSession(
                course_id=course.id,
                session_type=stype,
                delivery_mode='f2f',
                duration_hours=sessions[0].duration_hours,
                student_group_id=None,
            )
            db.session.add(new_session)
            sessions.append(new_session)
            changed = True

        # Trim: remove excess sessions if above target.
        # Only removes sessions that are fully unassigned with no timetable entries.
        while len(sessions) > target_count:
            last = sessions[-1]
            if not last.all_professor_ids and not last.student_group_id and not last.timetable_entries:
                db.session.delete(last)
                sessions.pop()
                changed = True
            else:
                break  # Leave assigned sessions alone - admin must clear them manually

    if changed:
        db.session.flush()

    # Keep Template 2's "Group" column correct for every session type this
    # module has - not just the f2f ones the split-sync above touches -
    # since stale/garbage values from before this logic existed need fixing
    # too, and this keeps it correct every time this page loads.
    all_types = {s.session_type for s in
                 ClassSession.query.filter_by(course_id=course.id).all()}
    for stype in all_types:
        _recompute_group_labels(course.id, stype)
    db.session.commit()

    # Reload sessions after potential expansion
    sessions = (ClassSession.query
                .filter_by(course_id=course_id)
                .order_by(ClassSession.delivery_mode, ClassSession.session_type)
                .all())

    professors = Professor.query.join(Professor.user).order_by(User.name).all()

    # Student groups: top-level for the course's year, plus their sub-groups
    top_groups = StudentGroup.query.filter_by(
        year_level=course.year_level, parent_id=None
    ).order_by(StudentGroup.group_label).all()

    # Build a flat list: top-level + sub-groups indented for the dropdown
    group_choices = []
    for g in top_groups:
        group_choices.append((g.id, g.group_label, False))
        for sub in sorted(g.sub_groups, key=lambda x: x.group_label):
            group_choices.append((sub.id, f'  {sub.group_label}', True))

    # Build compatible timeslots per session for the fixed-slot dropdown
    # Only show timeslots that match the session's type and duration
    all_timeslots = (TimeSlot.query
                     .order_by(TimeSlot.day_of_week, TimeSlot.start_time)
                     .all())

    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    all_timeslots.sort(key=lambda ts: (
        day_order.index(ts.day_of_week), ts.start_time
    ))

    # _slot_ok is a module-level helper (defined below, near the manual
    # timetable-editing routes that also need it) - reused here so both
    # places agree on what counts as a valid fixed-slot pin.
    compat_ts = {
        s.id: [ts for ts in all_timeslots if _slot_ok(ts, s)]
        for s in sessions
    }

    return render_template('admin/course_sessions.html',
                           course=course,
                           sessions=sessions,
                           professors=professors,
                           group_choices=group_choices,
                           compat_ts=compat_ts,
                           session_types=SESSION_TYPES)


@admin_bp.route('/courses/<int:course_id>/sessions/add', methods=['POST'])
@login_required
def course_session_add(course_id):
    course = Course.query.get_or_404(course_id)

    session_type = request.form.get('session_type', '').strip().lower()
    session_delivery_mode = request.form.get('session_delivery_mode', '').strip().lower()
    duration_raw = request.form.get('duration_hours', '').strip()

    if session_type not in SESSION_TYPES:
        flash('Please select a valid session type.', 'danger')
        return redirect(url_for('admin.course_sessions', course_id=course.id))
    if session_delivery_mode not in SESSION_DELIVERY_MODES:
        flash('Please select a valid delivery mode.', 'danger')
        return redirect(url_for('admin.course_sessions', course_id=course.id))
    if not duration_raw.isdigit() or int(duration_raw) <= 0:
        flash('Duration (hours) must be a positive whole number.', 'danger')
        return redirect(url_for('admin.course_sessions', course_id=course.id))

    existing = ClassSession.query.filter_by(course_id=course.id, session_type=session_type).first()
    if existing:
        flash(f'{course.module_code} already has a {session_type} session.', 'warning')
        return redirect(url_for('admin.course_sessions', course_id=course.id))

    db.session.add(ClassSession(
        course_id=course.id, session_type=session_type,
        delivery_mode=session_delivery_mode, duration_hours=int(duration_raw),
    ))
    db.session.flush()
    _recompute_group_labels(course.id, session_type)
    db.session.commit()
    flash(f'{session_type.capitalize()} session added to {course.module_code}.', 'success')
    return redirect(url_for('admin.course_sessions', course_id=course.id))


@admin_bp.route('/courses/<int:course_id>/sessions/<int:session_id>/assign',
                methods=['POST'])
@login_required
def session_assign(course_id, session_id):
    """Assign a session's student group, fixed timeslot, and professor(s).
    Validated in full BEFORE any write - this data feeds the solver
    directly (fixed_timeslot_id becomes a hard pin), so a tampered or
    stale form value must be rejected with a clear message rather than
    crash, silently corrupt the session, or pin an incompatible slot
    (found in a 2026-07-16 input-error-prevention audit - previously
    every field here was cast with a bare int() and written straight to
    the DB with no existence or compatibility check at all)."""
    from app.models.class_session_professor import ClassSessionProfessor

    session = ClassSession.query.get_or_404(session_id)
    student_group_raw = request.form.get('student_group_id', '').strip()
    fixed_ts_raw       = request.form.get('fixed_timeslot_id', '').strip()
    primary_raw        = request.form.get('professor_id_primary', '').strip()
    co_raws            = [r.strip() for r in request.form.getlist('professor_id_co') if r.strip()]

    errors = []

    group_id, err = _parse_id(student_group_raw, 'Student group')
    if err:
        errors.append(err)
    elif group_id is not None and StudentGroup.query.get(group_id) is None:
        errors.append('That student group no longer exists - please pick another.')

    fixed_ts_id, err = _parse_id(fixed_ts_raw, 'Fixed timeslot')
    if err:
        errors.append(err)
    elif fixed_ts_id is not None:
        ts = TimeSlot.query.get(fixed_ts_id)
        if ts is None:
            errors.append('That fixed timeslot no longer exists - please pick another.')
        elif not _slot_ok(ts, session):
            errors.append(
                f'{ts.day_of_week} {ts.period_label} does not match this session\'s '
                f'duration/type - it cannot be used as a fixed slot.'
            )

    primary_id, err = _parse_id(primary_raw, 'Primary professor')
    if err:
        errors.append(err)
    elif primary_id is not None and Professor.query.get(primary_id) is None:
        errors.append('That primary professor no longer exists - please pick another.')

    co_ids = []
    for co_raw in co_raws:
        co_id, err = _parse_id(co_raw, 'Co-teacher')
        if err:
            errors.append(err)
            continue
        if Professor.query.get(co_id) is None:
            errors.append('One of the selected co-teachers no longer exists - please review.')
            continue
        co_ids.append(co_id)

    # Same professor can't be picked twice (as primary + co-teacher, or in
    # two co-teacher rows) - would create duplicate/contradictory
    # ClassSessionProfessor rows with no way to tell them apart.
    all_prof_ids = ([primary_id] if primary_id is not None else []) + co_ids
    if len(all_prof_ids) != len(set(all_prof_ids)):
        errors.append('The same professor is selected more than once for this session.')

    if errors:
        for e in errors:
            flash(e, 'danger')
        return redirect(url_for('admin.course_sessions', course_id=course_id))

    session.student_group_id  = group_id
    session.fixed_timeslot_id = fixed_ts_id

    ClassSessionProfessor.query.filter_by(session_id=session.id).delete()

    order = 0
    if primary_id is not None:
        db.session.add(ClassSessionProfessor(
            session_id=session.id, professor_id=primary_id,
            is_primary=True, display_order=0
        ))
        order = 1
    for co_id in co_ids:
        db.session.add(ClassSessionProfessor(
            session_id=session.id, professor_id=co_id,
            is_primary=False, display_order=order
        ))
        order += 1

    db.session.commit()
    flash('Session updated.', 'success')
    return redirect(url_for('admin.course_sessions', course_id=course_id))


# ---------------------------------------------------------------------------
# Availability Declarations - admin classification
# ---------------------------------------------------------------------------

@admin_bp.route('/declarations', methods=['GET', 'POST'])
@login_required
def declarations():
    if request.method == 'POST':
        decl_id         = request.form.get('decl_id', '').strip()
        constraint_type = request.form.get('constraint_type', '').strip()

        decl_id_num = None
        if decl_id:
            try:
                decl_id_num = int(decl_id)
            except ValueError:
                decl_id_num = None

        if decl_id_num and constraint_type in ('strict', 'preferred'):
            decl = AvailabilityDeclaration.query.get_or_404(decl_id_num)
            decl.constraint_type = constraint_type
            decl.status          = 'classified'
            db.session.commit()
            flash(
                f'{decl.professor.user.name} - {decl.timeslot.day_of_week} '
                f'{decl.timeslot.period_label} classified as {constraint_type.capitalize()}.',
                'success'
            )
        else:
            flash('Invalid classification request.', 'danger')

        return redirect(url_for('admin.declarations'))

    pending    = (AvailabilityDeclaration.query
                  .filter_by(status='pending')
                  .order_by(AvailabilityDeclaration.submitted_at)
                  .all())
    classified = (AvailabilityDeclaration.query
                  .filter_by(status='classified')
                  .order_by(AvailabilityDeclaration.submitted_at.desc())
                  .all())

    return render_template('admin/declarations.html',
                           pending=pending,
                           classified=classified)


# ---------------------------------------------------------------------------
# Manual timetable editing - helpers
# ---------------------------------------------------------------------------

DAY_ORDER = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']


def _slot_ok(ts, s):
    """Return True if timeslot ts is compatible with session s."""
    start_m = ts.start_time.hour * 60 + ts.start_time.minute
    end_m   = ts.end_time.hour   * 60 + ts.end_time.minute
    if (end_m - start_m) // 60 != s.duration_hours:
        return False
    return ts.period_label.startswith('Lab') == (s.session_type == 'lab')


def _parse_id(raw, label):
    """Parse an optional numeric form field into (int_id, error_message).
    error_message is None on success; int_id is None for a blank field.
    Shared by every admin form route that accepts a dropdown-selected FK id,
    since a bare int() crashes on a tampered/stale non-numeric value instead
    of showing a clear error (input-error-prevention audit, 2026-07-16)."""
    if not raw:
        return None, None
    try:
        n = int(raw)
    except ValueError:
        return None, f'{label} was not a valid selection - please choose from the dropdown.'
    return n, None


def _check_week_conflicts(week_number, trimester, timeslot_id,
                          room_id, professor_id, student_group_id,
                          exclude_entry_id=None):
    """
    Check if placing a session at (week_number, timeslot_id) causes double-bookings.
    Returns a list of human-readable conflict strings. Empty list = no conflicts.
    """
    from app.engine.solver import _conflicting_group_ids
    conflicts = []

    q = (TimetableEntry.query
         .filter_by(trimester=trimester,
                    week_number=week_number,
                    timeslot_id=timeslot_id))
    if exclude_entry_id:
        q = q.filter(TimetableEntry.id != exclude_entry_id)
    competing = q.all()

    for other in competing:
        other_prof = other.effective_professor
        other_prof_id   = other_prof.id if other_prof else None
        other_group_id  = other.class_session.student_group_id

        if professor_id and other_prof_id == professor_id:
            conflicts.append(
                f'Professor double-booking: already assigned to '
                f'{other.class_session.course.module_code} '
                f'({other.class_session.session_type}) '
                f'at this slot in Week {week_number}.'
            )

        if room_id and other.room_id and other.room_id == room_id:
            conflicts.append(
                f'Room double-booking: {other.room.room_code} is already used by '
                f'{other.class_session.course.module_code} '
                f'at this slot in Week {week_number}.'
            )

        if student_group_id and other_group_id:
            if other_group_id in _conflicting_group_ids(student_group_id):
                other_label = other.class_session.student_group.group_label
                conflicts.append(
                    f'Student group double-booking: {other_label} is already in '
                    f'{other.class_session.course.module_code} '
                    f'at this slot in Week {week_number}.'
                )

    return conflicts


def _write_audit(user_id, trimester, action,
                 module_code, session_type, week_label,
                 old_ts, new_ts, old_room, new_room,
                 old_prof, new_prof):
    """Write one AuditLog entry."""
    db.session.add(AuditLog(
        user_id      = user_id,
        trimester    = trimester,
        action       = action,
        module_code  = module_code,
        session_type = session_type,
        week_label   = week_label,
        old_timeslot = old_ts,
        new_timeslot = new_ts,
        old_room     = old_room,
        new_room     = new_room,
        old_professor= old_prof,
        new_professor= new_prof,
    ))


# ---------------------------------------------------------------------------
# Manual timetable editing - routes
# ---------------------------------------------------------------------------

@admin_bp.route('/timetable/<trimester>/sessions/<int:session_id>/weeks')
@login_required
def timetable_session_weeks(trimester, session_id):
    session = ClassSession.query.get_or_404(session_id)

    entries = (TimetableEntry.query
               .filter_by(class_session_id=session_id, trimester=trimester)
               .order_by(TimetableEntry.week_number)
               .all())

    from app.models.academic_calendar import AcademicCalendar
    cal = {w.week_number: w for w in
           AcademicCalendar.query.filter_by(trimester=trimester).all()}

    return render_template('admin/timetable_session_weeks.html',
                           session=session,
                           entries=entries,
                           trimester=trimester,
                           cal=cal)


@admin_bp.route('/timetable/entries/<int:entry_id>/edit', methods=['GET', 'POST'])
@login_required
def timetable_edit_entry(entry_id):
    """Edit a single week's timetable entry."""
    entry   = TimetableEntry.query.get_or_404(entry_id)
    session = entry.class_session

    all_ts = TimeSlot.query.all()
    all_ts.sort(key=lambda ts: (DAY_ORDER.index(ts.day_of_week), ts.start_time))
    compat_slots = [ts for ts in all_ts if _slot_ok(ts, session)]

    all_rooms   = Room.query.filter_by(is_active=True).order_by(Room.room_code).all()
    professors  = Professor.query.join(Professor.user).order_by(User.name).all()

    conflicts   = []
    force_save  = False

    if request.method == 'POST':
        new_ts_id_raw   = request.form.get('timeslot_id', '').strip()
        new_room_id_raw = request.form.get('room_id', '').strip()
        new_prof_id_raw = request.form.get('professor_id', '').strip()
        force_save      = request.form.get('force_save') == '1'

        errors = []
        if not new_ts_id_raw:
            errors.append('Please select a timeslot.')

        new_ts_id, err = _parse_id(new_ts_id_raw, 'Timeslot')
        if err:
            errors.append(err)
        new_room_id, err = _parse_id(new_room_id_raw, 'Room')
        if err:
            errors.append(err)
        new_prof_id, err = _parse_id(new_prof_id_raw, 'Professor')
        if err:
            errors.append(err)

        if not errors and new_ts_id is not None:
            new_ts = TimeSlot.query.get(new_ts_id)
            if new_ts is None:
                errors.append('That timeslot no longer exists - please pick another.')
            elif not _slot_ok(new_ts, session):
                errors.append(
                    f'{new_ts.day_of_week} {new_ts.period_label} does not match this '
                    f'session\'s duration/type - it cannot be used here.'
                )
            if new_room_id is not None and Room.query.get(new_room_id) is None:
                errors.append('That room no longer exists - please pick another.')
            if new_prof_id is not None and Professor.query.get(new_prof_id) is None:
                errors.append('That professor no longer exists - please pick another.')

        if not errors:
            if not force_save:
                conflicts = _check_week_conflicts(
                    week_number      = entry.week_number,
                    trimester        = entry.trimester,
                    timeslot_id      = new_ts_id,
                    room_id          = new_room_id,
                    professor_id     = new_prof_id,
                    student_group_id = session.student_group_id,
                    exclude_entry_id = entry.id,
                )

            if not conflicts or force_save:
                # Build audit values
                old_ts_obj  = entry.timeslot
                new_ts_obj  = TimeSlot.query.get(new_ts_id)
                old_room_obj = entry.room
                new_room_obj = Room.query.get(new_room_id) if new_room_id else None
                old_prof_obj = entry.effective_professor
                new_prof_obj = Professor.query.get(new_prof_id) if new_prof_id else None

                def _ts_label(ts):
                    return f'{ts.day_of_week} {ts.period_label} ({ts.start_time.strftime("%H:%M")}–{ts.end_time.strftime("%H:%M")})' if ts else '-'

                _write_audit(
                    user_id      = current_user.id,
                    trimester    = entry.trimester,
                    action       = 'edit_week',
                    module_code  = session.course.module_code,
                    session_type = session.session_type,
                    week_label   = f'Week {entry.week_number}',
                    old_ts       = _ts_label(old_ts_obj),
                    new_ts       = _ts_label(new_ts_obj),
                    old_room     = old_room_obj.room_code if old_room_obj else 'Online',
                    new_room     = new_room_obj.room_code if new_room_obj else 'Online',
                    old_prof     = old_prof_obj.user.name if old_prof_obj else '-',
                    new_prof     = new_prof_obj.user.name if new_prof_obj else '-',
                )

                entry.timeslot_id           = new_ts_id
                entry.room_id               = new_room_id
                entry.override_professor_id = new_prof_id if new_prof_id != session.primary_professor_id else None
                entry.is_manually_edited    = True
                db.session.commit()

                flash(f'Week {entry.week_number} updated successfully.', 'success')
                return redirect(url_for('admin.timetable_session_weeks',
                                        trimester=entry.trimester,
                                        session_id=session.id))
        else:
            for e in errors:
                flash(e, 'danger')

    return render_template('admin/timetable_edit_entry.html',
                           entry=entry,
                           session=session,
                           compat_slots=compat_slots,
                           all_rooms=all_rooms,
                           professors=professors,
                           conflicts=conflicts,
                           edit_mode='single')


@admin_bp.route('/timetable/<trimester>/sessions/<int:session_id>/edit-all',
                methods=['GET', 'POST'])
@login_required
def timetable_edit_all_weeks(trimester, session_id):
    """Edit all weeks for a session at once."""
    session = ClassSession.query.get_or_404(session_id)

    entries = (TimetableEntry.query
               .filter_by(class_session_id=session_id, trimester=trimester)
               .all())

    if not entries:
        flash('No timetable entries found for this session.', 'danger')
        return redirect(url_for('admin.timetable', trimester=trimester))

    all_ts = TimeSlot.query.all()
    all_ts.sort(key=lambda ts: (DAY_ORDER.index(ts.day_of_week), ts.start_time))
    compat_slots = [ts for ts in all_ts if _slot_ok(ts, session)]

    all_rooms  = Room.query.filter_by(is_active=True).order_by(Room.room_code).all()
    professors = Professor.query.join(Professor.user).order_by(User.name).all()

    # Use first entry as the representative current values
    rep         = entries[0]
    conflicts   = []
    force_save  = False

    if request.method == 'POST':
        new_ts_id_raw   = request.form.get('timeslot_id', '').strip()
        new_room_id_raw = request.form.get('room_id', '').strip()
        new_prof_id_raw = request.form.get('professor_id', '').strip()
        force_save      = request.form.get('force_save') == '1'

        errors = []
        if not new_ts_id_raw:
            errors.append('Please select a timeslot.')

        new_ts_id, err = _parse_id(new_ts_id_raw, 'Timeslot')
        if err:
            errors.append(err)
        new_room_id, err = _parse_id(new_room_id_raw, 'Room')
        if err:
            errors.append(err)
        new_prof_id, err = _parse_id(new_prof_id_raw, 'Professor')
        if err:
            errors.append(err)

        if not errors and new_ts_id is not None:
            new_ts = TimeSlot.query.get(new_ts_id)
            if new_ts is None:
                errors.append('That timeslot no longer exists - please pick another.')
            elif not _slot_ok(new_ts, session):
                errors.append(
                    f'{new_ts.day_of_week} {new_ts.period_label} does not match this '
                    f'session\'s duration/type - it cannot be used here.'
                )
            if new_room_id is not None and Room.query.get(new_room_id) is None:
                errors.append('That room no longer exists - please pick another.')
            if new_prof_id is not None and Professor.query.get(new_prof_id) is None:
                errors.append('That professor no longer exists - please pick another.')

        if not errors:
            if not force_save:
                for e in entries:
                    week_conflicts = _check_week_conflicts(
                        week_number      = e.week_number,
                        trimester        = trimester,
                        timeslot_id      = new_ts_id,
                        room_id          = new_room_id,
                        professor_id     = new_prof_id,
                        student_group_id = session.student_group_id,
                        exclude_entry_id = e.id,
                    )
                    conflicts.extend(week_conflicts)

            if not conflicts or force_save:
                old_ts_obj   = rep.timeslot
                new_ts_obj   = TimeSlot.query.get(new_ts_id)
                old_room_obj = rep.room
                new_room_obj = Room.query.get(new_room_id) if new_room_id else None
                old_prof_obj = rep.effective_professor
                new_prof_obj = Professor.query.get(new_prof_id) if new_prof_id else None

                def _ts_label(ts):
                    return f'{ts.day_of_week} {ts.period_label} ({ts.start_time.strftime("%H:%M")}–{ts.end_time.strftime("%H:%M")})' if ts else '-'

                _write_audit(
                    user_id      = current_user.id,
                    trimester    = trimester,
                    action       = 'edit_all_weeks',
                    module_code  = session.course.module_code,
                    session_type = session.session_type,
                    week_label   = 'All weeks',
                    old_ts       = _ts_label(old_ts_obj),
                    new_ts       = _ts_label(new_ts_obj),
                    old_room     = old_room_obj.room_code if old_room_obj else 'Online',
                    new_room     = new_room_obj.room_code if new_room_obj else 'Online',
                    old_prof     = old_prof_obj.user.name if old_prof_obj else '-',
                    new_prof     = new_prof_obj.user.name if new_prof_obj else '-',
                )

                for e in entries:
                    e.timeslot_id           = new_ts_id
                    e.room_id               = new_room_id
                    e.override_professor_id = new_prof_id if new_prof_id != session.primary_professor_id else None
                    e.is_manually_edited    = True

                db.session.commit()
                flash(f'All weeks for {session.course.module_code} updated successfully.', 'success')
                return redirect(url_for('admin.timetable_session_weeks',
                                        trimester=trimester,
                                        session_id=session_id))
        else:
            for e in errors:
                flash(e, 'danger')

    return render_template('admin/timetable_edit_entry.html',
                           entry=rep,
                           session=session,
                           compat_slots=compat_slots,
                           all_rooms=all_rooms,
                           professors=professors,
                           conflicts=conflicts,
                           edit_mode='all',
                           trimester=trimester)


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Historical slot preference helper
# ---------------------------------------------------------------------------

def _build_historical_preferred(academic_year, trimester_num):
    """
    Build {class_session_id: timeslot_id} for soft-constraint guidance.

    Priority:
      1. Backbone entries for this AY+trimester (when regenerating an existing AY).
      2. All entries from the previous AY's equivalent trimester (normal case).
    """
    if not academic_year or len(academic_year) < 6:
        return {}

    tri_key = f'{academic_year}-T{trimester_num}'

    # Priority 1: backbone entries in this AY tell us the "real" schedule to target
    backbone = (TimetableEntry.query
                .filter_by(trimester=tri_key, is_backbone=True)
                .all())
    if backbone:
        preferred = {}
        for e in backbone:
            if e.class_session_id not in preferred:
                preferred[e.class_session_id] = e.timeslot_id
        return preferred

    # Priority 2: previous AY's equivalent trimester
    try:
        y1 = int(academic_year[2:4])
        y2 = int(academic_year[4:6])
        prev_ay = f'AY{y1 - 1:02d}{y2 - 1:02d}'
    except ValueError:
        return {}

    prev_tri_key = f'{prev_ay}-T{trimester_num}'
    prev_entries = (TimetableEntry.query
                    .filter_by(trimester=prev_tri_key)
                    .all())
    preferred = {}
    for e in prev_entries:
        if e.class_session_id not in preferred:
            preferred[e.class_session_id] = e.timeslot_id
    return preferred


# ---------------------------------------------------------------------------
# Timetable Flags
# ---------------------------------------------------------------------------

def _auto_create_flags(trimester, preferred_violations):
    """
    After solver run, create TimetableFlag records for each preferred violation.
    Skips violations that already have an open/acknowledged flag to avoid duplicates.
    Clears resolved flags for this trimester first (fresh run).
    """
    from app.models.timetable_flag import TimetableFlag

    # Delete any previously open flags for this trimester (solver re-run resets them)
    TimetableFlag.query.filter(
        TimetableFlag.status.in_(['open']),
        TimetableFlag.timetable_entry.has(trimester=trimester)
    ).delete(synchronize_session='fetch')

    for v in preferred_violations:
        if not v.get('declaration_id'):
            continue
        rep_entry = TimetableEntry.query.filter_by(
            class_session_id=v['class_session_id'],
            trimester=trimester,
        ).first()
        if rep_entry:
            db.session.add(TimetableFlag(
                timetable_entry_id = rep_entry.id,
                professor_id       = v['professor_id'],
                declaration_id     = v['declaration_id'],
                status             = 'open',
            ))
    db.session.commit()


@admin_bp.route('/timetable-flags')
@login_required
def timetable_flags():
    from app.models.timetable_flag import TimetableFlag

    trimester_filter = request.args.get('trimester', '')

    trimesters = [r[0] for r in
                  db.session.query(TimetableEntry.trimester).distinct()
                  .order_by(TimetableEntry.trimester).all()]

    # Default to most recent trimester
    if not trimester_filter and trimesters:
        trimester_filter = trimesters[-1]

    flags_query = (TimetableFlag.query
                   .join(TimetableFlag.timetable_entry)
                   .filter(TimetableEntry.trimester == trimester_filter)
                   .order_by(TimetableFlag.created_at.desc()))

    open_flags         = flags_query.filter(TimetableFlag.status == 'open').all()
    acknowledged_flags = flags_query.filter(TimetableFlag.status == 'acknowledged').all()
    resolved_flags     = flags_query.filter(TimetableFlag.status == 'resolved').all()

    return render_template('admin/timetable_flags.html',
                           open_flags=open_flags,
                           acknowledged_flags=acknowledged_flags,
                           resolved_flags=resolved_flags,
                           trimesters=trimesters,
                           active_trimester=trimester_filter)


@admin_bp.route('/timetable-flags/<int:flag_id>/notify', methods=['POST'])
@login_required
def flag_notify(flag_id):
    """Admin sets response deadline and marks notification as sent (email wired in Step 8)."""
    from app.models.timetable_flag import TimetableFlag
    from datetime import date as date_type

    flag     = TimetableFlag.query.get_or_404(flag_id)
    deadline = request.form.get('response_deadline', '').strip()

    if not deadline:
        flash('Please set a response deadline before sending.', 'danger')
        return redirect(url_for('admin.timetable_flags',
                                trimester=flag.timetable_entry.trimester))

    try:
        parsed_deadline = date_type.fromisoformat(deadline)
    except ValueError:
        flash('Response deadline was not a valid date - please use the date picker.', 'danger')
        return redirect(url_for('admin.timetable_flags',
                                trimester=flag.timetable_entry.trimester))

    flag.response_deadline = parsed_deadline
    db.session.commit()

    # Send email to professor
    from app.utils.email import send_flag_notification
    success, error = send_flag_notification(flag)

    if success:
        flag.notification_sent = True
        db.session.commit()
        flash(
            f'Email sent to {flag.professor.user.name} '
            f'({flag.professor.user.email}). '
            f'Deadline: {flag.response_deadline.strftime("%d %b %Y")}.',
            'success'
        )
    else:
        flash(
            f'Deadline saved but email could not be sent to '
            f'{flag.professor.user.name}. '
            f'SMTP error: {error}. '
            f'Please check your email configuration in config.py.',
            'danger'
        )

    return redirect(url_for('admin.timetable_flags',
                            trimester=flag.timetable_entry.trimester))


# ---------------------------------------------------------------------------
# Audit Log
# ---------------------------------------------------------------------------

@admin_bp.route('/audit-log')
@login_required
def audit_log():
    trimester_filter = request.args.get('trimester', '')

    trimesters = [r[0] for r in
                  db.session.query(AuditLog.trimester).distinct()
                  .order_by(AuditLog.trimester).all()]

    query = AuditLog.query.order_by(AuditLog.timestamp.desc())
    if trimester_filter:
        query = query.filter_by(trimester=trimester_filter)

    logs = query.all()

    return render_template('admin/audit_log.html',
                           logs=logs,
                           trimesters=trimesters,
                           active_trimester=trimester_filter)


@admin_bp.route('/audit-log/export')
@login_required
def audit_log_export():
    """Download audit log as CSV."""
    import csv
    import io
    from flask import Response

    trimester_filter = request.args.get('trimester', '')
    query = AuditLog.query.order_by(AuditLog.timestamp.desc())
    if trimester_filter:
        query = query.filter_by(trimester=trimester_filter)
    logs = query.all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Timestamp', 'Admin', 'Trimester', 'Module', 'Session Type',
                     'Scope', 'Old Timeslot', 'New Timeslot',
                     'Old Room', 'New Room', 'Old Professor', 'New Professor'])
    for log in logs:
        writer.writerow([
            log.timestamp.strftime('%Y-%m-%d %H:%M'),
            log.user.name if log.user else '',
            log.trimester or '',
            log.module_code or '',
            log.session_type or '',
            'All weeks' if log.action == 'edit_all_weeks' else (log.week_label or ''),
            log.old_timeslot or '', log.new_timeslot or '',
            log.old_room or '',    log.new_room or '',
            log.old_professor or '', log.new_professor or '',
        ])

    filename = f'audit_log{"_" + trimester_filter if trimester_filter else ""}.csv'
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ---------------------------------------------------------------------------
# System Info - constraints reference + assumed/self-input data disclosure
# ---------------------------------------------------------------------------

def _build_constraints_reference(sv, sv_shared_group_count):
    """The single source of truth for every hard/soft constraint's id, title,
    and description. Used by both system_info() (static reference) and
    timetable()'s post-generation summary (same text, live numbers attached)
    so the two pages can never describe a rule differently."""

    def fmt(t):
        return t.strftime('%H:%M')

    constraints = {
        'hard': [
            {'category': 'Conflict Prevention', 'icon': 'bi-shield-check', 'color': 'red', 'rows': [
                {'id': 'H1', 'title': 'No two classes in the same room at the same time',
                 'status': 'Implemented', 'value': 'Checked room by room, only for weeks the classes actually overlap'},
                {'id': 'H2', 'title': 'A professor cannot teach two classes at the same time',
                 'status': 'Implemented', 'value': 'Checked professor by professor, only for weeks the classes actually overlap'},
                {'id': 'H3', 'title': 'A student group cannot attend two classes at the same time',
                 'status': 'Implemented', 'value': 'Also checks for clashes between a group and its sub-groups'},
                {'id': 'H4', 'title': 'The room must be big enough for the class',
                 'status': 'Implemented', 'value': ''},
                {'id': 'H17', 'title': 'Modules shared across programmes are scheduled as one combined class',
                 'status': 'Implemented', 'value': f'{sv_shared_group_count} module(s) linked - same time and room, room size checked against everyone combined'},
            ]},
            {'category': 'Room & Delivery Mode', 'icon': 'bi-door-open', 'color': 'blue', 'rows': [
                {'id': 'H5', 'title': 'Online classes are never assigned a physical room',
                 'status': 'Implemented', 'value': 'Online classes simply have no room attached'},
                {'id': 'H6', 'title': 'In-person classes always get a real room, never left virtual',
                 'status': 'Implemented', 'value': 'Every in-person class must be matched to a room of the right type and size'},
            ]},
            {'category': 'Calendar & Teaching Weeks', 'icon': 'bi-calendar3', 'color': 'purple', 'rows': [
                {'id': 'H7', 'title': 'Odd/even week patterns (e.g. a class held only on odd weeks)',
                 'status': 'Implemented via data', 'value': 'Each class has its own list of which weeks it runs in, taken directly from the '
                            'uploaded file and followed exactly'},
                {'id': 'H8', 'title': 'No classes on public holidays or term breaks',
                 'status': 'Implemented', 'value': f'Uses the Singapore public holiday calendar, plus the '
                            f'term-break week (default: week {sorted(sv.DEFAULT_TERM_BREAK_WEEKS)})'},
                {'id': 'H16', 'title': 'University-wide modules: the two weekly classes must be on different days',
                 'status': 'Implemented', 'value': f'Applies only to {", ".join(sorted(sv.UNIWIDE_DAY_SEPARATED_MODULES))} '
                            '- the only modules the requirements doc names for this rule'},
            ]},
            {'category': 'Daily Time Windows', 'icon': 'bi-clock-history', 'color': 'green', 'rows': [
                {'id': 'H9', 'title': 'No classes before 9:00am',
                 'status': 'Built into the schedule', 'value': 'There simply are no time slots before 9:00am to choose from'},
                {'id': 'H10', 'title': 'No classes after 6:00pm',
                 'status': 'Implemented', 'value': f'No unlocked class may start at or after {fmt(sv.EVENING_CUTOFF)} - a handful of '
                            'real, source-confirmed evening classes (DSC-shared modules with other programmes) are the only exception, '
                            'locked to their exact real time'},
                {'id': 'H11', 'title': 'Wednesday afternoons are blocked',
                 'status': 'Implemented', 'value': f'No class may start at or after {fmt(sv.WED_AFTERNOON_CUTOFF)} on Wednesday'},
                {'id': 'H12', 'title': 'Friday 12:00pm-2:00pm is a protected window',
                 'status': 'Implemented', 'value': f'No class may start between {fmt(sv.FRI_BLOCK_START)} and {fmt(sv.FRI_BLOCK_END)} on Friday'},
                {'id': 'H13', 'title': 'Lunch break - flexible, not a fixed time',
                 'status': 'Implemented', 'value': f'Every student group gets at least 1 fully free hour somewhere between '
                            f'{fmt(sv.LUNCH_WINDOW_START)} and {fmt(sv.LUNCH_WINDOW_END)} each day - not one fixed hour for everyone'},
                {'id': 'H14', 'title': 'No classes on Saturday',
                 'status': 'Implemented', 'value': 'The system only knows about Monday to Friday time slots'},
                {'id': 'H15', 'title': 'No Friday classes after 5:00pm',
                 'status': 'Built into the schedule', 'value': 'Comes from the same time-slot list as the other Friday rules above'},
            ]},
        ],
        'soft': [
            {'category': 'Availability & Continuity', 'icon': 'bi-person-check', 'color': 'blue', 'rows': [
                {'id': 'S-avail', 'title': "Respect a professor's declared availability",
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_AVAILABILITY} - the highest of any soft rule'},
                {'id': 'S-pref-ts', 'title': "A professor's preferred time slot (read from their notes)",
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_PREFERRED_TS}'},
                {'id': 'S-hist', 'title': "Keep the same slot as last year's equivalent term, where possible",
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_HISTORICAL}'},
            ]},
            {'category': 'Workload & Comfort', 'icon': 'bi-battery-half', 'color': 'amber', 'rows': [
                {'id': 'S1', 'title': 'Avoid switching a professor between online and in-person back-to-back',
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_MODE_SWITCH_PROF}'},
                {'id': 'S2', 'title': 'Avoid switching a student group between online and in-person back-to-back',
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_MODE_SWITCH_GROUP}'},
                {'id': 'S3', 'title': "Avoid leaving a professor with more than a 2-hour gap in the same day",
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_PROF_IDLE_GAP} - triggers past a {sv.PROF_IDLE_GAP_THRESHOLD_HOURS}-hour gap'},
                {'id': 'S8', 'title': f'Avoid a student group having {sv.GROUP_BACKTOBACK_LIMIT_HOURS + 1}+ hours of classes back-to-back with no break',
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_GROUP_BACKTOBACK_HOURS} - the longer the unbroken run, the bigger the penalty'},
                {'id': 'S4', 'title': 'Avoid a student group having more than 4 consecutive teaching hours',
                 'status': 'Covered by S8 above', 'value': 'S8\'s stricter rule already catches every case this one would - '
                            'nothing extra needed'},
            ]},
            {'category': 'Scheduling Quality', 'icon': 'bi-stars', 'color': 'purple', 'rows': [
                {'id': 'S5', 'title': "Keep a student group's classes clustered, not spread across extra days",
                 'status': 'Implemented (groups only)', 'value': f'Priority {sv.WEIGHT_DAY_CLUSTER} - not yet applied to professors\' schedules'},
                {'id': 'S6', 'title': 'Prefer a room that\'s at least 60% full', 'status': 'Implemented',
                 'value': f'Priority {sv.WEIGHT_ROOM_UTIL} - target {int(sv.ROOM_UTIL_THRESHOLD * 100)}% full'},
                {'id': 'S7', 'title': "Avoid a group's very first or very last slot of the day",
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_EXTREMAL_SLOT}'},
                {'id': 'S9', 'title': 'Prefer classes to end by 5:00pm', 'status': 'Implemented',
                 'value': f'Priority {sv.WEIGHT_LATE_END} - target cutoff {fmt(sv.LATE_END_CUTOFF)}'},
                {'id': 'S10', 'title': 'Prefer a snugly-sized room, not just any room that fits',
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_ROOM_BEST_FIT} - a single very mismatched room doesn\'t '
                            'overwhelm the rest of the schedule\'s quality'},
                {'id': 'S11', 'title': 'Keep the same room for back-to-back classes (same professor or group)',
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_CONSISTENT_VENUE}'},
            ]},
            {'category': 'Pedagogical Ordering', 'icon': 'bi-mortarboard', 'color': 'green', 'rows': [
                {'id': 'S-lec-tut', 'title': "A module's lecture comes before its tutorial",
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_LEC_TUT_ORDER}'},
                {'id': 'S-lec-lab', 'title': "A module's lecture comes before its lab",
                 'status': 'Implemented', 'value': f'Priority {sv.WEIGHT_LEC_LAB_ORDER}'},
            ]},
        ],
    }
    return constraints


# Maps each soft constraint's id (from _build_constraints_reference) to a
# function computing (result text, violation count) from a solve() stats
# dict. S4 is intentionally absent - its card says "covered by S8 above",
# so it has no independent number to show.
_SOFT_STAT_MAP = {
    'S-avail':   lambda s: (f"{len(s.get('preferred_violations', []))} violated",
                             len(s.get('preferred_violations', []))),
    'S-pref-ts': lambda s: (f"{s.get('preferred_ts_honoured', 0)} honoured", 0),
    'S-hist':    lambda s: (f"{s.get('historical_honoured', 0)} kept, {s.get('historical_changed', 0)} changed",
                             s.get('historical_changed', 0)),
    'S1':        lambda s: (f"{s.get('mode_switches_prof', 0)} violated", s.get('mode_switches_prof', 0)),
    'S2':        lambda s: (f"{s.get('mode_switches_group', 0)} violated", s.get('mode_switches_group', 0)),
    'S3':        lambda s: (f"{s.get('prof_idle_gap_violations', 0)} violated", s.get('prof_idle_gap_violations', 0)),
    'S8':        lambda s: (f"{s.get('group_backtoback_violations', 0)} violated", s.get('group_backtoback_violations', 0)),
    'S5':        lambda s: (f"{s.get('day_spread_pairs', 0)} day-spread instances (informational, not a violation count)", 0),
    'S6':        lambda s: (f"{s.get('room_util_violations', 0)} violated", s.get('room_util_violations', 0)),
    'S7':        lambda s: (f"{s.get('extremal_slot_violations', 0)} violated", s.get('extremal_slot_violations', 0)),
    'S9':        lambda s: (f"{s.get('late_end_violations', 0)} violated", s.get('late_end_violations', 0)),
    'S10':       lambda s: (f"{s.get('room_best_fit_wasted_seats', 0)} wasted seats total (informational, not a violation count)", 0),
    'S11':       lambda s: (f"{s.get('consistent_venue_violations', 0)} violated", s.get('consistent_venue_violations', 0)),
    'S-lec-tut': lambda s: (f"{s.get('lec_tut_ordered', 0)} correctly ordered", 0),
    'S-lec-lab': lambda s: (f"{s.get('lec_lab_ordered', 0)} correctly ordered", 0),
}

# Rules whose "result" is informational (a count with no pass/fail meaning,
# e.g. day-spread instances or wasted seats) rather than a true violation -
# these never cost score points even though they have a number attached.
_SOFT_INFORMATIONAL_IDS = {'S5', 'S10', 'S-pref-ts', 'S-lec-tut', 'S-lec-lab', 'S-hist'}
# S-hist ("keep the same slot as last year") is backend-only machinery -
# it exists purely to bias the solver toward DSC's real backbone timing
# (see feedback_dsc_backbone.md), and only ever has a real target for a
# handful of backbone-linked sessions (everyone else has no "last year" to
# compare against at all). Counting it toward the score made it look like a
# general scheduling-quality problem when it isn't one - Brian: "remember
# backbone is backend? why is it in there" - made informational 2026-07-16
# so it no longer drags the score down or appears as a "top driver".

# Optimised Score calibration (see timetable_report()'s score calculation).
# Each violation is weighted by the rule's own solver priority (not a flat
# points-per-violation) - "ends late" (weight 15) costs more than "not the
# snuggest room fit" (weight 2). SCORE_CEILING_AVG_PENALTY is the average
# weighted-penalty-per-session at which the score reaches 0 - chosen from
# real observed data across all 3 trimesters (2026-07-12: T1 averaged 7.15,
# T2 4.24, T3 1.87 per session), so a ceiling of 15 gives genuine spread
# instead of every trimester this size instantly flooring to 0.
SCORE_CEILING_AVG_PENALTY = 15


def _build_constraint_summary(stats):
    """Post-generation breakdown: how many of each constraint were actually
    applied/violated in the given solve() stats, reusing the exact same
    rule text as System Info so the two pages never disagree. Also powers
    the Scoring Matrix on the Scheduling Report page (hard_groups/soft
    rows carry live weight + points-impact for full transparency)."""
    if not stats:
        return None
    from app.engine import solver as sv
    from app.models.shared_module_group import SharedModuleGroup

    ref = _build_constraints_reference(sv, SharedModuleGroup.query.count())
    live_weights = sv.get_effective_soft_weights()

    soft_groups = []
    total_violations = 0
    total_points = 0
    soft_rule_count = 0
    for group in ref['soft']:
        rows = []
        for r in group['rows']:
            fn = _SOFT_STAT_MAP.get(r['id'])
            if fn is None:
                continue  # S4 - no independent number, covered by S8
            soft_rule_count += 1
            result_text, violated = fn(stats)
            informational = r['id'] in _SOFT_INFORMATIONAL_IDS
            if not informational:
                total_violations += violated
            weight = live_weights.get(r['id']) or 0
            points = 0 if informational else violated * weight
            total_points += points
            rows.append({
                'id': r['id'], 'title': r['title'], 'result': result_text, 'violated': violated,
                'weight': live_weights.get(r['id']),
                'informational': informational,
                'points': points,
            })
        if rows:
            soft_groups.append({'category': group['category'], 'icon': group['icon'],
                                 'color': group['color'], 'rows': rows})

    hard_rule_count = sum(len(g['rows']) for g in ref['hard'])
    hard_applied = (stats.get('strict_constraints_applied', 0)
                     + stats.get('pins_applied', 0)
                     + stats.get('room_pins_applied', 0))
    hard_dropped = stats.get('pins_dropped', 0) + stats.get('room_pins_dropped', 0)

    return {
        'hard_rule_count': hard_rule_count,
        'hard_applied': hard_applied,
        'hard_dropped': hard_dropped,
        'hard_groups': ref['hard'],
        'soft_rule_count': soft_rule_count,
        'soft_total_violations': total_violations,
        'soft_total_points': total_points,
        'soft_groups': soft_groups,
    }


@admin_bp.route('/system-info')
@login_required
def system_info():
    from app.engine import solver as sv
    from sqlalchemy import exists as sa_exists
    from app.models.shared_module_group import SharedModuleGroup

    def fmt(t):
        return t.strftime('%H:%M')

    sv_shared_group_count = SharedModuleGroup.query.count()
    constraints = _build_constraints_reference(sv, sv_shared_group_count)

    total_course_count = Course.query.count()
    courses_with_year_range = Course.query.filter(Course.official_year_range.isnot(None)).count()

    # ---- Assumptions / self-input data disclosure ------------------------
    courses_missing_split, no_prof_sessions = _data_quality_gaps()

    no_group_sessions = ClassSession.query.filter(
        ClassSession.delivery_mode == 'f2f',
        ClassSession.student_group_id.is_(None),
        sa_exists().where(ClassSessionProfessor.session_id == ClassSession.id),
    ).all()

    no_prof_modules = sorted({
        f'{s.course.programme.code} {s.course.module_code}' for s in no_prof_sessions
    })

    no_fixed_room_sessions = ClassSession.query.filter(
        ClassSession.fixed_timeslot_id.isnot(None)
    ).count()

    long_sessions = ClassSession.query.filter(
        ClassSession.duration_hours > sv.LONG_SESSION_THRESHOLD_HOURS
    ).count()

    # Live check: which programmes in the DB aren't covered by the Template 2
    # export's Programme -> Sector/Campus mapping (T2_PROG_SECTOR)? These would
    # silently fall through to the PUNGGOL/PU default (all SIT programmes are
    # now consolidated at Punggol - confirmed by Brian 2026-07-10).
    all_prog_codes = {p.code for p in Programme.query.all()}
    unmapped_prog_codes = sorted(all_prog_codes - set(T2_PROG_SECTOR.keys()))

    cal_years = ', '.join(sorted(SIT_ACADEMIC_CALENDAR.keys()))
    cal_date_count = sum(len(v) for v in SIT_ACADEMIC_CALENDAR.values())

    # Live check: modules with 2+ quiz sessions sharing a teaching week (max 1
    # quiz/week per Ms. Yang's requirements doc - confirmed with Brian on
    # 2026-07-10 that "class" means one module's own class, not a student
    # group's full set of modules). This can't be fixed by the solver -
    # teaching_weeks is fixed input data - so it's surfaced here and as a
    # generation-blocking issue on the Timetable page (checker.py).
    from app.engine.checker import get_blocking_issues as _get_blocking_issues
    _blockers, _ = _get_blocking_issues()
    quiz_overlap_count = sum(1 for b in _blockers if 'quiz' in b.lower())

    # Cross-programme shared module linking (Common Modules / Programme Grouping)
    # stats - see bootstrap/28_load_common_modules.py for the actual matching logic.
    shared_groups_all = SharedModuleGroup.query.all()
    shared_sessions_linked = sum(len(g.class_sessions) for g in shared_groups_all)

    # Which programmes are sourced from the team's cleaned data vs the raw
    # upload - see bootstrap/32_load_cleaned_eng_data.py.
    CLEANED_DATA_PROGS = {'CVE', 'MEC', 'METS', 'EEE', 'ISE', 'RSE', 'SBE', 'DSC'}
    raw_data_progs = sorted(all_prog_codes - CLEANED_DATA_PROGS)
    null_weeks_cleaned = ClassSession.query.join(Course).join(Programme).filter(
        Programme.code.in_(CLEANED_DATA_PROGS),
        ClassSession.teaching_weeks.is_(None),
    ).count()
    fixed_room_count = ClassSession.query.filter(ClassSession.fixed_room_id.isnot(None)).count()

    # Live check: classes locked to a real, named room (fixed_room_id) whose
    # capacity is smaller than the group size actually shown for that class.
    # solver.py deliberately skips the capacity check for these (see
    # _room_compatible's require_capacity=False) - a human-confirmed real
    # room assignment from the cleaned files is treated as more trustworthy
    # than the capacity heuristic. Found 2026-07-18 while double-checking a
    # Template 2 export against every constraint: uses effective_group_size
    # (see ClassSession), not the raw student_group.intake_size, since a lab
    # split into simultaneous P1/P2 sections shares one StudentGroup record
    # and effective_group_size already corrects for that - this count is
    # what's left after that correction, genuinely unsplit single sessions.
    over_capacity_fixed_rooms = []
    for s in ClassSession.query.filter(ClassSession.fixed_room_id.isnot(None)).all():
        size = s.effective_group_size
        if s.fixed_room and s.fixed_room.capacity and size and size > s.fixed_room.capacity:
            over_capacity_fixed_rooms.append(s)

    # Change Log - dated, one-off audit-trail entries (bugs found & fixed, data
    # migrations). Kept separate from assumption_groups below: these describe
    # PAST corrections, not standing "this is how the system currently assumes
    # things" facts - mixing the two made the page hard to scan.
    changelog_groups = [
        {'category': 'Data Quality Fixes - 2026-07-10', 'icon': 'bi-wrench-adjustable', 'color': 'red', 'kind': 'value', 'rows': [
            {'label': 'Some class weeks were being read wrong', 'value': '82 classes corrected, 27 missing classes restored',
             'note': 'Some spreadsheet cells combined a week number with a date and time note (e.g. "Week 3, '
                     'Wed 17 Sep, 2-4pm"). The system was mistakenly reading every number in that text as a week '
                     'number, not just the real one. Fixed so only the actual week number is used. This also '
                     'explains why 27 classes had gone missing entirely: two different classes had accidentally '
                     'come out looking identical after the mistake, so the system kept only one of them. Both are '
                     'now correctly in place.'},
            {'label': 'Some Year 4 Mechatronics (METS) classes were mislabelled as Year 2', 'value': '4 modules corrected',
             'note': 'A previous fix for this exact issue was only applied in one of two places that needed it, '
                     'so it didn\'t take effect for regular imports. 4 modules (MET4004, MET4305, MET4505, '
                     'MET4604) were stored as Year 2 even though everything else about them correctly said '
                     'Year 4. Now fixed everywhere.'},
            {'label': "5 files' class weeks were missing entirely", 'value': 'EPE Years 1-3, METS Years 2-4',
             'note': 'Most uploaded files label the class-weeks column "Teaching Weeks", but these 5 files just '
                     'call it "Weeks". The system only recognised the longer name, so it silently skipped that '
                     'whole column in these files - every class in them had no week information at all until '
                     'this was fixed.'},
            {'label': 'Clarified what "one quiz per week" actually means', 'value': "Per module, not per student",
             'note': 'Confirmed with Brian: the rule limits each individual module to one quiz per week - it '
                     'does not mean a student can only have one quiz across all their modules in a week. Reading '
                     'it the wrong way, the system was flagging 67 problems that weren\'t real; read correctly, '
                     'it found 3 genuine ones (now down to 1 - see Data Gaps).'},
            {'label': 'Shared modules were missing real class data for most programmes', 'value': '62 classes created or corrected across 6 modules',
             'note': 'Six modules taken by students across several programmes at once (ENG1001, ENG1004, '
                     'ENG1005, ENG1008, ENG1010, ENG3001) were being skipped by the import completely, so most '
                     'programmes taking them had no timetable data at all. Also fixed: two programmes (MDME and '
                     'SBE) had been wrongly scheduled into the exact same lecture at the exact same time for '
                     'ENG1001, when they should each have their own separate lecture with a different lecturer.'},
            {'label': 'Export file now correctly shows Punggol as the campus', 'value': 'All programmes',
             'note': 'Confirmed with Brian: every SIT programme is now based at Punggol. The exported file used '
                     'to spread programmes across several old campus names from before the move. Checked '
                     'against the official sample file and corrected.'},
        ]},
        {'category': 'Cleaned Data Migration - 2026-07-10', 'icon': 'bi-arrow-repeat', 'color': 'red', 'kind': 'value', 'rows': [
            {'label': '8 programmes now use the verified data your team cleaned', 'value': f'{len(CLEANED_DATA_PROGS)} on cleaned data, {len(raw_data_progs)} still on the original upload',
             'note': 'CVE, MEC, METS, EEE, ISE, RSE, SBE and DSC now load from your team\'s cleaned, verified '
                     'files instead of the original raw upload. The other programmes don\'t have a cleaned '
                     f'version yet, so they still use the original data for now: {", ".join(raw_data_progs)}.'},
            {'label': 'Terms 2 and 3 are now loaded too', 'value': 'All 3 terms for these 8 programmes',
             'note': 'Originally only Term 1 was brought in. Extended to Terms 2 and 3 as well - the cleaned '
                     'files had this data all along, just not used yet. Two exceptions found along the way: MEC '
                     'and part of EEE\'s Term 2/3 sheets are just a module list without real class details (only '
                     '1 of 103 rows had an actual activity/time filled in) - nothing usable to bring in there '
                     'yet, that\'s a gap in the source file itself, not something missed.'},
            {'label': 'Exact room assignments from the cleaned files are now honoured', 'value': f'{fixed_room_count} classes locked to their named room',
             'note': 'The system can now "lock" a class to one specific room, the same way it could already lock '
                     'a class to one specific time. Matched the named venues in the cleaned files against actual '
                     'rooms in the system - about a third matched confidently and got locked; the rest use a '
                     'different short-form room naming (e.g. "TR1", "LAB1") that doesn\'t match how rooms are '
                     'named elsewhere in the system, or reference a room that doesn\'t exist in the room list at '
                     'all yet - left alone rather than guessed at.'},
            {'label': "Some activities were left out because they don't need a weekly time slot", 'value': '10 items excluded (assignments, capstone projects, work placements, and similar)',
             'note': 'Things like assignments, capstone projects, and work placements aren\'t regular weekly '
                     'classes, so they were left out - the same treatment already given to internships and '
                     'attachments. One extra case caught by hand: two work-placement modules were labelled as '
                     'ordinary classes in the file, and their real-world dates had been misread as week numbers '
                     '- spotted as suspicious and removed manually.'},
            {'label': 'A second shared-module group (EEE/ISE) is now connected too', 'value': '4 modules linked (ENG1101-ENG1104)',
             'note': 'Confirmed these 4 modules are genuinely the same class taught to EEE and ISE students '
                     'together (identical class weeks and identical named lecturer on both sides - not a '
                     'coincidence). Linked the same way ENG1001 already is. A 5th similar-looking module '
                     '(ENG1100) did NOT show this same pattern (different class structure and no lecturer named '
                     'on the EEE side), so it was left alone rather than assumed to match. EEE and ISE also '
                     'reference a much larger second family of shared modules in Terms 2/3 - not yet '
                     'investigated, flagged for a future pass.'},
            {'label': 'Old timetable entries for these 8 programmes will need regenerating', 'value': '230 old class records replaced, 21 manually-locked time slots lost',
             'note': 'Switching to the cleaned data meant removing the old class records for these 8 programmes '
                     'and creating fresh ones (some class codes changed in the process). 21 classes that an '
                     'admin had manually locked to a specific time slot lost that lock and would need to be set '
                     'again if still needed.'},
        ]},
        {'category': 'Filling In Missing Professors - 2026-07-10', 'icon': 'bi-person-check', 'color': 'red', 'kind': 'value', 'rows': [
            {'label': 'Classes are no longer dropped just for missing a professor', 'value': 'Room and time are still assigned',
             'note': 'Previously, any class with no professor was skipped entirely and simply didn\'t appear on '
                     'the timetable or in the Template 2 export. Now it still gets a room and a time slot - only '
                     'the staff name is left blank if genuinely unknown. This surfaces the gap instead of hiding '
                     'it by silently dropping the class.'},
            {'label': 'Filled in from the same module\'s own known lecturer', 'value': '47 classes across 33 modules',
             'note': 'Where a module had one component with a named lecturer (usually the lecture) but other '
                     'components (tutorial, lab, quiz) left blank in the source, that same lecturer was applied '
                     'to the blank ones. This is a real inference, not a guess: it\'s the one named person '
                     'already tied to that specific module, not picked at random. Includes the "All instructors" '
                     'ENG1001 quizzes (linked to each programme\'s own ENG1001 lecturer) and EPE2300\'s second '
                     'section (see below).'},
            {'label': 'EPE2300 split into two labelled sections', 'value': 'Section A and Section B',
             'note': 'The source data showed two complete, unlabelled parallel sections (different lecturers, '
                     'each with their own Lecture/Tutorial/Lab/Quiz) with no "Group A/B" marker to tell them '
                     'apart, which was earlier flagged as a possible false quiz-conflict. Labelled them '
                     'Section A (Thaiyal Naayagi Ramasamy / Khalid Seyed Saeed Ahsan Abidi) and Section B '
                     '(Anurag Sharma) based on the staff split already in the data - resolves the flag without '
                     'inventing anything not already stated.'},
            {'label': 'Quiz-clash check now aware of sections', 'value': 'Compares within the same section only',
             'note': 'Following the EPE2300 fix, the quiz check now treats differently-labelled sections of the '
                     'same module as separate classes, instead of comparing every quiz in the module against '
                     'every other regardless of section.'},
            {'label': 'Left honestly blank - no lecturer named anywhere in the source', 'value': f'{len(no_prof_sessions)} classes across {len(no_prof_modules)} modules',
             'note': 'For these, there was no real person to infer from - not even a same-module sibling class '
                     'with a name attached. Inventing a specific person\'s name here would mean attaching a real '
                     'identity to a class they never agreed to teach, which is different from an assumption - '
                     'it would be presenting invented information as fact. These are left with a blank staff '
                     'field and listed here for the programme coordinators to fill in: '
                     f'{", ".join(no_prof_modules)}.'},
        ]},
        {'category': 'Raw-Data Programmes Audited - 2026-07-10', 'icon': 'bi-search', 'color': 'blue', 'kind': 'value', 'rows': [
            {'label': "Re-checked the 8 programmes still on the original upload", 'value': 'ASE, CPC, EDE, EPE, ESE, MDME, NAME, SDE',
             'note': 'Ran the same week-number and year-level checks used earlier today against these 8 - they '
                     'came back clean, since the underlying parser fixes were applied system-wide, not just to '
                     'the 8 programmes that got switched to cleaned data.'},
            {'label': 'One more class weeks mistake found and fixed', 'value': 'ESE3112B (1 workshop)',
             'note': 'The source cell said "26 Sept 25", a real calendar date, not a week number. Worked out '
                     'which teaching week that date actually falls in using the official SIT term start date '
                     '(1 Sept 2025 = Week 1) - it\'s Week 4. Corrected from a nonsense value.'},
            {'label': '"CPC" is an empty programme entry', 'value': '0 modules, 0 classes',
             'note': 'This programme code exists in the system (from the Template 2 export mapping) but has no '
                     'course or class data at all - nothing to audit here yet. Worth checking with Brian whether '
                     'CPC should have real data or can be removed.'},
            {'label': "NAME's Year 3 file has almost no real data", 'value': 'Only 1 real row, and it\'s excluded from scheduling',
             'note': 'The uploaded file for NAME Year 3 is 47 rows long but only 1 has real content, and that '
                     'one is a Practicum placement (already correctly excluded from scheduling, same as other '
                     'internship-type activities). Not a bug - NAME genuinely has no Year 3 classroom data yet.'},
        ]},
        {'category': 'DSC Real Timetable Restored - 2026-07-10', 'icon': 'bi-exclamation-octagon', 'color': 'red', 'kind': 'value', 'rows': [
            {'label': 'DSC is the ONLY programme with a real, already-committed timetable', 'value': f'{TimetableEntry.query.filter_by(is_backbone=True).count()} real class times restored',
             'note': 'Confirmed with Brian: a finalised real timetable was only ever submitted for DSC (3 files, '
                     'one per term) - no other programme has been given one. So DSC is treated as a special '
                     'case, not the norm: its real schedule is the source of truth and the system must reproduce '
                     'it exactly, never recompute it. Every other programme has no real timetable to work from, '
                     'which is exactly why the system computes their schedules itself. Earlier today\'s general '
                     '"switch to cleaned data" step wrongly treated DSC the same as the other 7 programmes and '
                     'deleted this real timetable data by mistake - caught and fixed immediately once flagged.'},
            {'label': 'How this actually works', 'value': 'DSC classes strongly prefer their real time; every other programme is computed freely',
             'note': 'For DSC only, each class\'s real day/time/room (exactly as the professor submitted) is '
                     'loaded in as a strong preference, not an unbreakable lock - the system will reproduce it '
                     'exactly whenever nothing stops it, and only move a class when there is a genuine reason '
                     '(e.g. the same professor is also needed for another class at that time). All 15 other '
                     'programmes have no submitted timetable at all, so they continue to have their schedule '
                     'worked out by the system as normal - there is nothing to prefer for them. A few classes in '
                     'DSC\'s files use very short or very oddly-timed slots (e.g. a 1-hour Friday slot, evening '
                     'classes after 6pm for a related module) that don\'t match any time slot currently defined '
                     'in the system - those specific classes could not be preferred and are flagged, not '
                     'silently dropped.'},
            {'label': 'A few names and rooms in DSC\'s files could not be matched', 'value': '4 unmatched rooms, 3 unmatched staff names',
             'note': 'Some room codes (e.g. in the W3 building) and a couple of staff names in DSC\'s submitted '
                     'files don\'t match anything currently in the system\'s room or staff list - those specific '
                     'classes were still scheduled at their correct time, just without a room or staff name '
                     'attached yet.'},
        ]},
        {'category': 'DSC Backbone: Locking Removed, Data Bugs Fixed - 2026-07-10', 'icon': 'bi-wrench-adjustable', 'color': 'orange', 'kind': 'value', 'rows': [
            {'label': 'DSC\'s real timetable is no longer an unbreakable lock', 'value': 'Changed from a hard lock to a strong preference',
             'note': 'Previously, every DSC class was locked to its exact real-world time with zero flexibility. '
                     'This broke Trimester 2 entirely (the system could not generate any timetable for it) '
                     'because one professor\'s locked class left no possible time for a second class they also '
                     'teach. Fixed by changing DSC\'s real schedule into a very strong scheduling preference '
                     'instead - the system now reproduces it exactly whenever nothing else is affected, and only '
                     'shifts a class when there is a genuine reason to. Result: Trimester 1 reproduced its real '
                     'schedule exactly for 6 of 8 classes, Trimester 2 for 20 of 23, Trimester 3 for all 3.'},
            {'label': 'Two data-matching bugs found and fixed in the same pass', 'value': '2 bugs found via direct comparison against the original Excel files',
             'note': 'While investigating the Trimester 2 failure: (1) the professor-name-matching logic was '
                     'accidentally matching by any partial text overlap, which once wrongly linked a professor '
                     'named "Wang Yu" to a completely different professor named "Wang Fengyu" (because "Yu" '
                     'happens to appear inside "Fengyu") - fixed to only match whole names; (2) some of DSC\'s '
                     'parallel class sections (e.g. 4 separate seminar groups of the same module, each with a '
                     'different lecturer) were being merged into fewer database entries than actually exist, '
                     'which wrongly mixed up which lecturer taught which group - fixed the grouping logic and '
                     'confirmed against the original files that every professor is now correctly attached to '
                     'their own real section only.'},
        ]},
        {'category': 'Template 2 Export: Checked Against Ms. Yang\'s Own Reference File - 2026-07-10', 'icon': 'bi-file-earmark-check', 'color': 'green', 'kind': 'value', 'rows': [
            {'label': 'Every column name, order, and value format checked field-by-field', 'value': '31/31 columns confirmed correct',
             'note': 'Compared this system\'s Template 2 export directly against Ms. Yang\'s own reference file '
                     '("Worksheet in ITP Project Requirements (Template 2).xlsx"), including its built-in '
                     '"Class Type" code list (the official Activity Type/Class Type names) and its "Zone", '
                     '"Day", "Time", and "Course Code" lookup sheets. All 31 columns match her expected names, '
                     'order, and value formats (day abbreviations, HHMM time format, hostkey formats, zone '
                     'names, and so on).'},
            {'label': 'Two real mismatches found and fixed', 'value': '"Lab"->"Laboratory", workshop code "WRK"->"WOR"',
             'note': 'Her official Class Type list names the activity "Laboratory", not "Lab", and codes '
                     'workshop sessions as "WOR", not "WRK" (this system had used "WRK"). Both were silent '
                     'mismatches that would have shown up as wrong values in her import, not caught by any '
                     'internal test since our own code was internally consistent - only caught by comparing '
                     'directly against her reference file. Fixed both, and also fixed the "Term" column to be a '
                     'true number (e.g. 2510) instead of a text value, matching her file\'s exact column type.'},
            {'label': '"Room1" now filled in with the real assigned room', 'value': 'Was always blank - now shows the actual scheduled room',
             'note': 'Ms. Yang\'s reference file left this column blank in every example row, but that file is a '
                     'pre-scheduling requirements list (rooms not decided yet), while this export happens after '
                     'scheduling (rooms are known). Brian confirmed the real room should be included. One thing '
                     'to flag: the room codes used here (e.g. "E2-07-01") are this system\'s own internal '
                     'naming, not confirmed against a Punggol-specific room list from Ms. Yang\'s side - her '
                     'reference file\'s own room lookup sheet only has Dover-campus rooms, so there is nothing '
                     'to check the exact expected Punggol format against yet. "Room2" and "RoomGrouping" stay '
                     'blank - there is no data source for a second room per class anywhere in the system.'},
        ]},
        {'category': 'Full System Audit: Fabricated Professor Names Found and Removed - 2026-07-11', 'icon': 'bi-search', 'color': 'red', 'kind': 'value', 'rows': [
            {'label': '14 DSC class sessions had a made-up professor name, not from any source file', 'value': 'Found by cross-checking every name against all 3 real files',
             'note': 'Names like "Bob Chen", "Alice Wong", and "Frank Goh" were attached to some DSC classes even '
                     'though those names never appear anywhere in DSC\'s real submitted timetable. Traced to '
                     'leftover records from a brief, earlier mistake where DSC was processed like every other '
                     'programme before being switched to its real timetable (see the DSC entry above) - those old, '
                     'wrong names were never cleaned up, and a routine "fill in a likely name from a nearby class" '
                     'step then spread them further. Every affected class was checked against the original Excel '
                     'files: where a real name existed, it was used; where a class genuinely has no equivalent in '
                     'the real timetable, it is now correctly shown as unstaffed rather than guessed.'},
            {'label': 'Quiz sessions were being silently skipped on import', 'value': 'Fixed - quizzes and workshops now import correctly',
             'note': 'A labelling bug meant every quiz class in DSC\'s real timetable file (13 of them) was quietly '
                     'ignored during import, and workshop classes were mislabelled as tutorials. Fixed; several '
                     'real quiz classes are now correctly loaded with their real time and professor.'},
            {'label': '5 real quizzes needed new time slots the system didn\'t have', 'value': 'Fixed - added the missing time slots rather than drop the real data',
             'note': 'DSC3303\'s quiz is genuinely 1 hour long, and 4 shared-module quizzes (INF1009, INF2008, '
                     'SDE3001, SDE3002) are genuinely held 6pm-8pm in the evening - neither existed as a time slot '
                     'option in the system. Since Brian confirmed the real timetable file is accurate and these '
                     'classes really do happen, the system was updated to properly support them rather than leave '
                     'them out: added the 3 missing time slots, and locked each of these 5 classes to their exact '
                     'real time so they don\'t affect anyone else\'s schedule. The "no classes after 6pm" rule '
                     'still holds for every other class - these 5 are a disclosed, source-confirmed exception, not '
                     'a loosening of the rule generally.'},
        ]},
        {'category': 'Template 2 "Group" Column Fixed - 2026-07-15', 'icon': 'bi-file-earmark-spreadsheet', 'color': 'red', 'kind': 'value', 'rows': [
            {'label': 'Almost every exported class showed "Group: All", even genuinely split ones', 'value': '334 of 791 sessions corrected',
             'note': 'The value was set by a regex that only matched the literal word "Group" inside a free-text '
                     'Remarks cell - almost never present, so classes silently defaulted to "All" regardless of '
                     'whether they were actually split into parallel sections. The rare real matches often grabbed '
                     'garbage too (Remarks text that happened to contain "Group" followed by an unrelated word '
                     'produced values like "ASSESSMENTS", "SETTING", "WORK", "PROJECTS"). Checked against Ms. '
                     'Yang\'s own reference file to find the real convention (see "Group column letter per class '
                     'type" below), then recomputed every session\'s value from real sibling counts - see '
                     'bootstrap/39_fix_group_labels.py.'},
            {'label': 'New sessions now get a correct Group value automatically', 'value': 'Wired into session creation',
             'note': 'Recomputed any time the number of sessions of one type for a module changes - splitting via '
                     'the Sessions page\'s split count, or adding a new session type - so this can\'t silently '
                     'drift back to "All" for new data the way it did before.'},
        ]},
        {'category': 'Template 2 Silent Scheduling Gaps Fixed - 2026-07-16', 'icon': 'bi-file-earmark-spreadsheet', 'color': 'red', 'kind': 'value', 'rows': [
            {'label': 'Some sessions vanished from Template 2 with no explanation', 'value': '17 sessions in T1 affected, all now visible',
             'note': 'The solver picked each session\'s day/time with zero awareness of public holidays or '
                     'cancelled Events - fine for a normal weekly class (losing 1 of 13 weeks to a holiday is a '
                     'minor, visible gap), but a session with only 1-3 sparse teaching weeks (a one-off quiz, for '
                     'example) could have its single occurrence land on a blocked date and disappear entirely. '
                     'The solver\'s own stats still reported it as "scheduled" - the loss happened silently, one '
                     'step after solving. Fixed by teaching the solver to avoid any day that would wipe out every '
                     'occurrence of a session before it ever picks one; 7 of the 17 were rescheduled onto a '
                     'working day automatically.'},
            {'label': '10 sessions referenced a teaching week beyond the trimester\'s calendar', 'value': '10 removed, 1 replacement added',
             'note': 'A handful of sessions (mostly one-off quizzes) had teaching_weeks = 14 or 15, but this '
                     'system\'s academic calendar model is fixed at 13 weeks per trimester - those weeks simply '
                     'don\'t exist to schedule into, and the new check above would only ever report them as '
                     'skipped, never fix them (extending the calendar model itself was judged too risky this close '
                     'to the Template 2 deadline, since it\'s a system-wide assumption). Confirmed with Brian these '
                     'values were never real (matches the same conclusion bootstrap/29 already reached for other '
                     'quiz sessions on 15 July) - removed all 10 via bootstrap/40. One of them, ASE2210, had '
                     'ONLY these invalid sessions and nothing else, so removing them left the module with zero '
                     'sessions, which blocks generation outright. Added one basic replacement Lecture session '
                     '(bootstrap/41) - no hard constraints, same student group and standard weekly pattern every '
                     'other lecture uses - so ASE2210 stays part of the output instead of silently disappearing.'},
            {'label': '22 Year-3 ASE sessions had no student group at all - not even the group was missing, the whole cohort was', 'value': 'Created "ASE-Y3" group, intake 55 (estimated)',
             'note': 'ASE3106/3108/3109/3110 (all Trimester 1) had zero sessions with a student_group_id set. Unlike '
                     'ASE2210 (where a real ASE-Y2 group already existed to attach), there was no ASE-Y3 group '
                     'anywhere in the system to begin with - ASE only had Y1 (75 students) and Y2 (62 students). '
                     'Without a group the solver can\'t size a room, so it deliberately skips the class rather than '
                     'guess (previously shown as a "no student group" warning). Confirmed with Brian: create the '
                     'group so these 22 sessions get scheduled. intake_size=55 is a disclosed ESTIMATE, not real '
                     'data - continues ASE\'s observed Y1->Y2 decline, not measured. See bootstrap/44.'},
        ]},
    ]

    # Assumed / Self-Input Values - standing facts about how the system fills
    # gaps not covered by uploaded data. Unlike changelog_groups above, these
    # aren't tied to a date - they describe the system's current behaviour.
    assumption_groups = [
        {'category': 'Solver Scheduling Defaults', 'icon': 'bi-gear', 'color': 'blue', 'kind': 'value', 'rows': [
            {'label': 'Default week to treat as a term break', 'value': f'Week {sorted(sv.DEFAULT_TERM_BREAK_WEEKS)}',
             'note': 'Only used if a specific break week isn\'t chosen when generating a timetable.'},
            {'label': 'Wednesday afternoon cutoff', 'value': fmt(sv.WED_AFTERNOON_CUTOFF),
             'note': 'SIT\'s CCA (co-curricular activity) policy - a fixed school-wide rule, not something read '
                     'from any uploaded file.'},
            {'label': 'Friday protected window', 'value': f'{fmt(sv.FRI_BLOCK_START)}-{fmt(sv.FRI_BLOCK_END)}',
             'note': 'A fixed SIT-wide rule, not something read from any uploaded file.'},
            {'label': 'Lunch window (flexible, not a fixed block)', 'value': f'{fmt(sv.LUNCH_WINDOW_START)}-{fmt(sv.LUNCH_WINDOW_END)}',
             'note': 'Every class group needs at least 1 free hour somewhere in this window, on any day - it '
                     'isn\'t one fixed hour blocked for everyone (corrected per Ms. Yang, who clarified the '
                     'original "no class 12:00-13:00" reading was wrong).'},
            {'label': 'Preferred end-of-day cutoff', 'value': fmt(sv.LATE_END_CUTOFF),
             'note': 'Just a preference, not a hard rule - a class can still be scheduled later than this if '
                     'there\'s no earlier slot available.'},
        ]},
        {'category': 'Room & Capacity Rules', 'icon': 'bi-door-open', 'color': 'green', 'kind': 'value', 'rows': [
            {'label': 'Which room types can host which class types', 'value': 'A rule we set ourselves',
             'note': 'Lab classes need a lab room. Lectures, lectorials and quizzes need a lecture or seminar '
                     'room. Tutorials, seminars and workshops also use a seminar or lecture room. This matching '
                     'isn\'t based on any field in the uploaded data - it\'s a rule the system applies on its own.'},
            {'label': 'Class size used when no group is linked', 'value': '1 student',
             'note': 'If a class has no student group linked to it, the system assumes only 1 student when '
                     'checking room size and usage - almost certainly too low for a real class that\'s simply '
                     'missing its group link.'},
        ]},
        {'category': 'Academic Calendar & Timeslots', 'icon': 'bi-calendar3', 'color': 'purple', 'kind': 'value', 'rows': [
            {'label': 'SIT academic calendar (when each term starts)', 'value': f'{cal_date_count} dates ({cal_years})',
             'note': 'Typed in by hand from SIT\'s public academic calendar page, not from any uploaded file. '
                     'Should be double-checked at the start of every new academic year.'},
            {'label': 'Standard daily time slots', 'value': '35 slots (5 days x 7 periods)',
             'note': 'Set up by the project team from SIT\'s standard period-block reference sheet - fixed '
                     'unless SIT changes its daily class-period structure.'},
            {'label': 'Public holiday list', 'value': "A general Singapore public-holiday list",
             'note': 'Not a list curated specifically for SIT - it\'s the general Singapore public holiday '
                     'calendar, which can differ from SIT\'s own observed closures (e.g. school-specific '
                     'makeup or closure days).'},
        ]},
        {'category': 'Template 2 Export Mappings', 'icon': 'bi-file-earmark-spreadsheet', 'color': 'red', 'kind': 'value', 'rows': [
            {'label': 'Programme to campus mapping',
             'value': f'{len(T2_PROG_SECTOR)} mapped' + (f', {len(unmapped_prog_codes)} UNMAPPED' if unmapped_prog_codes else ', all covered'),
             'risk': bool(unmapped_prog_codes),
             'note': ('All SIT programmes are now based at the Punggol campus (confirmed by Brian 2026-07-10), '
                      'so every programme - listed here or not - exports as Punggol. '
                      f'Not explicitly listed yet (still fine, defaults to Punggol anyway): {", ".join(unmapped_prog_codes)}.'
                      if unmapped_prog_codes else
                      'Every current programme is explicitly listed, and any new one added later would also '
                      'default to Punggol automatically - no other campus is in use.')},
            {'label': 'Class-type abbreviation mapping', 'value': f'{len(T2_ACT_CODE)} class types',
             'note': 'e.g. lecture becomes "LEC", lectorial becomes "LET", tutorial becomes "TUT". Any class '
                     'type not on this list exports as "OTH".'},
            {'label': 'Cluster abbreviation mapping', 'value': f'{len(T2_CLUSTER_ABBR)} explicit + fallback',
             'note': 'A cluster name not on this list falls back to its first 3 letters, capitalised - which '
                     'could produce an odd-looking abbreviation for a brand-new cluster name.'},
            {'label': 'Class-length units', 'value': '1 hour = 3 units (20-minute blocks)',
             'note': 'The exported file expects class length in 20-minute units, so the system multiplies the '
                     'hour count by 3. This matches SIT\'s own template format, but isn\'t written down anywhere '
                     'in it - worked out from the sample file.'},
            {'label': 'Which classes get marked as "recorded"', 'value': 'Lectorials only',
             'note': 'Every other class type exports as "not recorded" - based on the assumption that only '
                     'lectorials are ever recorded.'},
            {'label': '"Term start week" column', 'value': 'Always set to 1',
             'note': 'Every exported row shows "1" here - it isn\'t calculated from the real calendar or term '
                     'break dates.'},
            {'label': '"Group" column letter per class type', 'value': 'L / T / P / S / W / Q',
             'note': 'Checked against Ms. Yang\'s own reference file: Lecture and Lectorial both use "L", '
                     'Tutorial "T", Workshop "W", and Quiz "Q" - all confirmed directly from real rows in that '
                     'file. Lab ("P") and Seminar ("S") have no real example there, so those two are a disclosed '
                     'best-fit, not a confirmed convention: P avoids colliding with Lecture\'s "L", S matches '
                     '"Seminar" itself. The number after the letter (e.g. T1, T2, T3) counts which parallel split '
                     'of that class type this is - a class with only one (unsplit) session exports as "All", '
                     'except Quiz, which the reference file always numbers even for a single quiz group.'},
        ]},
        {'category': 'Cross-Programme Coordination', 'icon': 'bi-diagram-3', 'color': 'red', 'kind': 'value', 'rows': [
            {'label': 'Classes shared across multiple programmes', 'value': f'{len(shared_groups_all)} shared group(s), {shared_sessions_linked} class(es) linked',
             'note': 'Built from two of your source files: one lists which modules are shared by which '
                     'programmes in free text, matched as closely as possible; the other spells out full class '
                     'schedules for shared modules, but its layout varies from module to module, so it was '
                     'typed up carefully by hand instead of read automatically, to avoid mistakes.'},
            {'label': 'Only the lecture is shared, not the tutorial/lab/quiz', 'value': '1 combined lecture per programme group',
             'note': 'When several programmes share a module, only their lecture is scheduled at the same time '
                     'for everyone - tutorials, labs, and quizzes stay separate per programme. This is a '
                     'judgment call, not something stated in the source files. Where a file only gave one row '
                     'covering several programmes for a tutorial/lab/quiz (no per-programme breakdown), the same '
                     'details were copied once for each programme rather than merged into a single class.'},
            {'label': 'Room size for shared classes', 'value': 'Big enough for every programme combined',
             'note': 'When several programmes share one class, the room chosen must fit everyone from all of '
                     'those programmes together - this isn\'t from any uploaded file, it\'s what made sense to us.'},
        ]},
        {'category': 'Module Reference Data', 'icon': 'bi-journal-bookmark', 'color': 'blue', 'kind': 'value', 'rows': [
            {'label': 'Official (SIT catalog) year range per module', 'value': f'{courses_with_year_range}/{total_course_count} filled in',
             'note': 'A module can legitimately span several years in SIT\'s own module catalog (sit.edu.sg) '
                     'even though it\'s only scheduled here for one specific year - Year Level (used for '
                     'scheduling) and Official Year Range (this field) are two different things and are not '
                     'always the same number. This is admin-entered reference text with no automated source - '
                     'it is never read by the solver and has no effect on scheduling. Editable on each module\'s '
                     'Edit page or in bulk via the Modules page\'s Import/Export.'},
        ]},
    ]

    # Data Gaps - live counts of missing/incomplete data needing attention.
    # Kept separate from assumption_groups: a gap is "we don't have this data
    # at all", not "we filled in a value" - a different kind of disclosure.
    data_gap_groups = [
        {'category': 'Data Gaps', 'icon': 'bi-exclamation-triangle', 'color': 'amber', 'kind': 'count', 'rows': [
            {'label': 'Classes with no professor assigned', 'count': len(no_prof_sessions),
             'note': 'Still scheduled - gets a room and time either way - but shows a blank staff field on '
                     'the timetable and Template 2 export until a professor is assigned.'},
            {'label': 'In-person classes with no student group assigned', 'count': len(no_group_sessions),
             'note': 'Also skipped entirely until a group is assigned.'},
            {'label': 'Modules missing a group-split setting', 'count': len(courses_missing_split),
             'note': 'Generation is blocked until an admin sets how many groups this module should be split '
                     'into - see the Dashboard warning.'},
            {'label': 'Classes with a manually-set time slot', 'count': no_fixed_room_sessions,
             'note': 'These were placed by an admin directly, not worked out by the system.'},
            {'label': 'Classes locked to a real room smaller than their group size', 'count': len(over_capacity_fixed_rooms),
             'note': f'{len(over_capacity_fixed_rooms)} remaining, all in Trimester 2/3 or one of the 7 '
                     f'programmes deferred from T1 generation - zero left in the actual T1 Template 2 export '
                     f'(2026-07-18: verified against the file directly). Spans CVE, ISE and METS. Two fixes '
                     f'applied, both authorized by Brian as a normal, expected pattern ("some modules will have '
                     f'split like P1, P2... the professor teaches all of them... we are allowed to do splits on '
                     f'our end if needed"): (1) ClassSession.effective_group_size divides the shown Class Size '
                     f'fairly across confirmed-concurrent parallel sections (same course+session_type+group, '
                     f'teaching weeks that actually overlap) instead of every section claiming the full cohort '
                     f'- fixed cases that already had a P1/P2-style split on record; (2) for T1\'s 11 remaining '
                     f'un-split sessions, generated the missing parallel sections directly (bootstrap/53, same '
                     f'professor teaching every section, matching Ms. Yang\'s own real sample data\'s split '
                     f'pattern) and placed them into free slots (bootstrap/55) - 3 modules (ENG1102/1103/1104, '
                     f'one very busy shared ISE cohort) had no free calendar slots left for a 4-way split, so '
                     f'those instead moved into one larger real, unbooked room (E2-08-01, cap 100) rather than '
                     f'splitting at all. Re-verified zero professor/room/group conflicts across the whole T1 '
                     f'schedule afterward. The remaining T2/T3 cases follow the same fixable pattern, just not '
                     f'yet applied - not part of this submission\'s Template 2 file.'},
            {'label': f'Single classes longer than {sv.LONG_SESSION_THRESHOLD_HOURS} hours', 'count': long_sessions,
             'note': 'For information only - a class\'s length is fixed before scheduling even starts, so this '
                     'isn\'t something the system controls. Some engineering labs are legitimately this long; '
                     'worth reviewing case by case.'},
            {'label': 'Modules with more than one quiz in the same week', 'count': quiz_overlap_count,
             'note': 'This blocks timetable generation completely - each module is only allowed one quiz per '
                     'week, and the system can\'t move a quiz to a different week since that comes from the '
                     'uploaded data, not a scheduling choice. Confirmed with Brian that this rule applies per '
                     'module, not per student. The one case still open (EPE2300) might not be a real problem: '
                     'the source file shows two separate, unlabelled groups of students (different lecturers) '
                     'with no way to tell them apart - not fixed, since we didn\'t want to invent a label the '
                     'data itself didn\'t give us.'},
            {'label': 'Modules shared between two intake years under one record', 'count': 2,
             'note': 'MME3201A and SBE3113A are each referenced by two different years\' files, but the '
                     'database can currently only store one year per module - a pre-existing limitation, not '
                     'something from today\'s changes. Neither has week information either way, so nothing is '
                     'at risk; left as-is until there\'s a decision on how to properly support a module offered '
                     'across two cohort years.'},
            {'label': 'No way yet to detect overlapping elective choices', 'count': 1,
             'note': 'Ms. Yang\'s requirements mention checking for overlaps between elective options a student '
                     'might pick - but nowhere in any uploaded file (raw or cleaned) is there a flag marking '
                     'which modules are electives, or which electives are alternatives to each other. The '
                     'closest thing found: DSC\'s Term 2 file reserves two generic "Elective 1" / "Elective 2" '
                     'time slots (each with its own dedicated lecturer already assigned) rather than naming '
                     'specific elective modules at all - this sidesteps the overlap question entirely rather '
                     'than answering it, and doesn\'t generalise to how other programmes might handle electives. '
                     'Still needs a real source file from Brian to implement properly.'},
            {'label': '"Lectures fully online" statement contradicts the real data', 'count': 1,
             'note': 'The requirements doc has one line saying lectures are always online with no physical room '
                     'needed - but across the entire dataset (raw and cleaned, all 16 programmes), lectures are '
                     'explicitly marked f2f (in-person) about as often as online, row by row. The system trusts '
                     'each class\'s own stated delivery mode from the uploaded data instead of this blanket rule, '
                     'since the data itself is more specific and consistent than the general statement.'},
            {'label': 'CVE3371A workshop has no Teaching Weeks value in the source file', 'count': 2,
             'note': 'Checked directly against the CVE cleaned-data source file (found 2026-07-17): the Teaching '
                     'Weeks column is blank for both CVE3371A rows, and for its f2f sibling CVE3371B too - not an '
                     'import or parsing issue, the cell itself is empty. CVE3371B still gets scheduled normally '
                     'because the solver treats a missing value as "every teaching week" for classes it actually '
                     'schedules - but CVE3371A is Online-Asynchronous, which the solver deliberately never '
                     'schedules at all (no room/time needed), so it never gets that same default and shows a '
                     'blank Tri Week in the Template 2 export instead. Not fixed with a guessed value, same '
                     'reasoning as the elective-overlap and "lectures fully online" gaps above - would need the '
                     'real weeks pattern from Ms. Yang\'s team.'},
            {'label': 'Template 2 Location sheet: 18 rooms use a placeholder code, not a real venue', 'count': 18,
             'note': 'The Location lookup sheet added to Template 2 (2026-07-18, so Room1 codes resolve against '
                     'it) includes every active room in this system, but 18 of them are internal placeholders '
                     'like "[CVE-Lab-1]" - used when a real lab room code was never given for that programme, '
                     'not an actual bookable venue. Also, only Name/Host Key/Capacity are populated for every '
                     'room - Department, Zone, and the other facilities-management columns are Ms. Yang\'s own '
                     'fields with no equivalent source data on our side, left blank rather than guessed.'},
        ]},
        {'category': 'Deferred From T1 Generation - Not A Data Gap', 'icon': 'bi-hourglass-split', 'color': 'blue', 'kind': 'count', 'rows': [
            {'label': 'Sessions temporarily excluded from Trimester 1 scheduling', 'count': 325,
             'note': 'Trimester 1 (537 sessions across 15 programmes, many professors teaching across '
                     'several programmes at once) is too large and interconnected for the CP-SAT solver to '
                     'find a valid, conflict-free schedule within any practical time budget - tested '
                     'extensively (budgets up to 25 minutes, warm-started search, sub-problem solving all '
                     'failed the same way). Ms. Yang\'s own requirements only call for a minimum of 20 '
                     'programme-year schedules for Template 2 submission, not full coverage, so 7 of the 15 '
                     'programmes (EDE, EEE, EPE, ESE, MDME, MEC, NAME - 325 sessions) are marked '
                     '"deferred_from_solve" and excluded from this generation pass. The remaining 8 '
                     'programmes (SDE, DSC, SBE, ISE, CVE, METS, RSE, ASE) cover 24 programme-year '
                     'schedules - DSC included, verified zero real conflicts. This is a deliberate, '
                     'disclosed scope decision, not missing or fabricated data - every deferred session\'s '
                     'course/module/professor data is untouched and complete, it simply isn\'t scheduled '
                     'yet. Reversible: clearing the flag and re-running generation restores full scope '
                     'whenever a longer time budget or an improved solver strategy is available.'},
        ]},
    ]

    return render_template('admin/system_info.html',
                           constraints=constraints,
                           assumption_groups=assumption_groups,
                           data_gap_groups=data_gap_groups,
                           changelog_groups=changelog_groups)


# ---------------------------------------------------------------------------
# Constraint Settings - admin on/off + priority override for soft constraints
# only (hard constraints stay fixed - they exist to guarantee a usable
# schedule and are never adjustable). Confirmed with Brian 2026-07-11.
# ---------------------------------------------------------------------------

# Relative priority dial for Constraint Settings - an admin adjusts each
# rule's importance relative to its own baseline (Higher/Lower) instead of
# typing a bare number no one can interpret ("what does 15 mean?"). Each
# tier is a multiplier on the rule's own default weight, so the relative
# relationship between rules (e.g. "ends late" already matters more than
# "not the snuggest room fit") is preserved regardless of which rule is
# being adjusted.
PRIORITY_TIERS = [
    ('much_lower', 0.25, 'Much Lower'),
    ('lower', 0.5, 'Lower'),
    ('default', 1.0, 'Default'),
    ('higher', 2.0, 'Higher'),
    ('much_higher', 4.0, 'Much Higher'),
]
_PRIORITY_TIER_MULTIPLIERS = {key: mult for key, mult, _ in PRIORITY_TIERS}


def _tier_weight(default, multiplier):
    return max(1, round(default * multiplier))


@admin_bp.route('/constraint-settings', methods=['GET', 'POST'])
@login_required
def constraint_settings():
    from app.engine import solver as sv
    from app.models.solver_setting import SolverSetting
    from app.models.shared_module_group import SharedModuleGroup

    existing = {row.constraint_id: row for row in SolverSetting.query.all()}

    if request.method == 'POST':
        for cid in sv.SOFT_CONSTRAINT_DEFAULTS:
            enabled = request.form.get(f'enabled_{cid}') == 'on'
            tier = request.form.get(f'tier_{cid}', 'default')
            default = sv.SOFT_CONSTRAINT_DEFAULTS[cid]

            row = existing.get(cid)
            if not row:
                row = SolverSetting(constraint_id=cid)
                db.session.add(row)

            row.enabled = enabled
            if tier == 'default' or tier not in _PRIORITY_TIER_MULTIPLIERS:
                row.weight_override = None
            else:
                row.weight_override = _tier_weight(default, _PRIORITY_TIER_MULTIPLIERS[tier])

        db.session.commit()
        flash('Constraint settings saved - they take effect the next time a trimester is generated.', 'success')
        return redirect(url_for('admin.constraint_settings'))

    ref = _build_constraints_reference(sv, SharedModuleGroup.query.count())

    rows_by_category = []
    for group in ref['soft']:
        rows = []
        for r in group['rows']:
            cid = r['id']
            if cid not in sv.SOFT_CONSTRAINT_DEFAULTS:
                continue  # S4 - covered by S8, no independent weight
            row = existing.get(cid)
            default = sv.SOFT_CONSTRAINT_DEFAULTS[cid]
            weight_override = row.weight_override if row else None

            if weight_override is None:
                selected_tier = 'default'
            else:
                selected_tier = min(
                    PRIORITY_TIERS,
                    key=lambda t: abs(_tier_weight(default, t[1]) - weight_override)
                )[0]

            enabled = row.enabled if row else True
            rows.append({
                'id': cid,
                'title': r['title'],
                'default': default,
                'enabled': enabled,
                'selected_tier': selected_tier,
                'effective_weight': (0 if not enabled
                                      else (weight_override if weight_override is not None else default)),
                'is_customised': (not enabled) or selected_tier != 'default',
            })
        if rows:
            rows_by_category.append({'category': group['category'], 'icon': group['icon'],
                                      'color': group['color'], 'rows': rows})

    return render_template('admin/constraint_settings.html',
                           rows_by_category=rows_by_category, priority_tiers=PRIORITY_TIERS)


# ---------------------------------------------------------------------------
# Data Import / Export
# ---------------------------------------------------------------------------

@admin_bp.route('/import')
@login_required
def data_import():
    return render_template('admin/data_import.html')


# ---------------------------------------------------------------------------
# Template 1 import - upload → preview → confirm
# ---------------------------------------------------------------------------

@admin_bp.route('/import/template1', methods=['GET', 'POST'])
@login_required
def import_template1():
    import os, uuid, re, tempfile
    import pandas as pd
    from app.engine.template1_parser import (
        load_module_sheet, PROG_NAMES, SKIP_NAMES, SKIP_SHEETS,
    )

    UPLOAD_DIR = os.path.join(tempfile.gettempdir(), 'sit_template1_uploads')
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    def _clean_staff(staff_list):
        out = []
        for name, sid in staff_list:
            n = '' if (name is None or (isinstance(name, float) and math.isnan(name))) else str(name).strip()
            s = '' if (sid is None or (isinstance(sid, float) and math.isnan(sid))) else str(sid).strip()
            if n and n.lower() not in SKIP_NAMES:
                out.append({'name': n, 'sid': s})
        return out

    def _get_or_create_programme(code):
        prog = Programme.query.filter_by(code=code).first()
        if not prog:
            prog = Programme(code=code, name=PROG_NAMES.get(code, code), cluster='ENG')
            db.session.add(prog)
            db.session.flush()
        return prog

    def _get_or_create_student_group(prog, year_level, intake_size):
        label = f'{prog.code}-Y{year_level}'
        sg = StudentGroup.query.filter_by(
            programme_id=prog.id, year_level=year_level,
            group_label=label, parent_id=None,
        ).first()
        if not sg:
            sg = StudentGroup(
                programme_id=prog.id, year_level=year_level,
                group_label=label, intake_size=intake_size or 30, parent_id=None,
            )
            db.session.add(sg)
            db.session.flush()
        elif intake_size and intake_size > 0 and sg.intake_size != intake_size:
            sg.intake_size = intake_size
        return sg

    def _get_or_create_professor(name_raw, sid_raw):
        name = '' if (name_raw is None or (isinstance(name_raw, float) and math.isnan(name_raw))) else str(name_raw).strip()
        sid  = '' if (sid_raw  is None or (isinstance(sid_raw,  float) and math.isnan(sid_raw)))  else str(sid_raw).strip()
        if not name or name.lower() in SKIP_NAMES:
            return None
        if re.fullmatch(r't\d{1,2}', name.lower()):
            return None
        name = name.title()
        sid  = re.sub(r'[^\w]', '', sid)
        if re.fullmatch(r't\d{1,2}', sid, re.IGNORECASE):
            sid = ''
        if sid:
            prof = Professor.query.filter_by(staff_id=sid).first()
            if prof:
                return prof
        user = User.query.filter_by(name=name, role='professor').first()
        if user:
            prof = Professor.query.filter_by(user_id=user.id).first()
            if prof:
                if sid and not prof.staff_id:
                    prof.staff_id = sid
                return prof
        if not sid:
            base = abs(hash(name)) % 9999
            sid = f'ENG{base:04d}'
            attempt = 0
            while Professor.query.filter_by(staff_id=sid).first():
                attempt += 1
                sid = f'ENG{(base + attempt) % 9999:04d}'
        email_local = re.sub(r'[^a-z0-9]', '.', name.lower()).strip('.')
        email_local = re.sub(r'\.+', '.', email_local)
        email = f'{email_local}@sit.edu.sg'
        attempt = 0
        while User.query.filter_by(email=email).first():
            attempt += 1
            email = f'{email_local}.{attempt}@sit.edu.sg'
        user = User(name=name, email=email, role='professor')
        user.set_password('SIT@2526')
        db.session.add(user)
        db.session.flush()
        prof = Professor(user_id=user.id, staff_id=sid, department='ENG')
        db.session.add(prof)
        db.session.flush()
        return prof

    def _get_or_create_course(module_code, title, prog, year_level, delivery_mode, tri):
        course = Course.query.filter_by(
            module_code=module_code, programme_id=prog.id, trimester=tri,
        ).first()
        if not course:
            course = Course(
                programme_id=prog.id, module_code=module_code,
                title=title or module_code, year_level=year_level,
                trimester=tri, delivery_mode=delivery_mode,
                sessions_per_week=1, total_hours=0,
            )
            db.session.add(course)
            db.session.flush()
        return course

    def _parse_file(fpath, fname, all_slots):
        xl = pd.ExcelFile(fpath)
        sessions = []
        for sheet in xl.sheet_names:
            if sheet.strip().lower() in SKIP_SHEETS:
                continue
            try:
                df_raw = pd.read_excel(fpath, sheet_name=sheet, header=None)
            except Exception:
                continue
            for rec in load_module_sheet(df_raw, all_slots, fname_hint=fname):
                sessions.append(rec)
        return sessions

    if request.method == 'GET':
        return render_template('admin/import_template1.html')

    trimester_raw = request.form.get('trimester', '1').strip()
    if trimester_raw not in ('1', '2', '3'):
        flash('Trimester was not a valid selection - please choose from the dropdown.', 'danger')
        return redirect(url_for('admin.import_template1'))
    trimester    = int(trimester_raw)
    confirm      = request.form.get('confirm') == '1'
    token        = request.form.get('token', '').strip()
    orig_filename = request.form.get('orig_filename', '')

    if confirm and token:
        fpath = os.path.join(UPLOAD_DIR, f'{token}.xlsx')
        if not os.path.exists(fpath):
            flash('Upload session expired - please re-upload.', 'danger')
            return redirect(url_for('admin.import_template1'))

        all_slots = TimeSlot.query.all()
        recs = _parse_file(fpath, orig_filename, all_slots)

        seen = set()
        for cs in ClassSession.query.join(Course).all():
            seen.add((cs.course.module_code, cs.course.programme_id,
                      cs.session_type, cs.group_label or 'All', cs.teaching_weeks or ''))

        created = 0
        skipped = 0
        try:
            for rec in recs:
                if not rec['prog_code'] or not rec['year_level']:
                    skipped += 1
                    continue
                prog   = _get_or_create_programme(rec['prog_code'])
                sg     = _get_or_create_student_group(prog, rec['year_level'], rec['class_size'])
                course = _get_or_create_course(
                    rec['module_code'], rec['module_title'],
                    prog, rec['year_level'], rec['delivery_mode'], trimester,
                )
                sig = (rec['module_code'], prog.id, rec['session_type'],
                       rec['group_label'], rec['teaching_weeks'] or '')
                if sig in seen:
                    skipped += 1
                    continue
                seen.add(sig)
                cs = ClassSession(
                    course_id=course.id,
                    session_type=rec['session_type'],
                    delivery_mode=rec['delivery_mode'],
                    is_async=rec['is_async'],
                    duration_hours=rec['duration_hours'],
                    student_group_id=sg.id,
                    trimester=trimester,
                    teaching_weeks=rec['teaching_weeks'],
                    group_label=rec['group_label'],
                    preferred_timeslot_id=rec['pref_slot_id'],
                )
                db.session.add(cs)
                db.session.flush()
                created += 1
                is_first = True
                for staff_name, staff_id in rec['staff']:
                    prof = _get_or_create_professor(staff_name, staff_id)
                    if not prof:
                        continue
                    already = ClassSessionProfessor.query.filter_by(
                        session_id=cs.id, professor_id=prof.id).first()
                    if not already:
                        db.session.add(ClassSessionProfessor(
                            session_id=cs.id, professor_id=prof.id,
                            is_primary=is_first,
                            display_order=0 if is_first else 1,
                        ))
                        if is_first:
                            is_first = False
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Import failed: {e}', 'danger')
            return redirect(url_for('admin.import_template1'))

        try:
            os.unlink(fpath)
        except OSError:
            pass

        flash(f'Import complete: {created} sessions created, {skipped} skipped.', 'success')
        return redirect(url_for('admin.import_template1'))

    # Step 1: upload + parse for preview
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Please select a file.', 'danger')
        return render_template('admin/import_template1.html', trimester=trimester)

    token = str(uuid.uuid4())
    fpath = os.path.join(UPLOAD_DIR, f'{token}.xlsx')
    file.save(fpath)

    all_slots = TimeSlot.query.all()
    try:
        recs = _parse_file(fpath, file.filename, all_slots)
    except Exception as e:
        try:
            os.unlink(fpath)
        except OSError:
            pass
        flash(f'Could not parse file: {e}', 'danger')
        return render_template('admin/import_template1.html', trimester=trimester)

    preview = []
    warn_count = 0
    for rec in recs:
        staff_clean = _clean_staff(rec['staff'])
        row_warnings = []
        if not staff_clean:
            row_warnings.append('No staff assigned')
        if rec['prog_code'] not in PROG_NAMES:
            row_warnings.append(f'Unknown programme code: {rec["prog_code"]}')
        if not rec['teaching_weeks']:
            row_warnings.append('Teaching weeks not specified')
        if row_warnings:
            warn_count += 1
        preview.append({
            'module_code':    rec['module_code'],
            'module_title':   rec['module_title'],
            'session_type':   rec['session_type'],
            'group_label':    rec['group_label'],
            'duration_hours': rec['duration_hours'],
            'teaching_weeks': rec['teaching_weeks'] or '-',
            'is_async':       rec['is_async'],
            'delivery_mode':  rec['delivery_mode'],
            'prog_code':      rec['prog_code'],
            'year_level':     rec['year_level'],
            'staff':          staff_clean,
            'pref_slot_id':   rec['pref_slot_id'],
            'warnings':       row_warnings,
        })

    return render_template('admin/import_template1.html',
                           preview=preview,
                           warn_count=warn_count,
                           token=token,
                           orig_filename=file.filename,
                           trimester=trimester)


# ---------------------------------------------------------------------------
# Timetable - async solver (background task + polling)
# ---------------------------------------------------------------------------

def _purge_old_tasks():
    """Remove task entries older than _TASK_TTL to keep the dict tidy."""
    cutoff = _time.time() - _TASK_TTL
    stale = [tid for tid, t in _solver_tasks.items() if t['started_at'] < cutoff]
    for tid in stale:
        _solver_tasks.pop(tid, None)


def _run_solver_task(app, task_id, action, form_data):
    """Background thread: runs solver and writes result into _solver_tasks."""
    from datetime import date as _date
    from app.engine.solver import solve as _solve
    from app.engine.checker import get_blocking_issues

    def _update(status, message, **kw):
        _solver_tasks[task_id].update({'status': status, 'message': message, **kw})

    with app.app_context():
        try:
            academic_year = form_data.get('academic_year', '').strip().upper()
            break_raw = form_data.get('term_break_weeks', '7').strip()
            term_break_weeks = {int(p) for p in break_raw.split(',') if p.strip().isdigit()} or {7}
            preserve = form_data.get('preserve_existing') == 'on'

            if action == 'generate':
                tri_raw = form_data.get('trimester_num', '').strip()
                trimester_num = int(tri_raw) if tri_raw in ('1', '2', '3') else None
                trimester = f'{academic_year}-T{trimester_num}' if trimester_num else ''
                start_raw = form_data.get('start_date', '').strip()
                if not start_raw and academic_year and trimester_num:
                    start_raw = SIT_ACADEMIC_CALENDAR.get(academic_year, {}).get(trimester_num, '')

                _update('running', f'Solving {trimester}…', trimester=trimester)

                pinned_slots = None
                if preserve:
                    existing = TimetableEntry.query.filter_by(trimester=trimester, is_backbone=False).all()
                    seen = set()
                    pinned_slots = {e.class_session_id: e.timeslot_id
                                    for e in existing if e.class_session_id not in seen
                                    and not seen.add(e.class_session_id)} or None

                # Backbone (real, professor-submitted) entries are fed to the
                # solver as a strongly-weighted soft preference via
                # historical_preferred, NOT a hard pin. A hard pin cannot
                # yield even when a session it locks is genuinely
                # incompatible with another hard rule elsewhere (e.g. the
                # same professor's second, unpinned class needing a slot) -
                # that turned the whole trimester infeasible (found
                # 2026-07-10). The soft version reproduces the real
                # timetable whenever nothing conflicts, since matching it
                # costs nothing and deviating only adds objective cost, but
                # lets the solver move a class when there's a real conflict.
                historical_preferred = _build_historical_preferred(academic_year, trimester_num)

                success, message, stats = _solve(
                    trimester, _date.fromisoformat(start_raw), term_break_weeks,
                    trimester_num=trimester_num, academic_year=academic_year,
                    pinned_slots=pinned_slots, historical_preferred=historical_preferred,
                )
                if success:
                    _auto_create_flags(trimester, stats.get('preferred_violations', []))
                    _save_solve_run(trimester, stats)
                _update('done', message, success=success, stats=stats,
                        trimester=trimester, academic_year=academic_year)

        except Exception as exc:
            _solver_tasks[task_id].update({
                'status': 'error',
                'success': False,
                'message': f'Solver error: {exc}',
            })


@admin_bp.route('/timetable/solve-async', methods=['POST'])
@login_required
def timetable_solve_async():
    """Start a background solver run and return a task_id for polling."""
    from app.engine.checker import get_blocking_issues
    from flask import current_app

    _purge_old_tasks()

    action = request.form.get('action', '')
    if action != 'generate':
        return jsonify({'error': 'Invalid action.'}), 400

    # Quick validation before spawning thread
    academic_year = request.form.get('academic_year', '').strip().upper()
    if not academic_year:
        return jsonify({'error': 'Academic year is required.'}), 400

    tri_raw = request.form.get('trimester_num', '').strip()
    trimester_num = int(tri_raw) if tri_raw in ('1', '2', '3') else None
    if not trimester_num:
        return jsonify({'error': 'Trimester number is required.'}), 400
    tri_blockers, _ = get_blocking_issues(trimester_num=trimester_num)
    if tri_blockers:
        return jsonify({'error': f'{len(tri_blockers)} blocking issue(s): ' + tri_blockers[0]}), 400
    start_raw = request.form.get('start_date', '').strip() or \
                SIT_ACADEMIC_CALENDAR.get(academic_year, {}).get(trimester_num, '')
    if not start_raw:
        return jsonify({'error': 'Start date is required.'}), 400
    trimester = f'{academic_year}-T{trimester_num}'

    # Check no task already running for this trimester
    for t in _solver_tasks.values():
        if t.get('status') == 'running' and t.get('academic_year') == academic_year:
            return jsonify({'error': f'A solver run for {academic_year} is already in progress.'}), 409

    task_id = str(uuid.uuid4())
    _solver_tasks[task_id] = {
        'status': 'running',
        'message': 'Starting…',
        'started_at': _time.time(),
        'success': None,
        'stats': {},
        'trimester': trimester,
        'academic_year': academic_year,
    }

    form_data = request.form.to_dict()  # flat snapshot before the request context closes
    app = current_app._get_current_object()
    t = threading.Thread(target=_run_solver_task, args=(app, task_id, action, form_data), daemon=True)
    t.start()

    return jsonify({'task_id': task_id, 'trimester': trimester})


@admin_bp.route('/solver/status/<task_id>')
@login_required
def solver_status(task_id):
    """Poll endpoint for async solver progress."""
    task = _solver_tasks.get(task_id)
    if not task:
        return jsonify({'status': 'unknown', 'message': 'Task not found.'}), 404
    elapsed = int(_time.time() - task['started_at'])
    return jsonify({
        'status':        task['status'],
        'message':       task['message'],
        'elapsed':       elapsed,
        'success':       task.get('success'),
        'stats':         task.get('stats', {}),
        'trimester':     task.get('trimester', ''),
        'academic_year': task.get('academic_year', ''),
    })


def _compute_soft_violations(entries, trimester):
    """For each entry (already deduplicated to one row per session), work
    out which soft constraints its placement violates - reusing the same
    rule definitions as solver.py/the Scoring Matrix so this view can never
    disagree with them. Returns {entry.id: [reason, ...]}, used to amber-
    highlight rows on the Timetable List view (Brian, 2026-07-12: "different
    color for different constraints... red if hard, yellow for soft").

    Scoped to checks that are meaningful on a single row without having to
    replicate the solver's exact CP-SAT sliding-window math: ends-after-
    cutoff, first/last slot of the day, an under-filled room, an adjacent-
    slot mode switch or room change for the same professor/group, and a
    preferred-availability declaration the solver had to override (read
    from this trimester's persisted solve stats). Professor idle-gap and
    group back-to-back stacking (S3/S8) aren't included yet - flagging
    those correctly needs the same multi-session daily-span logic
    solver.py uses, which doesn't reduce cleanly to a per-row check; a
    future pass could add them. Hard constraints are deliberately not
    checked here - a completed schedule can't violate one (see the Scoring
    Matrix's own explainer), so there's nothing to flag red for a normal
    generated row. A genuinely conflicting manually-edited row would show
    up via the existing Flags/blocking-issues machinery instead.
    """
    from app.engine import solver as sv
    from collections import defaultdict
    from app.models.timeslot import TimeSlot

    violations = {e.id: [] for e in entries}
    if not entries:
        return violations

    def _add(entry_id, reason):
        if reason not in violations[entry_id]:
            violations[entry_id].append(reason)

    # ---- Ends after the preferred cutoff (S9) ----
    cutoff_label = sv.LATE_END_CUTOFF.strftime('%H:%M')
    for e in entries:
        if e.timeslot.end_time > sv.LATE_END_CUTOFF:
            _add(e.id, f'Ends after {cutoff_label}')

    # ---- First/last slot of the day (S7) - global slot catalog, cheap ----
    day_starts = defaultdict(list)
    for ts in TimeSlot.query.all():
        day_starts[ts.day_of_week].append((ts.start_time, ts.id))
    extremal_ts_ids = set()
    for lst in day_starts.values():
        lst.sort()
        extremal_ts_ids.add(lst[0][1])
        extremal_ts_ids.add(lst[-1][1])
    for e in entries:
        if e.timeslot_id in extremal_ts_ids and e.class_session.student_group_id:
            _add(e.id, 'First or last class slot of the day for this group')

    # ---- Room under target utilisation (S6) ----
    threshold_pct = int(sv.ROOM_UTIL_THRESHOLD * 100)
    for e in entries:
        if e.room_id and e.room and e.class_session.student_group:
            group_size = e.class_session.student_group.intake_size
            if e.room.capacity and (group_size / e.room.capacity) < sv.ROOM_UTIL_THRESHOLD:
                pct = round(100 * group_size / e.room.capacity)
                _add(e.id, f'Room only {pct}% full (target {threshold_pct}%+)')

    # ---- Adjacent-slot mode switch / room change, same professor or group (S1/S2/S11) ----
    def _adjacent(a, b):
        return a.timeslot.day_of_week == b.timeslot.day_of_week and (
            a.timeslot.end_time == b.timeslot.start_time or b.timeslot.end_time == a.timeslot.start_time
        )

    by_prof = defaultdict(list)
    by_group = defaultdict(list)
    for e in entries:
        for prof_id in e.class_session.all_professor_ids:
            by_prof[prof_id].append(e)
        if e.class_session.student_group_id:
            by_group[e.class_session.student_group_id].append(e)

    def _check_adjacency(groups, subject):
        for lst in groups.values():
            for i in range(len(lst)):
                for j in range(len(lst)):
                    if i == j or not _adjacent(lst[i], lst[j]):
                        continue
                    ei, ej = lst[i], lst[j]
                    if not sv._weeks_overlap(ei.class_session, ej.class_session):
                        continue  # never actually adjacent in real calendar time
                    if ei.class_session.delivery_mode != ej.class_session.delivery_mode:
                        _add(ei.id, f'Back-to-back online/in-person switch ({subject})')
                    elif ei.room_id and ej.room_id and ei.room_id != ej.room_id:
                        _add(ei.id, f'Different room than the adjacent class ({subject})')

    _check_adjacency(by_prof, 'same professor')
    _check_adjacency(by_group, 'same group')

    # ---- Preferred-availability declaration overridden (S-avail) ----
    stats = _load_solve_run(trimester)
    for pv in (stats or {}).get('preferred_violations', []):
        sid = pv.get('class_session_id')
        ts_id = pv.get('timeslot_id')
        prof_name = pv.get('professor', 'The assigned professor')
        for e in entries:
            if e.class_session_id == sid and e.timeslot_id == ts_id:
                _add(e.id, f'{prof_name} asked to avoid this slot')

    return violations


# ---------------------------------------------------------------------------
# Timetable - generate and view
# ---------------------------------------------------------------------------

@admin_bp.route('/timetable', methods=['GET', 'POST'])
@login_required
def timetable():
    from app.engine.checker import get_blocking_issues
    from app.engine.solver import solve
    from datetime import date
    from collections import defaultdict

    issues, issue_warnings = get_blocking_issues()
    result   = None
    stats    = {}
    trimester = request.args.get('trimester', '')
    source    = request.args.get('source', '')   # 'bb' | 'generated' | ''

    # On fresh GET with no trimester param, default to T1 of the most recent AY
    if not trimester and request.method == 'GET':
        _latest = (db.session.query(TimetableEntry.trimester)
                   .distinct()
                   .order_by(TimetableEntry.trimester.desc())
                   .first())
        if _latest:
            trimester = f'{_latest[0][:6]}-all'

    # Compute default AY for the form (SIT AY starts August)
    today = date.today()
    if today.month >= 8:
        ay_default = f'AY{str(today.year)[2:]}{str(today.year + 1)[2:]}'
    else:
        ay_default = f'AY{str(today.year - 1)[2:]}{str(today.year)[2:]}'

    def _next_ay(ay):
        s, e = int(ay[2:4]), int(ay[4:6])
        return f'AY{s+1:02d}{e+1:02d}'

    selected_tri = None   # remembered for re-populating the form after POST

    if request.method == 'POST':
        action    = request.form.get('action', '')
        trimester = request.form.get('trimester', '').strip()
        start_raw = request.form.get('start_date', '').strip()

        if action == 'generate':
            # New fields: academic_year + trimester_num; build the internal key
            academic_year = request.form.get('academic_year', '').strip().upper()
            tri_raw       = request.form.get('trimester_num', '').strip()
            trimester_num = int(tri_raw) if tri_raw.isdigit() and tri_raw in ('1', '2', '3') else None
            selected_tri  = trimester_num
            if academic_year and trimester_num:
                trimester = f'{academic_year}-T{trimester_num}'

            # Re-check issues scoped to the selected trimester only
            if trimester_num:
                tri_blockers, tri_warns = get_blocking_issues(trimester_num=trimester_num)
            else:
                tri_blockers, tri_warns = issues, issue_warnings

            # Auto-populate start date from official SIT calendar if not provided
            if not start_raw and academic_year and trimester_num:
                start_raw = SIT_ACADEMIC_CALENDAR.get(academic_year, {}).get(trimester_num, '')

            if tri_blockers:
                flash('Resolve all blocking issues for this trimester before generating.', 'danger')
                for iss in tri_blockers[:5]:
                    flash(iss, 'warning')
            elif not academic_year:
                flash('Academic year is required (e.g. AY2526).', 'danger')
            elif not trimester_num:
                flash('Trimester number (1, 2, or 3) is required.', 'danger')
            elif not start_raw:
                flash('Start date is required (AY not in known SIT calendar - enter manually).', 'danger')
            else:
                try:
                    start_date = date.fromisoformat(start_raw)
                    # Parse term break weeks (comma-separated, e.g. "7" or "7,14")
                    break_raw = request.form.get('term_break_weeks', '7').strip()
                    term_break_weeks = set()
                    for part in break_raw.split(','):
                        part = part.strip()
                        if part.isdigit():
                            term_break_weeks.add(int(part))
                    if not term_break_weeks:
                        term_break_weeks = {7}   # safe fallback
                    preserve = request.form.get('preserve_existing') == 'on'
                    pinned_slots = None
                    if preserve:
                        existing = TimetableEntry.query.filter_by(trimester=trimester, is_backbone=False).all()
                        seen_sess = set()
                        pinned_slots = {}
                        for e in existing:
                            if e.class_session_id not in seen_sess:
                                seen_sess.add(e.class_session_id)
                                pinned_slots[e.class_session_id] = e.timeslot_id
                        if not pinned_slots:
                            pinned_slots = None

                    # Backbone entries go in as a soft preference (via
                    # historical_preferred below), not a hard pin - see the
                    # matching comment in _run_solver_task for why a hard pin
                    # made Tri2 infeasible (found 2026-07-10).
                    historical_preferred = _build_historical_preferred(academic_year, trimester_num)

                    success, message, stats = solve(
                        trimester, start_date, term_break_weeks,
                        trimester_num=trimester_num,
                        academic_year=academic_year,
                        pinned_slots=pinned_slots,
                        historical_preferred=historical_preferred,
                    )
                    result = {'success': success, 'message': message}
                    if success:
                        flash(message, 'success')
                        # Auto-create TimetableFlag records for preferred violations
                        _auto_create_flags(trimester, stats.get('preferred_violations', []))
                        _save_solve_run(trimester, stats)
                    else:
                        flash(message, 'danger')
                except Exception as e:
                    flash(f'Solver error: {str(e)}', 'danger')

        elif action == 'publish':
            TimetableEntry.query.filter_by(trimester=trimester, is_published=False).update(
                {'is_published': True}
            )
            db.session.commit()
            flash(f'Timetable for {trimester} published. Professors and students can now view it.', 'success')

        elif action == 'reset':
            from app.models.academic_calendar import AcademicCalendar
            from app.models.timetable_flag import TimetableFlag
            from app.models.flag_response import FlagResponse

            # Only delete solver-generated (non-backbone) entries
            non_backbone_ids = [
                e.id for e in TimetableEntry.query.filter_by(trimester=trimester, is_backbone=False).all()
            ]
            if non_backbone_ids:
                flag_ids = [
                    f.id for f in TimetableFlag.query
                    .filter(TimetableFlag.timetable_entry_id.in_(non_backbone_ids))
                    .all()
                ]
                if flag_ids:
                    FlagResponse.query.filter(FlagResponse.flag_id.in_(flag_ids)).delete(synchronize_session=False)
                    TimetableFlag.query.filter(TimetableFlag.id.in_(flag_ids)).delete(synchronize_session=False)
                TimetableEntry.query.filter(TimetableEntry.id.in_(non_backbone_ids)).delete(synchronize_session=False)

            backbone_count = TimetableEntry.query.filter_by(trimester=trimester, is_backbone=True).count()
            if backbone_count == 0:
                AcademicCalendar.query.filter_by(trimester=trimester).delete()

            db.session.commit()
            note = f' ({backbone_count} backbone entries preserved)' if backbone_count else ''
            flash(f'Solver entries for {trimester} cleared ({len(non_backbone_ids)} deleted{note}).', 'info')
            trimester = ''

        elif action == 'reset_ay':
            # Clear solver-generated entries for all 3 trimesters (backbone preserved)
            from app.models.academic_calendar import AcademicCalendar
            from app.models.timetable_flag import TimetableFlag
            from app.models.flag_response import FlagResponse
            clear_ay = request.form.get('clear_ay', '').strip()
            if clear_ay:
                tri_keys = [f'{clear_ay}-T1', f'{clear_ay}-T2', f'{clear_ay}-T3']
                non_backbone_ids = [
                    e.id for e in TimetableEntry.query.filter(
                        TimetableEntry.trimester.in_(tri_keys),
                        TimetableEntry.is_backbone == False
                    ).all()
                ]
                if non_backbone_ids:
                    flag_ids = [
                        f.id for f in TimetableFlag.query
                        .filter(TimetableFlag.timetable_entry_id.in_(non_backbone_ids))
                        .all()
                    ]
                    if flag_ids:
                        FlagResponse.query.filter(FlagResponse.flag_id.in_(flag_ids)).delete(synchronize_session=False)
                        TimetableFlag.query.filter(TimetableFlag.id.in_(flag_ids)).delete(synchronize_session=False)
                    TimetableEntry.query.filter(TimetableEntry.id.in_(non_backbone_ids)).delete(synchronize_session=False)
                deleted = len(non_backbone_ids)
                # Only clear calendar if no backbone entries remain for that trimester
                for tk in tri_keys:
                    if TimetableEntry.query.filter_by(trimester=tk, is_backbone=True).count() == 0:
                        AcademicCalendar.query.filter_by(trimester=tk).delete()
                db.session.commit()
                flash(f'All timetables for {clear_ay} cleared ({deleted} entries deleted).', 'info')

    # Load timetable entries for display
    entries = []
    trimesters = [r[0] for r in
                  db.session.query(TimetableEntry.trimester).distinct().order_by(TimetableEntry.trimester).all()]

    existing_ays = sorted(set(
        t[:6] for t in trimesters if t.startswith('AY') and len(t) >= 6
    ))
    ay_options = sorted(set(existing_ays + [ay_default, _next_ay(ay_default)]))

    prog_filter = request.args.get('prog', '')

    # Check whether the active AY has backbone entries (for source toggle)
    _active_ay_pfx = trimester[:6] if trimester and trimester.startswith('AY') and len(trimester) >= 6 else ''
    has_backbone_ay = bool(_active_ay_pfx and db.session.query(TimetableEntry.id).filter(
        TimetableEntry.trimester.startswith(_active_ay_pfx),
        TimetableEntry.is_backbone == True
    ).limit(1).scalar())
    # Default to 'generated' when backbone exists and no explicit source chosen
    if has_backbone_ay and not source:
        source = 'generated'

    if trimester:
        if trimester.endswith('-all'):
            ay_prefix = trimester[:-4]  # 'AY2627'
            q = TimetableEntry.query.filter(TimetableEntry.trimester.startswith(ay_prefix))
            if source == 'bb':
                q = q.filter(TimetableEntry.is_backbone == True)
            elif source == 'generated':
                q = q.filter(TimetableEntry.is_backbone == False)
            entries = (q.join(TimetableEntry.timeslot)
                       .order_by(TimetableEntry.trimester, TimeSlot.day_of_week, TimeSlot.period_label)
                       .all())
        else:
            q = TimetableEntry.query.filter_by(trimester=trimester)
            if source == 'bb':
                q = q.filter(TimetableEntry.is_backbone == True)
            elif source == 'generated':
                q = q.filter(TimetableEntry.is_backbone == False)
            entries = (q.join(TimetableEntry.timeslot)
                       .order_by(TimeSlot.day_of_week, TimeSlot.period_label)
                       .all())

    # Programmes present in current view (for active highlighting)
    active_prog_codes = set(
        e.class_session.course.programme.code
        for e in entries
        if e.class_session.course.programme
    )
    # All programmes that have any timetable entry ever (always shown in filter)
    from app.models.programme import Programme
    all_programmes = sorted(
        set(
            (e[0], e[1]) for e in
            db.session.query(Programme.code, Programme.name)
            .join(Programme.courses)
            .join(Course.class_sessions)
            .join(ClassSession.timetable_entries)
            .distinct()
            .all()
        ),
        key=lambda x: x[0]
    )
    programmes = all_programmes  # kept for template compatibility

    # Apply programme filter
    if prog_filter:
        entries = [e for e in entries
                   if e.class_session.course.programme
                   and e.class_session.course.programme.code == prog_filter]

    # Deduplicate: show one row per session (not one per week)
    seen = set()
    unique_entries = []
    for e in entries:
        if e.class_session_id not in seen:
            seen.add(e.class_session_id)
            unique_entries.append(e)

    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    if trimester.endswith('-all'):
        unique_entries.sort(key=lambda e: (
            e.trimester,
            day_order.index(e.timeslot.day_of_week),
            e.timeslot.start_time
        ))
    else:
        unique_entries.sort(key=lambda e: (
            day_order.index(e.timeslot.day_of_week),
            e.timeslot.start_time
        ))

    # ---------------------------------------------------------------------------
    # Weekly view support
    # ---------------------------------------------------------------------------
    from app.models.academic_calendar import AcademicCalendar

    DAYS_ALL = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    view_mode = 'list' if trimester.endswith('-all') else request.args.get('view', 'list')
    week_number     = request.args.get('week', 1, type=int)
    calendar_weeks  = []
    hour_grid       = []
    current_cal_week = None
    prev_week_num   = None
    next_week_num   = None

    if trimester:
        calendar_weeks = (AcademicCalendar.query
                          .filter_by(trimester=trimester)
                          .order_by(AcademicCalendar.week_number)
                          .all())

        all_week_nums = [cw.week_number for cw in calendar_weeks] if calendar_weeks else list(range(1, 14))
        if week_number not in all_week_nums:
            week_number = all_week_nums[0]

        current_idx      = all_week_nums.index(week_number)
        prev_week_num    = all_week_nums[current_idx - 1] if current_idx > 0 else None
        next_week_num    = all_week_nums[current_idx + 1] if current_idx < len(all_week_nums) - 1 else None
        current_cal_week = next((cw for cw in calendar_weeks if cw.week_number == week_number), None)

        wq = (TimetableEntry.query
              .join(TimetableEntry.class_session)
              .join(TimetableEntry.timeslot)
              .filter(
                  TimetableEntry.trimester == trimester,
                  TimetableEntry.week_number == week_number,
              ))
        if source == 'bb':
            wq = wq.filter(TimetableEntry.is_backbone == True)
        elif source == 'generated':
            wq = wq.filter(TimetableEntry.is_backbone == False)
        week_entries_raw = wq.all()

        if prog_filter:
            week_entries_raw = [e for e in week_entries_raw
                                if e.class_session.course.programme
                                and e.class_session.course.programme.code == prog_filter]

        # Continuous hourly time-axis grid, not one row per named period -
        # lab periods (e.g. "Lab PM2" 13:00-15:00) run longer than lecture
        # periods (e.g. "P2" 12:00-14:00) but start at the same clock time,
        # so a period-label-per-row layout made adjacent rows look like
        # separate non-overlapping blocks when they actually overlapped in
        # real time (found 2026-07-12, Brian: "why is lab PM2 a time?").
        # Rows are now real clock hours; a session spans `rowspan` rows
        # (its real duration), and each day column tracks how many hours
        # are still covered by an earlier rowspan so those rows don't get
        # a second cell.
        all_slot_hours = TimeSlot.query.all()
        min_hour = min((ts.start_time.hour for ts in all_slot_hours), default=9)
        max_hour = max((ts.end_time.hour for ts in all_slot_hours), default=18)

        entries_by_day_hour = {day: defaultdict(list) for day in DAYS_ALL}
        for entry in week_entries_raw:
            d = entry.timeslot.day_of_week
            if d in entries_by_day_hour:
                entries_by_day_hour[d][entry.timeslot.start_time.hour].append(entry)

        occupied_until = {day: min_hour for day in DAYS_ALL}
        for h in range(min_hour, max_hour):
            row = {'hour': h, 'label': f'{h:02d}:00', 'end_label': f'{h + 1:02d}:00', 'cells': {}}
            for day in DAYS_ALL:
                if occupied_until[day] > h:
                    row['cells'][day] = None   # covered by an earlier rowspan - render nothing here
                    continue
                day_entries = entries_by_day_hour[day].get(h, [])
                if day_entries:
                    span = max(1, max(
                        e.timeslot.end_time.hour - e.timeslot.start_time.hour for e in day_entries
                    ))
                    row['cells'][day] = {'rowspan': span, 'entries': day_entries}
                    occupied_until[day] = h + span
                else:
                    row['cells'][day] = {'rowspan': 1, 'entries': []}
                    occupied_until[day] = h + 1
            hour_grid.append(row)

    # Year levels available in this trimester (for filter buttons)
    year_levels = sorted(set(
        e.class_session.course.year_level
        for e in unique_entries
        if e.class_session.course.year_level
    )) if unique_entries else []

    # The primary generate flow is async (posts to solve-async, then this
    # page loads as a plain GET) so `stats` is usually still {} here even
    # right after a successful generation - fall back to the last solve
    # persisted for this trimester so the summary still shows up.
    display_stats = stats or _load_solve_run(trimester)
    constraint_summary = _build_constraint_summary(display_stats)

    # Per-row soft-constraint highlighting - List view only for now (Weekly's
    # grid cells are too small for a useful multi-line tooltip yet).
    soft_violations = (
        _compute_soft_violations(unique_entries, trimester)
        if view_mode == 'list' and unique_entries else {}
    )

    # The Generate form always targets whatever single trimester is currently
    # being viewed (the trimester selector above the timetable), rather than
    # its own separate AY/Tri picker - having two independent trimester
    # pickers on one page was the single biggest source of "which trimester
    # am I even looking at" confusion (Brian, 2026-07-15). "All Tri" has no
    # single trimester to generate for, so gen_trimester stays None there.
    gen_ay = gen_tri_num = gen_trimester = None
    gen_start_date = ''
    if trimester and trimester[-4:] != '-all' and '-T' in trimester:
        _gen_ay_part, _gen_tri_part = trimester.split('-T')
        if _gen_tri_part in ('1', '2', '3'):
            gen_ay = _gen_ay_part
            gen_tri_num = int(_gen_tri_part)
            gen_trimester = trimester
            gen_start_date = SIT_ACADEMIC_CALENDAR.get(gen_ay, {}).get(gen_tri_num, '')

    return render_template('admin/timetable.html',
                           gen_ay=gen_ay,
                           gen_tri_num=gen_tri_num,
                           gen_trimester=gen_trimester,
                           gen_start_date=gen_start_date,
                           issues=issues,
                           issue_warnings=issue_warnings,
                           trimesters=trimesters,
                           active_trimester=trimester,
                           entries=unique_entries,
                           soft_violations=soft_violations,
                           stats=display_stats,
                           constraint_summary=constraint_summary,
                           result=result,
                           view_mode=view_mode,
                           week_number=week_number,
                           calendar_weeks=calendar_weeks,
                           current_cal_week=current_cal_week,
                           prev_week_num=prev_week_num,
                           next_week_num=next_week_num,
                           hour_grid=hour_grid,
                           days_all=DAYS_ALL,
                           year_levels=year_levels,
                           programmes=programmes,
                           active_prog_codes=active_prog_codes,
                           prog_filter=prog_filter,
                           ay_default=ay_default,
                           ay_options=ay_options,
                           selected_tri=selected_tri,
                           source=source,
                           has_backbone_ay=has_backbone_ay)


# ---------------------------------------------------------------------------
# Scheduling Administration Report - a fuller, standalone view of how one
# trimester's timetable was generated: run overview, an optimised score, and
# a breakdown of the actual scheduled sessions by source/day/type/mode.
# Own design (not a copy of any reference) - built to fit this app's own
# data model rather than force-fitting someone else's labels.
# ---------------------------------------------------------------------------

@admin_bp.route('/timetable/report')
@login_required
def timetable_report():
    from app.models.solve_run import SolveRun
    from app.models.programme import Programme

    trimester = request.args.get('trimester', '')
    if not trimester:
        _latest = (db.session.query(TimetableEntry.trimester)
                   .distinct().order_by(TimetableEntry.trimester.desc()).first())
        trimester = _latest[0] if _latest else ''

    trimesters = sorted(set(
        t[0] for t in db.session.query(TimetableEntry.trimester).distinct().all()
    ), reverse=True)

    if not trimester:
        return render_template('admin/timetable_report.html', trimester='',
                                trimesters=trimesters, report=None)

    all_entries = (
        TimetableEntry.query
        .filter_by(trimester=trimester)
        .join(TimetableEntry.class_session)
        .join(ClassSession.course)
        .options(
            db.joinedload(TimetableEntry.timeslot),
            db.joinedload(TimetableEntry.room),
            db.joinedload(TimetableEntry.class_session).joinedload(ClassSession.course).joinedload(Course.programme),
            db.joinedload(TimetableEntry.class_session).joinedload(ClassSession.student_group),
        )
        .all()
    )

    # One row per session (not one per week) for every count/breakdown below -
    # a weekly class recurring 13 times must count once, not 13 times. Prefer
    # the solver-generated entry over a backbone one when a session has both
    # (mirrors the Template 2 export fix, 2026-07-16) - backbone is a soft
    # preference the solver already used to bias its own result, not a
    # second, competing "real" schedule; picking whichever happened to come
    # back first from the DB could silently show a stale backbone day/time
    # for the handful of sessions where the solver moved off it.
    by_session = {}
    for e in all_entries:
        existing = by_session.get(e.class_session_id)
        if existing is None or (existing.is_backbone and not e.is_backbone):
            by_session[e.class_session_id] = e
    sessions = list(by_session.values())

    def _count_by(key_fn, label_fn=None):
        counts = {}
        for e in sessions:
            k = key_fn(e)
            if k is None:
                continue
            counts[k] = counts.get(k, 0) + 1
        label_fn = label_fn or (lambda k: k)
        rows = [{'label': label_fn(k), 'count': v} for k, v in counts.items()]
        rows.sort(key=lambda r: -r['count'])
        max_count = max((r['count'] for r in rows), default=0)
        for r in rows:
            r['pct'] = round(100 * r['count'] / max_count) if max_count else 0
        return rows

    DAY_ORDER = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    day_rows = _count_by(lambda e: e.timeslot.day_of_week if e.timeslot else None)
    day_rows.sort(key=lambda r: DAY_ORDER.index(r['label']) if r['label'] in DAY_ORDER else 99)
    max_day = max((r['count'] for r in day_rows), default=0)
    for r in day_rows:
        r['pct'] = round(100 * r['count'] / max_day) if max_day else 0

    class_type_rows = _count_by(
        lambda e: e.class_session.session_type,
        lambda k: T2_CLASS_TYPE.get(k, k.capitalize()),
    )
    delivery_rows = _count_by(
        lambda e: e.class_session.delivery_mode,
        lambda k: {'f2f': 'Face-to-face', 'online': 'Online', 'hybrid': 'Hybrid'}.get(k, k.capitalize()),
    )

    rooms_used = len({e.room_id for e in sessions if e.room_id})
    staff_assigned = (
        db.session.query(ClassSessionProfessor.professor_id)
        .join(ClassSession, ClassSessionProfessor.session_id == ClassSession.id)
        .filter(ClassSession.id.in_([e.class_session_id for e in sessions]))
        .distinct().count()
    ) if sessions else 0
    teaching_weeks = len({e.week_number for e in all_entries if e.week_number})
    programmes_covered = len({
        e.class_session.course.programme_id for e in sessions
        if e.class_session.course.programme_id
    })
    modules_covered = len({e.class_session.course_id for e in sessions})
    student_groups_covered = len({
        e.class_session.student_group_id for e in sessions
        if e.class_session.student_group_id
    })

    solve_row = SolveRun.query.filter_by(trimester=trimester).first()
    stats = {}
    if solve_row:
        import json as _json
        try:
            stats = _json.loads(solve_row.stats_json)
        except (TypeError, ValueError):
            stats = {}
    constraint_summary = _build_constraint_summary(stats) if stats else None

    hist_honoured = stats.get('historical_honoured')
    hist_changed = stats.get('historical_changed')
    historical_match = None
    if hist_honoured is not None and (hist_honoured + (hist_changed or 0)) > 0:
        historical_match = f'{hist_honoured}/{hist_honoured + hist_changed}'

    # Top drivers - the 1-2 soft rules actually responsible for most of the
    # point deficit, named in plain language up front rather than making an
    # admin dig through the full 15-rule Scoring Matrix to find out why the
    # score is what it is (Brian: "we definitely need to improve the
    # scoring... 47 is very low" - the honest answer is 2 rules explain most
    # of it, not a scattered mess).
    top_issues = []
    if constraint_summary:
        all_rows = [r for g in constraint_summary['soft_groups'] for r in g['rows'] if r['points'] > 0]
        all_rows.sort(key=lambda r: r['points'], reverse=True)
        for r in all_rows[:2]:
            top_issues.append({'title': r['title'], 'violated': r['violated'], 'points': r['points']})

    soft_total = constraint_summary['soft_total_violations'] if constraint_summary else 0
    soft_total_points = constraint_summary['soft_total_points'] if constraint_summary else 0
    preference_problems = len(stats.get('preferred_violations', []))
    # Self-defined scoring, disclosed on the page - not an official metric.
    # Each violation is weighted by its own solver priority (see
    # SCORE_CEILING_AVG_PENALTY), then averaged per session so a large
    # trimester with more sessions to check isn't unfairly penalised versus
    # a small one - a flat "-N points per violation" used to floor every
    # trimester this size straight to 0, unable to tell a decent schedule
    # from a bad one (found 2026-07-12). Hard constraints can never be
    # violated (the solve would have failed), so that term is always -0.
    # Full per-rule breakdown (weight + points impact) is in
    # constraint_summary, rendered as the Scoring Matrix below.
    sessions_count = len(sessions) or 1
    avg_weighted_penalty = soft_total_points / sessions_count
    score = round(max(0, 100 * (1 - avg_weighted_penalty / SCORE_CEILING_AVG_PENALTY)))

    report = {
        'trimester': trimester,
        'solver_status': solve_row.solver_status if solve_row else None,
        'generated_at': (solve_row.updated_at or solve_row.created_at) if solve_row else None,
        'has_data': bool(sessions),
        'top_issues': top_issues,
        'tiles': {
            'sessions_scheduled': len(sessions),
            'rooms_used': rooms_used,
            'staff_assigned': staff_assigned,
            'teaching_weeks': teaching_weeks,
            'soft_violations': soft_total,
        },
        'overview': {
            'programmes': programmes_covered,
            'modules': modules_covered,
            'student_groups': student_groups_covered,
            'hard_rule_count': constraint_summary['hard_rule_count'] if constraint_summary else None,
            'soft_rule_count': constraint_summary['soft_rule_count'] if constraint_summary else None,
            # "Backbone match" (kept-vs-changed slot vs last year/DSC's real
            # timetable) isn't a universal quality metric - only DSC has a
            # real submitted schedule to match against, everyone else is
            # compared to the solver's own prior output. Shown here as
            # context, not a top-level score tile, and only when there's
            # actually something to report.
            'historical_match': historical_match,
        },
        'score': {
            'available': solve_row is not None,
            'value': score,
            'hard_conflicts': 0,
            'soft_violations': soft_total,
            'soft_total_points': soft_total_points,
            'avg_weighted_penalty': round(avg_weighted_penalty, 2),
            'ceiling': SCORE_CEILING_AVG_PENALTY,
            'preference_problems': preference_problems,
        },
        'breakdown': {
            'day': day_rows,
            'class_type': class_type_rows,
            'delivery_mode': delivery_rows,
        },
        'constraint_summary': constraint_summary,
    }

    return render_template('admin/timetable_report.html',
                           trimester=trimester, trimesters=trimesters, report=report)


# ---------------------------------------------------------------------------
# Timetable Similarity - mirrored timetable report across trimesters
# ---------------------------------------------------------------------------

@admin_bp.route('/timetable/similarity')
@login_required
def timetable_similarity():
    """Cross-AY comparison: did the historical soft constraint keep slots consistent?"""
    # Available AYs from timetable entries
    all_ays = sorted(set(
        e[0] for e in db.session.query(TimetableEntry.academic_year).distinct().all()
        if e[0]
    ))

    def _slot_label(ts):
        return f'{ts.day_of_week[:3]} {ts.start_time.strftime("%H:%M")}–{ts.end_time.strftime("%H:%M")}'

    # Build combined option list: value='AY2526:backbone', label='AY2526BB'
    # Backbone variant uses BB suffix; generated variant uses plain AY name.
    similarity_options = []  # list of (value, label)
    for ay in all_ays:
        has_bb  = TimetableEntry.query.filter_by(academic_year=ay, is_backbone=True).first()  is not None
        has_gen = TimetableEntry.query.filter_by(academic_year=ay, is_backbone=False).first() is not None
        if has_bb:
            similarity_options.append((f'{ay}:backbone',  f'{ay}BB'))
        if has_gen:
            similarity_options.append((f'{ay}:generated', ay))
        if not has_bb and not has_gen:
            similarity_options.append((f'{ay}:all', ay))

    def _parse_opt(val):
        """Split 'AY2526:backbone' → (ay, source). Default source = 'all'."""
        if val and ':' in val:
            ay, src = val.split(':', 1)
            return ay, src
        return val or '', 'all'

    # Smart defaults: base → first backbone option, compare → first generated option
    def _default_opt(prefer_backbone):
        for v, _ in similarity_options:
            ay, src = _parse_opt(v)
            if prefer_backbone and src == 'backbone':
                return v
        for v, _ in similarity_options:
            ay, src = _parse_opt(v)
            if not prefer_backbone and src == 'generated':
                return v
        return similarity_options[0][0] if similarity_options else ''

    default_base    = _default_opt(True)
    default_compare = _default_opt(False)
    # If both defaults resolve to the same option, push compare to the next distinct one
    if default_base == default_compare and len(similarity_options) >= 2:
        default_compare = similarity_options[1][0]

    base_val    = request.args.get('base',    default_base)
    compare_val = request.args.get('compare', default_compare)
    base_ay,    base_source    = _parse_opt(base_val)
    compare_ay, compare_source = _parse_opt(compare_val)

    # Human-readable labels for headings/alerts
    _opt_labels = dict(similarity_options)
    base_label    = _opt_labels.get(base_val,    base_ay)
    compare_label = _opt_labels.get(compare_val, compare_ay)

    cross_rows = []

    if base_ay and compare_ay:
        def _build_map(trimester_key, source='all'):
            q = TimetableEntry.query.filter_by(trimester=trimester_key)
            if source == 'backbone':
                q = q.filter_by(is_backbone=True)
            elif source == 'generated':
                q = q.filter_by(is_backbone=False)
            entries = q.all()
            slot_map = {}
            for e in entries:
                key = (e.class_session.course.module_code, e.class_session.session_type)
                if key not in slot_map:
                    ts = e.timeslot
                    slot_map[key] = {
                        'label':            _slot_label(ts),
                        'day':              ts.day_of_week,
                        'start':            ts.start_time.strftime('%H:%M'),
                        'room':             e.room.room_code if e.room else None,
                        'timeslot_id':      ts.id,
                        'room_id':          e.room_id,
                        'class_session_id': e.class_session_id,
                    }
            return slot_map

        def _reason_tag(class_session_id, base_timeslot_id, base_room_id, compare_tri_key):
            """Infer why the solver moved a session off its historical slot.
            Returns (tag, explanation) - explanation is a specific, data-driven sentence for the admin."""
            cs = ClassSession.query.get(class_session_id)
            base_ts = TimeSlot.query.get(base_timeslot_id)
            slot_str = (f"{base_ts.day_of_week} {base_ts.start_time.strftime('%H:%M')}–{base_ts.end_time.strftime('%H:%M')}"
                        if base_ts else "the backbone slot")
            if cs:
                for prof in cs.all_professors:
                    hard = AvailabilityDeclaration.query.filter_by(
                        professor_id=prof.id,
                        timeslot_id=base_timeslot_id,
                        constraint_type='strict',
                    ).first()
                    if hard:
                        return ('prof_hard_conflict',
                                f"Prof. {prof.user.name} has a strict unavailability at {slot_str} - the solver cannot override this hard constraint.")
                    soft = AvailabilityDeclaration.query.filter_by(
                        professor_id=prof.id,
                        timeslot_id=base_timeslot_id,
                    ).first()
                    if soft:
                        return ('prof_unavailable',
                                f"Prof. {prof.user.name} is unavailable at {slot_str}, so the solver found a different slot.")
                if cs.student_group_id:
                    group_clash = (TimetableEntry.query
                                   .join(TimetableEntry.class_session)
                                   .filter(
                                       TimetableEntry.trimester == compare_tri_key,
                                       TimetableEntry.timeslot_id == base_timeslot_id,
                                       ClassSession.student_group_id == cs.student_group_id,
                                       TimetableEntry.class_session_id != class_session_id,
                                   ).first())
                    if group_clash:
                        clash_cs = group_clash.class_session
                        sg = cs.student_group.group_label if cs.student_group else 'the same group'
                        return ('group_overlap',
                                f"Group {sg} already has {clash_cs.course.module_code} {clash_cs.session_type} at {slot_str}.")
            if base_room_id:
                room_clash = TimetableEntry.query.filter_by(
                    trimester=compare_tri_key,
                    room_id=base_room_id,
                    timeslot_id=base_timeslot_id,
                ).first()
                if room_clash:
                    rc_cs = room_clash.class_session
                    base_room = Room.query.get(base_room_id)
                    room_code = base_room.room_code if base_room else 'the backbone room'
                    return ('room_conflict',
                            f"{room_code} is occupied by {rc_cs.course.module_code} {rc_cs.session_type} at {slot_str}.")
            return ('rescheduled',
                    'No specific constraint identified - the solver chose a different slot through optimisation.')

        # Module codes that existed in the base AY (filtered by source) - used to
        # distinguish truly-new modules from ones that simply had no backbone entry
        base_ay_mods_q = TimetableEntry.query.filter(TimetableEntry.academic_year == base_ay)
        if base_source == 'backbone':
            base_ay_mods_q = base_ay_mods_q.filter_by(is_backbone=True)
        elif base_source == 'generated':
            base_ay_mods_q = base_ay_mods_q.filter_by(is_backbone=False)
        base_ay_mods = set(
            e.class_session.course.module_code.upper() for e in base_ay_mods_q.all()
        )
        # Also include all courses that exist in the system (curriculum-level check)
        all_curriculum_mods = set(c.module_code.upper() for c in Course.query.all())

        for tri_num in [1, 2, 3]:
            base_map    = _build_map(f'{base_ay}-T{tri_num}',    base_source)
            compare_map = _build_map(f'{compare_ay}-T{tri_num}', compare_source)
            if not base_map and not compare_map:
                continue

            all_keys = sorted(set(base_map.keys()) | set(compare_map.keys()))
            for mod, stype in all_keys:
                bd = base_map.get((mod, stype))
                cd = compare_map.get((mod, stype))
                base_slot    = bd['label'] if bd else None
                compare_slot = cd['label'] if cd else None

                if base_slot and compare_slot:
                    consistency = 'same' if base_slot == compare_slot else 'different'
                elif base_slot:
                    consistency = 'base_only'
                else:
                    consistency = 'compare_only'

                reason     = ''
                reason_tag = ''
                # A module is only "truly new" if it doesn't exist in the curriculum at all.
                # If it exists in the DB but is missing from the backbone, that's a backbone
                # import gap - not a new module.
                is_truly_new = (mod.upper() not in all_curriculum_mods
                                and mod.upper() not in base_ay_mods)
                is_backbone_gap = (mod.upper() in all_curriculum_mods
                                   and mod.upper() not in base_ay_mods)

                if consistency == 'different':
                    parts = []
                    if bd['day'] != cd['day']:
                        parts.append(f"Day: {bd['day'][:3]} → {cd['day'][:3]}")
                    if bd['start'] != cd['start']:
                        parts.append(f"Time: {bd['start']} → {cd['start']}")
                    if bd['room'] != cd['room']:
                        parts.append(f"Room: {bd['room'] or '?'} → {cd['room'] or '?'}")
                    reason = ' | '.join(parts) if parts else 'Slot changed'
                    reason_tag, explanation = _reason_tag(
                        bd['class_session_id'],
                        bd['timeslot_id'],
                        bd['room_id'],
                        f'{compare_ay}-T{tri_num}',
                    )
                elif consistency == 'base_only':
                    reason = f'Not scheduled in {compare_ay}'
                    explanation = f'Present in {base_label} but not scheduled in {compare_label}. The module may have been dropped or excluded from the new run.'
                elif consistency == 'compare_only':
                    if is_truly_new:
                        reason = f'New module added in {compare_ay}'
                        explanation = f'Not in {base_label} at all - genuinely new to the curriculum, scheduled without a historical reference.'
                    else:
                        reason = f'Missing from {base_label} (backbone import gap)'
                        explanation = f'Module exists in the curriculum but was not captured in the {base_label} import. It was scheduled freely in {compare_label} with no historical slot to follow.'
                else:  # same
                    explanation = ''

                cross_rows.append({
                    'tri_num'         : tri_num,
                    'module_code'     : mod,
                    'session_type'    : stype,
                    'base_slot'       : base_slot,
                    'compare_slot'    : compare_slot,
                    'consistency'     : consistency,
                    'reason'          : reason,
                    'reason_tag'      : reason_tag,
                    'is_truly_new'    : is_truly_new,
                    'is_backbone_gap' : is_backbone_gap,
                    'explanation'     : explanation,
                })

    same_count = sum(1 for r in cross_rows if r['consistency'] == 'same')
    diff_count = sum(1 for r in cross_rows if r['consistency'] == 'different')

    # Per-trimester breakdown
    tri_stats = {}
    for tri_num in [1, 2, 3]:
        s = sum(1 for r in cross_rows if r['tri_num'] == tri_num and r['consistency'] == 'same')
        d = sum(1 for r in cross_rows if r['tri_num'] == tri_num and r['consistency'] == 'different')
        total = s + d
        tri_stats[tri_num] = {
            'same':  s,
            'diff':  d,
            'total': total,
            'pct':   round(s / total * 100) if total > 0 else None,
        }

    # Programme codes for the filter chips (derived from module code prefix)
    prog_codes = sorted(set(r['module_code'][:3] for r in cross_rows))

    return render_template('admin/timetable_similarity.html',
                           similarity_options=similarity_options,
                           base_val=base_val,
                           compare_val=compare_val,
                           base_ay=base_ay,
                           compare_ay=compare_ay,
                           base_label=base_label,
                           compare_label=compare_label,
                           cross_rows=cross_rows,
                           same_count=same_count,
                           diff_count=diff_count,
                           tri_stats=tri_stats,
                           prog_codes=prog_codes)


# ---------------------------------------------------------------------------
# Timetable export - flat XLSX download
# ---------------------------------------------------------------------------

@admin_bp.route('/timetable/export')
@login_required
def timetable_export():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    trimester_filter = request.args.get('trimester', '')

    q = TimetableEntry.query
    if trimester_filter:
        q = q.filter_by(trimester=trimester_filter)
    entries = q.order_by(
        TimetableEntry.trimester,
        TimetableEntry.week_number,
    ).all()

    # Session-type colour palette (hex fill colours)
    TYPE_COLOURS = {
        'lecture':  'D6E4F0',
        'lectorial':'D6E4F0',
        'tutorial': 'D5F5E3',
        'lab':      'FEF9E7',
        'seminar':  'F9EBEA',
    }
    DEFAULT_COLOUR = 'F2F3F4'

    HDR_FILL = PatternFill('solid', fgColor='2E4057')
    HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
    BOLD     = Font(bold=True, size=9)
    SMALL    = Font(size=9)
    CENTRE   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT     = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
    THIN     = Side(style='thin', color='BBBBBB')
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    def _make_sheet(wb, title, sheet_entries):
        """Build one grid sheet: rows=weeks, columns=Day×Period."""
        from app.models.academic_calendar import AcademicCalendar

        # Gather unique time slots from the entries (ordered)
        from app.models.timeslot import TimeSlot as TS
        slot_ids = sorted(set(e.timeslot_id for e in sheet_entries),
                          key=lambda sid: (
                              DAYS.index(next(e.timeslot.day_of_week for e in sheet_entries if e.timeslot_id == sid)),
                              next(e.timeslot.start_time for e in sheet_entries if e.timeslot_id == sid)
                          ))
        # Build ordered list of (day, start, end, period_label) tuples
        col_slots = []
        seen_cols = set()
        for e in sorted(sheet_entries, key=lambda x: (
                DAYS.index(x.timeslot.day_of_week) if x.timeslot.day_of_week in DAYS else 9,
                x.timeslot.start_time)):
            ts = e.timeslot
            key = (ts.day_of_week, ts.start_time, ts.end_time)
            if key not in seen_cols:
                seen_cols.add(key)
                col_slots.append((ts.day_of_week, ts.start_time, ts.end_time, ts.period_label))

        if not col_slots:
            return

        # Gather weeks
        weeks = sorted(set(e.week_number for e in sheet_entries))
        trimester = sheet_entries[0].trimester if sheet_entries else ''
        cal_map = {
            cw.week_number: cw
            for cw in AcademicCalendar.query.filter_by(trimester=trimester).all()
        } if trimester else {}

        ws = wb.create_sheet(title=title)

        # Header row 1: Day spans
        ws.cell(1, 1, 'Week').font = HDR_FONT
        ws.cell(1, 1).fill = HDR_FILL
        ws.cell(1, 1).alignment = CENTRE
        ws.cell(2, 1, 'Date').font = HDR_FONT
        ws.cell(2, 1).fill = HDR_FILL
        ws.cell(2, 1).alignment = CENTRE

        col = 2
        day_start_col = {}
        for day, start, end, period in col_slots:
            if day not in day_start_col:
                day_start_col[day] = col
            label = f'{start.strftime("%H:%M")}–{end.strftime("%H:%M")}'
            c = ws.cell(1, col, day if col == day_start_col[day] else '')
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = CENTRE
            c2 = ws.cell(2, col, label)
            c2.font = HDR_FONT
            c2.fill = HDR_FILL
            c2.alignment = CENTRE
            ws.column_dimensions[get_column_letter(col)].width = 22
            col += 1

        ws.column_dimensions['A'].width = 12

        # Merge day header cells
        col = 2
        for day in DAYS:
            day_cols = [i + 2 for i, (d, _, _, _) in enumerate(col_slots) if d == day]
            if len(day_cols) > 1:
                ws.merge_cells(start_row=1, start_column=day_cols[0],
                               end_row=1, end_column=day_cols[-1])
            col += len(day_cols)

        # Build lookup: (week, day, start_time) -> entry list
        from collections import defaultdict
        cell_map = defaultdict(list)
        for e in sheet_entries:
            cell_map[(e.week_number, e.timeslot.day_of_week, e.timeslot.start_time)].append(e)

        # Data rows
        data_row = 3
        for wk in weeks:
            cw = cal_map.get(wk)
            week_label = f'Week {wk}'
            if cw and cw.start_date:
                date_label = cw.start_date.strftime('%d %b')
            else:
                date_label = ''

            wc = ws.cell(data_row, 1, week_label)
            wc.font = BOLD
            wc.alignment = CENTRE
            wc.border = BORDER
            if cw and cw.is_term_break:
                wc.fill = PatternFill('solid', fgColor='FFF3CD')

            dc = ws.cell(data_row + 1, 1, date_label)
            dc.font = SMALL
            dc.alignment = CENTRE
            dc.border = BORDER

            col = 2
            for day, start, end, period in col_slots:
                cell_entries = cell_map.get((wk, day, start), [])
                if cell_entries:
                    e = cell_entries[0]
                    cs  = e.class_session
                    crs = cs.course
                    prof = (e.override_professor.user.name if e.override_professor
                            else (cs.primary_professor.user.name if cs.primary_professor else ''))
                    room = e.room.room_code if e.room else 'Online'
                    text = f'{crs.module_code}\n{cs.session_type.capitalize()}\n{room}'
                    if prof:
                        text += f'\n{prof}'
                    colour = TYPE_COLOURS.get(cs.session_type, DEFAULT_COLOUR)
                    fill = PatternFill('solid', fgColor=colour)
                    c = ws.cell(data_row, col, text)
                    c.fill = fill
                    c.font = SMALL
                    c.alignment = LEFT
                    c.border = BORDER
                    ws.row_dimensions[data_row].height = 52
                else:
                    c = ws.cell(data_row, col, '')
                    c.border = BORDER
                col += 1

            # Second sub-row for date label already written; border remaining cells
            for col2 in range(2, col):
                ws.cell(data_row + 1, col2).border = BORDER

            data_row += 2

        ws.freeze_panes = 'B3'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Group entries by trimester then year level
    from collections import defaultdict
    tri_year_map = defaultdict(lambda: defaultdict(list))
    for e in entries:
        yr = e.class_session.course.year_level or 0
        tri_year_map[e.trimester][yr].append(e)

    for tri_key in sorted(tri_year_map.keys()):
        yr_map = tri_year_map[tri_key]
        for yr in sorted(yr_map.keys()):
            sheet_entries = yr_map[yr]
            yr_label = f'Y{yr}' if yr else 'All'
            short_tri = tri_key.replace('AY', '').replace('-T', ' T')
            sheet_title = f'{short_tri} {yr_label}'[:31]  # Excel limit
            _make_sheet(wb, sheet_title, sheet_entries)

    if not wb.sheetnames:
        ws = wb.create_sheet('No data')
        ws['A1'] = 'No timetable entries found.'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'timetable_{trimester_filter or "all"}.xlsx'
    return send_file(buf, download_name=filename,
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------------------------------------------------------------------------
# Template 2 export - SIT upload format (flat row-per-session-pattern)
# ---------------------------------------------------------------------------

@admin_bp.route('/timetable/export-template2')
@login_required
def timetable_export_template2():
    import io, os, re
    import openpyxl
    from openpyxl.styles import Font
    from flask import send_file
    from collections import defaultdict
    from datetime import time as dtime

    trimester_filter = request.args.get('trimester', '')

    HEADERS = [
        'Module', 'Class Type', 'Template', 'Group', 'Day', 'Start', 'End',
        'Class Size', 'Sector', 'RoomGrouping', 'Room1', 'Room2', 'StaffGrouping',
        'Staff1', 'Staff2', 'Tri Week', 'Recording Mode', 'Remark',
        'FMTS Tri Start Week', 'Activity Hostkey', 'SIS Module Code', 'Term',
        'Activity Type', 'Duration', 'Staff Suitability ID', 'SIS Staff ID',
        'SIS Staff ID 2', 'Zone Hoskey', 'Location Suitability ID',
        'Location Hostkey', 'Location Hostkey 2',
    ]

    def _term_code(tri_str):
        m = re.match(r'AY(\d{2})\d{2}-T(\d)', tri_str or '')
        return f'{m.group(1)}{m.group(2)}0' if m else '2510'

    # Resolve trimester int from string 'AY2526-T1' → 1
    tri_int = None
    if trimester_filter:
        m = re.match(r'AY\d{4}-T(\d)', trimester_filter)
        if m:
            tri_int = int(m.group(1))

    # Validation gate: genuine blocking issues (e.g. quiz-overlap) always stop
    # the export outright - an export built on top of them would just carry
    # the same problem into Ms. Yang's system. Warnings (missing data that
    # doesn't break scheduling logic) don't block, but require an explicit
    # confirm click rather than exporting silently.
    from app.engine.checker import get_blocking_issues as _get_blocking_issues
    _blockers, _warnings = _get_blocking_issues(trimester_num=tri_int)
    if _blockers:
        flash(
            f'Export blocked - {len(_blockers)} issue(s) must be fixed first: '
            + '; '.join(_blockers[:3]) + (f' (+{len(_blockers) - 3} more)' if len(_blockers) > 3 else ''),
            'danger'
        )
        return redirect(url_for('admin.timetable', trimester=trimester_filter))

    if _warnings and request.args.get('confirmed') != '1':
        return render_template('admin/export_confirm.html',
                               trimester=trimester_filter, warnings=_warnings)

    # Load sessions with all relationships eager-loaded
    q = (
        ClassSession.query
        .join(ClassSession.course)
        .options(
            db.joinedload(ClassSession.course).joinedload(Course.programme),
            db.joinedload(ClassSession.professor_assignments)
                .joinedload(ClassSessionProfessor.professor)
                .joinedload(Professor.user),
            db.joinedload(ClassSession.student_group),
            db.joinedload(ClassSession.timetable_entries)
                .joinedload(TimetableEntry.timeslot),
            db.joinedload(ClassSession.timetable_entries)
                .joinedload(TimetableEntry.room),
            db.joinedload(ClassSession.fixed_room),
        )
    )
    if tri_int is not None:
        q = q.filter(ClassSession.trimester == tri_int)
    # Sessions deliberately deferred from this generation pass (see
    # bootstrap/48-49, System Info's "Deferred from T1 generation" note)
    # have no TimetableEntry at all - exclude them from the export outright
    # rather than showing a confusing all-blank row that looks like a gap.
    q = q.filter(ClassSession.deferred_from_solve.is_(False))
    all_sessions = q.order_by(Course.module_code, ClassSession.session_type).all()

    # Build slot map: cs.id → (timeslot | None, sorted_weeks_list, room | None)
    slot_map = {}
    for cs in all_sessions:
        te_list_all = [e for e in cs.timetable_entries
                       if not trimester_filter or e.trimester == trimester_filter]
        # Prefer solver-generated entries over backbone ones - mirrors the
        # Timetable page's own default Source filter ("default to generated
        # when backbone exists"). Backbone data is a soft preference the
        # solver already used to bias its own result; once generation has
        # run, the generated entry is the current schedule. Mixing both
        # sources for the same session (found 2026-07-16) produced random
        # blank Room1 cells and, for sessions where the solver picked a
        # different day than backbone, a Tri Week list merged from both.
        generated = [e for e in te_list_all if not e.is_backbone]
        te_list = generated if generated else te_list_all
        if te_list:
            ts = te_list[0].timeslot
            weeks = sorted(set(e.week_number for e in te_list if e.week_number != 7))
            room = te_list[0].room
            slot_map[cs.id] = (ts, weeks, room)
        elif cs.fixed_timeslot_id and cs.fixed_timeslot:
            weeks = ([int(w) for w in cs.teaching_weeks.split(',') if w.strip()]
                     if cs.teaching_weeks else [])
            slot_map[cs.id] = (cs.fixed_timeslot, weeks, cs.fixed_room)
        else:
            slot_map[cs.id] = (None, [], None)

    # Sort for stable Template numbering
    DAYS_ORD = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4}

    def _sort_key(cs):
        ts, _, _ = slot_map[cs.id]
        d = DAYS_ORD.get(ts.day_of_week, 9) if ts else 9
        s = ts.start_time if ts else dtime(0, 0)
        return (cs.course.module_code, cs.session_type,
                cs.teaching_weeks or '', d, s, cs.group_label or 'All')

    all_sessions.sort(key=_sort_key)

    # Assign Template numbers and Activity Hostkey suffixes
    tmpl_state = defaultdict(lambda: {'patterns': {}, 'counter': 0})
    act_state   = {}   # (mod_key, group, weeks_str) → activity sequential number
    act_counter = defaultdict(int)

    term_code = _term_code(trimester_filter)
    rows = []

    for cs in all_sessions:
        ts, weeks, room = slot_map[cs.id]
        prog = cs.course.programme
        sector, campus_abbr = T2_PROG_SECTOR.get(prog.code, T2_PROG_SECTOR_DEFAULT)
        cluster_abbr = T2_CLUSTER_ABBR.get(prog.cluster, prog.cluster[:3].upper())
        mod_code = cs.course.module_code
        group = cs.group_label or 'All'
        weeks_str = cs.teaching_weeks or (','.join(str(w) for w in weeks) if weeks else '')

        mod_key = (mod_code, cs.session_type)

        # Template number: each unique (weeks_str, timeslot) within mod_key = new template.
        # Unscheduled sessions use cs.id as tiebreaker so parallel unscheduled sessions
        # (e.g. two lectures per week not yet solved) each get their own template number.
        if ts:
            slot_pat = (weeks_str, ts.day_of_week, ts.start_time)
        else:
            slot_pat = (weeks_str, None, cs.id)
        tmpl_s = tmpl_state[mod_key]
        if slot_pat not in tmpl_s['patterns']:
            tmpl_s['counter'] += 1
            tmpl_s['patterns'][slot_pat] = tmpl_s['counter']
        tmpl_num = tmpl_s['patterns'][slot_pat]

        # Activity number (for hostkey suffix): unique per (mod_key, group, weeks_str)
        act_tuple = (mod_key, group, weeks_str)
        if act_tuple not in act_state:
            act_counter[mod_key] += 1
            act_state[act_tuple] = act_counter[mod_key]
        act_num = act_state[act_tuple]
        act_code = T2_ACT_CODE.get(cs.session_type, 'OTH')
        act_sfx  = '' if act_num == 1 else str(act_num)

        hostkey     = f'{mod_code}-{term_code}-{cluster_abbr}-UGRD-{campus_abbr}-{act_code}{act_sfx}/{group}'
        sis_mod     = f'{mod_code}-{term_code}-{cluster_abbr}-UGRD-{campus_abbr}'

        # Staff - handle names stored as "Prof A\nProf B" (single DB record, two people)
        profs = cs.all_professors
        staff1_name = profs[0].user.name if profs else ''
        staff1_id   = profs[0].staff_id  if profs else ''
        staff2_name = profs[1].user.name if len(profs) > 1 else ''
        staff2_id   = profs[1].staff_id  if len(profs) > 1 else ''
        if '\n' in staff1_name and not staff2_name:
            parts = staff1_name.split('\n', 1)
            staff1_name, staff2_name = parts[0].strip(), parts[1].strip()

        # Timeslot
        if ts:
            day_str   = T2_DAY_ABBR.get(ts.day_of_week, ts.day_of_week[:3])
            start_str = ts.start_time.strftime('%H%M')
            end_str   = ts.end_time.strftime('%H%M')
        else:
            day_str = start_str = end_str = ''

        # Remark - cross-programme shared-module note, matching Ms. Yang's
        # reference file convention ("w <other linked module code>") for a
        # class that's really one shared session under multiple programmes'
        # own module numbering. Only set when we have a real SharedModuleGroup
        # link - never guessed for ordinary, non-shared sessions.
        remark = ''
        if cs.shared_module_group_id:
            other_mods = sorted({
                s.course.module_code for s in cs.shared_module_group.class_sessions
                if s.course.module_code != mod_code
            })
            if other_mods:
                remark = 'w ' + ', '.join(other_mods)

        rows.append({
            'Module':                 mod_code,
            'Class Type':             T2_CLASS_TYPE.get(cs.session_type, cs.session_type.capitalize()),
            'Template':               tmpl_num,
            'Group':                  group,
            'Day':                    day_str,
            'Start':                  start_str,
            'End':                    end_str,
            'Class Size':             cs.effective_group_size if cs.student_group else '',
            'Sector':                 sector,
            'RoomGrouping':           '',
            'Room1':                  room.room_code if room else '',
            'Room2':                  '',
            'StaffGrouping':          '',
            'Staff1':                 staff1_name,
            'Staff2':                 staff2_name,
            'Tri Week':               weeks_str,
            'Recording Mode':         'A0' if cs.session_type == 'lectorial' else '',
            'Remark':                 remark,
            'FMTS Tri Start Week':    1,
            'Activity Hostkey':       hostkey,
            'SIS Module Code':        sis_mod,
            'Term':                   int(term_code),
            'Activity Type':          act_code,
            'Duration':               cs.duration_hours * 3,
            'Staff Suitability ID':   '',
            'SIS Staff ID':           staff1_id,
            'SIS Staff ID 2':         staff2_id,
            'Zone Hoskey':            sector,
            'Location Suitability ID': '',
            'Location Hostkey':       '',
            'Location Hostkey 2':     '',
        })

    # Build Excel workbook using the reference template as base so that all
    # lookup sheets, auto-filter, data-validation dropdowns, and header
    # formatting are preserved exactly as Ms. Yang's template.
    import os
    base_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                             'static', 'template2_base.xlsx')
    wb = openpyxl.load_workbook(base_path)
    ws = wb['Timetable']

    # The base template's 'Staff' lookup sheet (used for the Staff1/Staff2
    # dropdown validation) ships with Ms. Yang's own full SIT staff directory
    # (~2,300 real names + real staff IDs) - copied through unmodified by
    # every export until now, even after staff_id was mocked on the
    # Timetable data sheet itself (2026-07-17: "replace all of the staff id
    # with mock id" - real IDs flagged as likely NDA/confidential). Replace
    # its contents with only the staff who actually appear in this system,
    # using the same mock IDs already on their Professor record, instead of
    # carrying through ~2,000 unrelated real staff records with real IDs.
    if 'Staff' in wb.sheetnames:
        staff_ws = wb['Staff']
        if staff_ws.max_row > 1:
            staff_ws.delete_rows(2, staff_ws.max_row - 1)
        for row_idx, prof in enumerate(
            Professor.query.join(Professor.user).order_by(User.name).all(), start=2
        ):
            staff_ws.cell(row_idx, 1, prof.user.name)
            staff_ws.cell(row_idx, 2, prof.staff_id)

    # 'Sheet2' and 'Sheet3' are ALSO real SIT-wide staff directories (found
    # 2026-07-18, one export cycle after the 'Staff' sheet fix above - same
    # kind of leak, real names + real staff IDs, just hiding under generic
    # unlabelled sheet names I hadn't checked yet: 2,066 and 1,413 real
    # records respectively, covering staff across the whole university, not
    # just this system's professors). Same treatment as 'Staff': clear and
    # rebuild from only the professors that actually appear here, using
    # their existing mock staff_id. Only Name/Host Key are populated - the
    # richer columns these sheets have (Email, Department, Shared With,
    # Primary/Other Suitabilities, Contract/Maximum Periods) are Ms. Yang's
    # own facilities-scheduling fields with no equivalent data on our side,
    # left blank rather than guessed.
    for sheet_name in ('Sheet2', 'Sheet3'):
        if sheet_name not in wb.sheetnames:
            continue
        extra_ws = wb[sheet_name]
        if extra_ws.max_row > 1:
            extra_ws.delete_rows(2, extra_ws.max_row - 1)
        for row_idx, prof in enumerate(
            Professor.query.join(Professor.user).order_by(User.name).all(), start=2
        ):
            extra_ws.cell(row_idx, 1, prof.user.name)
            extra_ws.cell(row_idx, 2, prof.staff_id)

    # The base template's 'Location' lookup sheet only lists Ms. Yang's own
    # Dover-campus venues (e.g. 'DV-AP-LT1A') - none of this system's real
    # Punggol-campus room codes (e.g. 'E2-01-01') are in it at all, so a
    # downstream system validating Room1 against this sheet would reject
    # every single row (found 2026-07-18). Ms. Yang's own Dover rows are left
    # untouched (real venue data, not confidential, and may still be needed
    # elsewhere) - this only appends our own rooms so they're recognised too.
    # Only Name/Host Key/Capacity are populated from data this system
    # actually has; Department/Zone/Suitabilities/etc are Ms. Yang's own
    # facilities-management fields with no equivalent here, left blank rather
    # than guessed (see System Info for the disclosure). 18 of these rooms
    # use a bracketed placeholder code (e.g. '[CVE-Lab-1]') rather than a
    # real physical room number - included since they're genuinely what
    # appears in Room1, but they are NOT real bookable venues.
    if 'Location' in wb.sheetnames:
        loc_ws = wb['Location']
        next_row = loc_ws.max_row + 1
        for room in Room.query.filter_by(is_active=True).order_by(Room.room_code).all():
            loc_ws.cell(next_row, 1, room.room_code)
            loc_ws.cell(next_row, 2, room.room_code)
            loc_ws.cell(next_row, 3, str(room.capacity))
            next_row += 1

    # Clear existing data rows (keep header row 1 with its formatting).
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    DATA_FONT = Font(size=11)

    if rows:
        for ri, row in enumerate(rows, 2):
            for ci, h in enumerate(HEADERS, 1):
                c = ws.cell(ri, ci, row.get(h, ''))
                c.font = DATA_FONT
    else:
        ws.cell(2, 1, 'No sessions found for this trimester.').font = DATA_FONT

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'template2_{trimester_filter or "all"}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------------------------------------------------------------------------
# Timetable summary - plain-English overview via LLM
# ---------------------------------------------------------------------------

@admin_bp.route('/timetable/summary')
@login_required
def timetable_summary():
    from flask import jsonify
    import anthropic as _anthropic

    trimester = request.args.get('trimester', '')
    if not trimester:
        return jsonify({'error': 'No trimester specified'}), 400

    entries = TimetableEntry.query.filter_by(trimester=trimester).all()
    if not entries:
        return jsonify({'error': 'No timetable entries found for this trimester'}), 404

    # Build stats for the prompt
    total_sessions   = len(set(e.class_session_id for e in entries))
    total_weeks      = len(set(e.week_number for e in entries))
    rooms_used       = sorted(set(e.room.room_code for e in entries if e.room))
    module_codes     = sorted(set(e.class_session.course.module_code for e in entries))
    session_types    = {}
    for e in entries:
        st = e.class_session.session_type
        session_types[st] = session_types.get(st, 0) + 1
    type_summary = ', '.join(f'{v} {k}s' for k, v in sorted(session_types.items()))
    profs_set = set()
    for e in entries:
        for p in e.class_session.all_professors:
            profs_set.add(p.name)

    stats_text = (
        f"Trimester: {trimester}\n"
        f"Modules scheduled: {len(module_codes)} ({', '.join(module_codes)})\n"
        f"Unique class sessions: {total_sessions}\n"
        f"Session type breakdown: {type_summary}\n"
        f"Weeks covered: {total_weeks}\n"
        f"Rooms utilised: {len(rooms_used)} ({', '.join(rooms_used[:10])}{'...' if len(rooms_used) > 10 else ''})\n"
        f"Professors teaching: {len(profs_set)}\n"
    )

    api_key = current_app.config.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'ANTHROPIC_API_KEY not configured in config.py'}), 500

    client = _anthropic.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model='claude-haiku-4-5-20251001',
        max_tokens=300,
        messages=[{
            'role': 'user',
            'content': (
                'You are a timetabling administrator writing a brief internal summary of a generated academic timetable. '
                'Write 3–4 concise sentences in plain English. Do not use bullet points. '
                'Do not mention AI, machine learning, algorithms, or any automated system. '
                'Just describe what was scheduled factually, as a human administrator would write it.\n\n'
                f'Timetable statistics:\n{stats_text}'
            ),
        }],
    )
    summary = msg.content[0].text.strip()
    return jsonify({'summary': summary})


# ---------------------------------------------------------------------------
# Events - planned events as hard constraints
# ---------------------------------------------------------------------------

@admin_bp.route('/events')
@login_required
def events():
    from app.models.event import Event
    from app.models.timetable_entry import TimetableEntry
    from datetime import date

    all_events = Event.query.order_by(Event.event_date).all()

    # For each event, count how many published timetable entries are affected
    impact = {}
    for ev in all_events:
        affected = (TimetableEntry.query
                    .join(TimetableEntry.class_session)
                    .join(TimetableEntry.timeslot)
                    .filter(TimetableEntry.is_published == True)
                    .all())
        count = 0
        for entry in affected:
            session_date = entry.class_session  # placeholder - detailed check in template
            count_check = _event_affects_entry(ev, entry)
            if count_check:
                count += 1
        impact[ev.id] = count

    return render_template('admin/events.html', events=all_events, impact=impact)


def _event_affects_entry(event, entry):
    """Return True if an event blocks a specific timetable entry."""
    from datetime import timedelta
    _day_offset = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4}
    cal = entry.academic_calendar_week if hasattr(entry, 'academic_calendar_week') else None

    # We need the entry's actual date - requires the calendar week
    from app.models.academic_calendar import AcademicCalendar
    cal_week = AcademicCalendar.query.filter_by(
        trimester=entry.trimester,
        week_number=entry.week_number
    ).first()
    if not cal_week:
        return False

    day_offset = _day_offset.get(entry.timeslot.day_of_week, 0)
    entry_date = cal_week.start_date + timedelta(days=day_offset)

    if entry_date != event.event_date:
        return False

    if event.is_full_day:
        return True

    # Check specific timeslots
    blocked = event.blocked_timeslot_ids
    return entry.timeslot_id in blocked


@admin_bp.route('/events/add', methods=['GET', 'POST'])
@login_required
def event_add():
    from app.models.event import Event
    programmes = Programme.query.order_by(Programme.code).all()
    timeslots  = TimeSlot.query.order_by(TimeSlot.day_of_week, TimeSlot.start_time).all()

    if request.method == 'POST':
        name         = request.form.get('name', '').strip()
        description  = request.form.get('description', '').strip()
        event_date   = request.form.get('event_date', '').strip()
        is_full_day  = request.form.get('is_full_day') == 'on'
        timeslot_ids = ','.join(request.form.getlist('timeslot_ids'))
        scope        = request.form.get('scope', 'school_wide')
        programme_id = request.form.get('programme_id', '').strip() or None
        outcome      = request.form.get('outcome', 'cancel')
        trimester    = request.form.get('trimester', '').strip() or None
        academic_year= request.form.get('academic_year', '').strip() or None
        is_recurring = request.form.get('is_recurring') == 'on'

        errors = []
        if not name:       errors.append('Event name is required.')

        from datetime import date as date_cls
        parsed_date = None
        if not event_date:
            errors.append('Event date is required.')
        else:
            try:
                parsed_date = date_cls.fromisoformat(event_date)
            except ValueError:
                errors.append('Event date was not a valid date - please use the date picker.')

        if scope not in ('school_wide', 'programme', 'course'):
            errors.append('Scope was not a valid selection - please choose from the dropdown.')
        if outcome not in ('cancel', 'reschedule'):
            errors.append('Outcome was not a valid selection - please choose from the dropdown.')

        programme = None
        if scope == 'programme' and not programme_id:
            errors.append('Please select a programme, since this event\'s scope is programme-specific.')
        if programme_id:
            try:
                programme = Programme.query.get(int(programme_id))
            except ValueError:
                errors.append('Programme was not a valid selection - please choose from the dropdown.')
            else:
                if programme is None:
                    errors.append('That programme no longer exists - please pick another.')

        trimester_num = None
        if trimester:
            try:
                trimester_num = int(trimester)
            except ValueError:
                errors.append('Trimester was not a valid selection - please choose from the dropdown.')
            else:
                if trimester_num not in (1, 2, 3):
                    errors.append('Trimester must be 1, 2, or 3.')

        if not is_full_day and not timeslot_ids:
            errors.append('Please select at least one timeslot, or mark this event as full-day.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/event_add.html',
                                   programmes=programmes, timeslots=timeslots, form=request.form)

        ev = Event(
            name         = name,
            description  = description or None,
            event_date   = parsed_date,
            is_full_day  = is_full_day,
            timeslot_ids = timeslot_ids if not is_full_day else None,
            scope        = scope,
            programme_id = programme.id if programme else None,
            outcome      = outcome,
            trimester    = trimester_num,
            academic_year= academic_year or None,
            is_recurring = is_recurring,
        )
        db.session.add(ev)
        db.session.commit()
        flash(f'Event "{name}" added successfully.', 'success')
        return redirect(url_for('admin.events'))

    return render_template('admin/event_add.html',
                           programmes=programmes, timeslots=timeslots, form={})


@admin_bp.route('/events/<int:event_id>/delete', methods=['POST'])
@login_required
def event_delete(event_id):
    from app.models.event import Event
    ev = Event.query.get_or_404(event_id)
    name = ev.name
    db.session.delete(ev)
    db.session.commit()
    flash(f'Event "{name}" deleted.', 'success')
    return redirect(url_for('admin.events'))
